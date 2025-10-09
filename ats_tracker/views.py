import os
import io
import json
import csv
import mysql.connector
import textract
from .parser import ResumeParser

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.hashers import check_password, make_password
from django.contrib.auth import logout
from django.core.paginator import Paginator
from django.db import connection
from django.http import (
    JsonResponse, HttpResponseBadRequest, FileResponse, Http404, HttpResponse
)
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.utils.html import escape
from django.views.decorators.csrf import csrf_exempt
from django.template.loader import render_to_string

from datetime import datetime, timedelta, date
from dateutil.relativedelta import relativedelta

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook


from .authentication import anonymous_required, login_required, role_required
from .db_initializer import ATSDatabaseInitializer
from .utils import Constants, DataOperations, MessageProviders, DataValidators



@anonymous_required
def login_view(request):
    """
    Login view for user authentication.

    """
    initializer = ATSDatabaseInitializer()
    initializer.initialize()
    initializer.close()

    username = request.POST.get('username', '').strip()
    password = request.POST.get('password', '').strip()
    name = username  # Default name

    if request.method == 'POST':
        if not username or not password:
            return render(request, 'login.html', {'error': 'Please enter both username/email and password.'})

        valid_user = validate_user(username, password)
        print("login_view -> Valid User:", valid_user)
        if valid_user and valid_user[0] is not None:
            user_id, db_username, role, status = valid_user

            # Check for existing active session
            conn = DataOperations.get_db_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute(
                "SELECT session_id, expires_at FROM user_sessions WHERE user_id=%s AND expires_at > %s",
                (user_id, datetime.now())
            )
            existing_session = cursor.fetchone()
            if existing_session:
                cursor.close()
                conn.close()
                return render(request, 'login.html', {'error': 'User already logged in elsewhere.'})

            # Create new session
            import uuid
            session_id = str(uuid.uuid4())
            session_expiry = settings.SESSION_COOKIE_AGE  # seconds
            expires_at = datetime.now() + timedelta(seconds=session_expiry)
            cursor.execute(
                "INSERT INTO user_sessions (session_id, user_id, expires_at) VALUES (%s, %s, %s)",
                (session_id, user_id, expires_at)
            )

            # Fetch name from hr_team_members table against email
            cursor.execute("SELECT first_name, last_name FROM hr_team_members WHERE email=%s", [username])
            row = cursor.fetchone()
            if row:
                first_name, last_name = row['first_name'], row['last_name']
                name = f"{first_name} {last_name}"

            conn.commit()
            cursor.close()
            conn.close()

            request.session['user_id'] = user_id
            request.session['username'] = db_username
            request.session['role'] = role
            request.session['authenticated'] = True
            request.session['email'] = username
            request.session['name'] = name
            request.session['session_id'] = session_id  # Store session_id

            return redirect('home')
        else:
            return render(request, 'login.html', {'error': 'Invalid credentials or inactive user.'})
    print("login_view -> User Name:", name)
    return render(request, 'login.html', {'error': 'Click Submit to Proceed'})

# function to validate username and password with users table
def validate_user(username, password):
    """
    Validate user credentials against the database.

    """
    print("validate_user -> Username:", username)
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT user_id, username, email, password_hash, role, is_active
        FROM users
        WHERE username=%s OR email=%s
        LIMIT 1
    """, [username, username])
    row = cursor.fetchone()
    print("validate_user -> Database row:", row)
    #compare the hashed password with the stored hash
    if row:
        print("validate_user -> User found in database")
        user_id, db_username, db_email, db_hash, role, is_active = row
        if check_password(password, db_hash) and is_active:
            status = True
            return user_id, db_username, role, status
    cursor.close()
    conn.close()
    return None, None, None, None

def home(request):
    """Home view for authenticated users."""
    session = request.session
    print(session)
    name = request.session.get('name', 'Guest')
    return render(request, 'home.html', {'name': name})




@role_required('Admin')
def add_member(request):
    """
    View to add a new HR team member.

    """
    print("add_member -> Request method:", request.method)
    from .utils import Constants
    message = ''
    error = ''
    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip()
        phone = request.POST.get('phone', '').strip()
        role = request.POST.get('role', '').strip()
        date_joined = request.POST.get('date_joined', '').strip()
        status = request.POST.get('status', 'active').strip()

        # Basic in-view validation
        if not first_name or not last_name or not email or not role or not date_joined:
            error = "All fields except phone are required."
        elif not DataValidators.validate_names(first_name) or not DataValidators.validate_names(last_name):
            error = "Names can only contain letters, spaces, and hyphens."
        elif '@' not in email or len(email) > 100 or not DataValidators.is_valid_email(email):
            error = "Invalid email address."
        elif len(first_name) > 50 or len(last_name) > 50:
            error = "First and last name should be under 50 characters."
        elif status not in ('active', 'inactive', 'on_leave'):
            error = "Invalid status."
        elif phone and (len(phone) > 20) or not DataValidators.is_valid_mobile_number(phone):
            error = "Please Enter a valid phone number."
        # if date is too old, it should not accept
        elif date_joined > str(date.today()):
            error = "Date joined cannot be in the future."
        elif date_joined < Constants.MIN_JOIN_DATE:
            error = f"Date joined cannot be before {Constants.MIN_JOIN_DATE}."
        else:
            try:
                # Add HR team member
                connection = DataOperations.get_db_connection()
                cursor = connection.cursor()
                cursor.execute("""
                    INSERT INTO hr_team_members
                    (first_name, last_name, email, phone, role, date_joined, status)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, [first_name, last_name, email, phone, role, date_joined, status])

                # Create user login for HR member
                from django.contrib.auth.hashers import make_password
                from .utils import Constants
                default_password = Constants.DEFAULT_PASSWORD
                password_hash = make_password(default_password)
                cursor.execute("""
                    INSERT INTO users (username, email, password_hash, role, is_active)
                    VALUES (%s, %s, %s, %s, %s)
                """, [email, email, password_hash, "User", True])
                
                message = f"Member {escape(first_name)} {escape(last_name)} added successfully and login created!"
                connection.commit()
                # creating login for the new member
            except Exception as e:
                if connection:
                    connection.rollback()
                if 'Duplicate entry' in str(e):
                    error = "A member with this email already exists."
                else:
                    error = f"Failed to add member: {str(e)}"
            finally:
                DataOperations.close_db_connection(connection, cursor)
    name = request.session.get('name', 'Guest')
    return render(request, 'add_member.html', {
        'message': message,
        'name': name,
        'error': error
    })

@role_required(['Admin', 'Team_Lead'])
def manage_members(request):
    """
    View to manage HR team members with searching and sorting.
    """
    members_list = []
    error = ''
    
    # Get search and sort parameters
    search_query = request.GET.get('q', '')
    sort_by = request.GET.get('sort', 'full_name')
    sort_dir = request.GET.get('dir', 'asc')

    # Validate sort_by to prevent SQL injection
    valid_sort_columns = ['full_name', 'email', 'role']
    if sort_by not in valid_sort_columns:
        sort_by = 'full_name'
    
    # Validate sort_dir
    if sort_dir not in ['asc', 'desc']:
        sort_dir = 'asc'

    try:
        role = request.session.get('role', 'Guest')
        user_id = request.session.get('user_id', None)
        conn = DataOperations.get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        sql_query = """
            SELECT DISTINCT
                htm.emp_id,
                CONCAT(htm.first_name, ' ', htm.last_name) AS full_name,
                htm.email,
                htm.role,
                htm.status
            FROM
                hr_team_members htm
            INNER JOIN users u ON htm.email = u.email AND u.is_active = TRUE
        """
        params = []

        if role == 'Team_Lead':
            emp_id = DataOperations.get_emp_id_from_user_id(user_id)
            sql_query += """
                INNER JOIN
                    team_members tm ON htm.emp_id = tm.emp_id
                INNER JOIN
                    teams t ON tm.team_id = t.team_id
                WHERE
                    t.lead_emp_id = %s
            """
            params.append(emp_id)
        
        if search_query:
            search_term = f"%{search_query}%"
            if role == 'Team_Lead':
                sql_query += " AND (CONCAT(htm.first_name, ' ', htm.last_name) LIKE %s OR htm.email LIKE %s)"
            else:
                sql_query += " WHERE (CONCAT(htm.first_name, ' ', htm.last_name) LIKE %s OR htm.email LIKE %s)"
            params.extend([search_term, search_term])

        sql_query += f" ORDER BY {sort_by} {sort_dir.upper()};"
        
        cursor.execute(sql_query, params)
        members_list = cursor.fetchall()

    except Exception as e:
        error = f"Failed to fetch members: {str(e)}"
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'conn' in locals():
            conn.close()

    paginator = Paginator(members_list, 10)  # Show 10 members per page.
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    name = request.session.get('name', 'Guest')
    if error:
        messages.error(request, error)
    
    return render(request, 'manage_members.html', {
        'members': page_obj,
        'name': name,
        'error': error,
        'search_query': search_query,
        'sort_by': sort_by,
        'sort_dir': sort_dir,
        'next_sort_dir': 'desc' if sort_dir == 'asc' else 'asc'
    })

@role_required(['Admin', 'Team_Lead'])
def change_member_status(request):
    """
    API endpoint to change the status of a team member.

    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method.'}, status=400)

    member_id = request.POST.get('member_id')
    new_status = request.POST.get('new_status')

    if not member_id or not new_status:
        return JsonResponse({'error': 'Member ID and new status are required.'}, status=400)
    
    if new_status not in ('active', 'on_leave'):
        return JsonResponse({'error': 'Invalid status value.'}, status=400)
    elif new_status == 'inactive':
        return JsonResponse({'error': 'Cannot set status to inactive via this endpoint.'}, status=400)

    # if user is team lead, then check emp_id is belong to the team where user is lead of
    role = request.session.get('role', 'Guest')
    if role == 'Team_Lead':
        user_id = request.session.get('user_id', None)
        emp_id = DataOperations.get_emp_id_from_user_id(user_id)
        conn = DataOperations.get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT COUNT(*) as count
            FROM hr_team_members htm
            INNER JOIN team_members tm ON htm.emp_id = tm.emp_id
            INNER JOIN teams t ON tm.team_id = t.team_id
            WHERE t.lead_emp_id = %s AND htm.emp_id = %s
        """, [emp_id, member_id])
        row = cursor.fetchone()
        if row['count'] == 0:
            cursor.close()
            conn.close()
            return JsonResponse({'error': 'You do not have permission to change this member\'s status.'}, status=403)
        cursor.close()
        conn.close()
    try:
        conn = DataOperations.get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("UPDATE hr_team_members SET status=%s WHERE emp_id=%s", [new_status, member_id])
        if cursor.rowcount == 0:
            return JsonResponse({'error': 'Member not found.'}, status=404)
        conn.commit()
        cursor.execute("SELECT first_name, last_name FROM hr_team_members WHERE emp_id=%s", [member_id])
        row = cursor.fetchone()
        if row:
            messages.success(request, f"Member status updated successfully: {row['first_name']} {row['last_name']}")
        return JsonResponse({'message': 'Member status updated successfully.'})
    except Exception as e:
        if conn:
            conn.rollback()
        messages.error(request, f"Failed to update status: {str(e)}")
        return JsonResponse({'error': f'Failed to update status: {str(e)}'}, status=500)
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'conn' in locals():
            conn.close()

@role_required('Admin')
def create_team(request):
    """
    View to create a new team.

    """
    message = error = None
    members = []
    teams = []
    print("create_team -> Request method:", request.POST)
    if request.method == "POST":
        team_name = request.POST.get("team_name", "").strip()
        selected_members = request.POST.getlist("members")
        team_lead = request.POST.get("team_lead", "").strip()
        # Deduplicate member IDs
        selected_members = list(set(selected_members))
        if team_lead not in selected_members:
            selected_members.append(team_lead)
        
        if not team_name or not selected_members:
            error = "Team name and at least one member are required."
        elif not team_lead:
            error = "Team lead must be selected."
        else:
            try:
                conn = DataOperations.get_db_connection()
                cursor = conn.cursor(dictionary=True)
                # Check for duplicate team name
                cursor.execute("SELECT COUNT(*) as count FROM teams WHERE team_name=%s", [team_name])
                t = cursor.fetchone()
                print("create_team -> Duplicate check result:", t, type(t), t['count'] if t else 'N/A', type(t['count']) if t else 'N/A')
                if t['count'] > 0:
                    error = "A team with this name already exists."
                else:
                    cursor.execute("INSERT INTO teams (team_name, lead_emp_id) VALUES (%s, %s)", [team_name, team_lead])
                    team_id = cursor.lastrowid

                    for emp_id in selected_members:
                        cursor.execute("INSERT INTO team_members (team_id, emp_id) VALUES (%s, %s)", [team_id, emp_id])

                    for emp_id in selected_members:
                        user_id = DataOperations.get_user_id_from_emp_id(emp_id)
                        if user_id and DataOperations.get_user_settings(user_id).get('notifications_enabled', False):
                            MessageProviders.send_notification(user_id, "Team Update", f"You have been added to the team '{team_name}'", created_by="system", notification_type="Team")
                    
                    message = f"Team '{team_name}' created successfully."
                    conn.commit()
            except Exception as e:
                if conn:
                    conn.rollback()
                error = f"Failed to create team: {str(e)}"
            finally:
                if 'cursor' in locals():
                    cursor.close()
                if 'conn' in locals():
                    conn.close()
    # Fetch available members
    try:
        conn = DataOperations.get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT emp_id, first_name, last_name, email, phone FROM hr_team_members WHERE status='active'")
        members = cursor.fetchall()
        # Fetch teams and their strength
        cursor.execute("""
            SELECT t.team_id, t.team_name, t.created_at, COUNT(tm.emp_id) as strength
            FROM teams t
            LEFT JOIN team_members tm ON t.team_id = tm.team_id
            GROUP BY t.team_id, t.team_name, t.created_at
            ORDER BY t.created_at DESC
        """)
        teams = cursor.fetchall()
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'conn' in locals():
            conn.close()
    name = request.session.get('name', 'Guest')
    return render(request, "create_team.html", {
        "members": members,
        "teams": teams,
        "message": message,
        "name": name,
        "error": error
    })

def team_members(request, team_id):
    """
    View to get members of a specific team.
    """
    try:
        conn = DataOperations.get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT m.emp_id, m.first_name, m.last_name, m.email, m.phone, m.role, m.date_joined, m.status
            FROM hr_team_members m
            JOIN team_members tm ON m.emp_id = tm.emp_id
            WHERE tm.team_id = %s
        """, [team_id])
        members = cursor.fetchall()
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'conn' in locals():
            conn.close()
    return JsonResponse({"members": members})

def view_edit_teams(request):
    """
    View to edit existing teams.
    """
    teams = []
    
    try:
        user_id = request.session.get('user_id', None)
        name = request.session.get('name', 'Guest')
        role = request.session.get('role', 'Guest')
        conn = DataOperations.get_db_connection()
        cursor = conn.cursor(dictionary=True)
        if user_id and role!='Guest':
            cursor.execute("""
                    SELECT emp_id FROM hr_team_members
                            WHERE email=(SELECT email from users WHERE user_id=%s)
                            LIMIT 1;
                """, [user_id])
            row = cursor.fetchone()
            emp_id=row['emp_id']

            if role == 'Admin':
                cursor.execute("""
                    SELECT t.team_id, t.team_name, t.created_at, COUNT(tm.emp_id) as strength
                    FROM teams t
                    LEFT JOIN team_members tm ON t.team_id = tm.team_id
                    GROUP BY t.team_id, t.team_name, t.created_at
                    ORDER BY t.created_at DESC
                """)

            elif role == 'Team_Lead':
                cursor.execute("""
                    SELECT t.team_id, t.team_name, t.created_at, COUNT(tm.emp_id) as strength
                    FROM teams t
                    LEFT JOIN team_members tm ON t.team_id = tm.team_id
                    WHERE t.lead_emp_id=%s OR tm.emp_id=%s
                    GROUP BY t.team_id, t.team_name, t.created_at
                    ORDER BY t.created_at DESC
                """, [emp_id, emp_id])

            elif role == 'User':
                cursor.execute("""
                    SELECT t.team_id, t.team_name, t.created_at, COUNT(tm.emp_id) as strength
                    FROM teams t
                    LEFT JOIN team_members tm ON t.team_id = tm.team_id
                    WHERE tm.emp_id=%s
                    GROUP BY t.team_id, t.team_name, t.created_at
                    ORDER BY t.created_at DESC;
                """, [emp_id])
            
            else:
                pass

            teams = cursor.fetchall()
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'conn' in locals():
            conn.close()

    return render(request, 'view_edit_teams.html', {'teams': teams,
                                                    'name': name, 
                                                    'user_role': role
                                                    })


def team_members_api(request, team_id):
    """
    API endpoint to get members of a specific team.
    """
    members = []
    available_members = []
    try:
        conn = DataOperations.get_db_connection()
        cursor = conn.cursor(dictionary=True)
        # Get team members
        cursor.execute("""
            SELECT m.emp_id, m.first_name, m.last_name, m.email, m.role
            FROM hr_team_members m
            INNER JOIN team_members tm ON m.emp_id = tm.emp_id
            WHERE tm.team_id = %s
        """, [team_id])
        members = cursor.fetchall()
        # Get available members not in this team
        cursor.execute("""
            SELECT m.emp_id, m.first_name, m.last_name, m.email
            FROM hr_team_members m
            WHERE m.emp_id NOT IN (
                SELECT emp_id FROM team_members WHERE team_id = %s
            )
        """, [team_id])
        available_members = cursor.fetchall()
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'conn' in locals():
            conn.close()
    return JsonResponse({'members': members, 'available_members': available_members})

@csrf_exempt
@role_required('Admin', is_api=True)
def add_member_api(request, team_id):
    """
    API endpoint to add a member to a specific team.
    """
    success = False
    try:
        data = json.loads(request.body)
        emp_id = data.get('emp_id')
        # Get user id of the member (from users table, by matching email from hr_team_members)
        user_id = DataOperations.get_user_id_from_emp_id(emp_id)
        if not emp_id:
            return HttpResponseBadRequest("emp_id required")
        conn = DataOperations.get_db_connection()
        cursor = conn.cursor()
        # Insert into team_members, ignore if already exists
        cursor.execute("""
            INSERT IGNORE INTO team_members (team_id, emp_id) VALUES (%s, %s)
        """, [team_id, emp_id])



        # Get team name
        cursor.execute("SELECT team_name FROM teams WHERE team_id=%s", [team_id])
        team_row = cursor.fetchone()
        team_name = team_row[0] if team_row else None

        
        if user_id and DataOperations.get_user_settings(user_id).get('notifications_enabled', False):
            MessageProviders.send_notification(user_id, "Team Update", f"You have been added to the team '{team_name}'", created_by="system", notification_type="Team")
        
        conn.commit()
    except Exception as e:
        if conn:
            conn.rollback()
        return HttpResponseBadRequest(str(e))
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'conn' in locals():
            conn.close()

    return team_members_api(request, team_id)

@csrf_exempt
@role_required(['Admin'], is_api=True)
def remove_member_api(request, team_id):
    """
    API endpoint to remove a member from a specific team.
    """
    try:
        data = json.loads(request.body)
        emp_id = data.get('emp_id')
        if not emp_id:
            return HttpResponseBadRequest("emp_id required")
        conn = DataOperations.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            DELETE FROM team_members WHERE team_id = %s AND emp_id = %s
        """, [team_id, emp_id])

        # Get user id of the member (from users table, by matching email from hr_team_members)
        user_id = DataOperations.get_user_id_from_emp_id(emp_id)
        if not user_id:
            return HttpResponseBadRequest("User not found for the given emp_id")
        # Get team name
        cursor.execute("SELECT team_name FROM teams WHERE team_id=%s", [team_id])
        team_row = cursor.fetchone()
        team_name = team_row[0] if team_row else None
        if user_id and DataOperations.get_user_settings(user_id).get('notifications_enabled', False):
            MessageProviders.send_notification(user_id, "Team Update", f"You have been removed from the team '{team_name}'", created_by="system", notification_type="Team")
        conn.commit()
    except Exception as e:
        if conn:
            conn.rollback()
        return HttpResponseBadRequest(str(e))
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'conn' in locals():
            conn.close()
    return team_members_api(request, team_id)


def generate_jd_id():
    """
    Generate a new job description ID.
    """
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT jd_id FROM recruitment_jds ORDER BY created_at DESC LIMIT 1")
    last = cursor.fetchone()
    if last:
        num = int(last[0][2:]) + 1
    else:
        num = 1
    return f"JD{num:02d}"

@login_required
def create_jd_view(request):
    """
    View to create a new job description.
    """
    print("create_jd -> Request method:", request.method)
    companies = []
    try:
        conn = DataOperations.get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT company_id, company_name FROM customers ORDER BY company_name")
        companies = cursor.fetchall()
        
    except Exception as e:
        print("create_jd_view -> Error fetching companies:", str(e))
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'conn' in locals():
            conn.close()

    return render(request, "jd_create.html", {
        "companies": companies
    })

@csrf_exempt
@login_required
def create_jd(request):
    """
    View to create a new job description.
    """
    print("create_jd -> Request method:", request.method)
    message = error = None
    if request.method == "POST":
        jd_id = generate_jd_id()
        company_id = request.POST.get("company_id")
        jd_summary = request.POST["jd_summary"]
        jd_description = request.POST["jd_description"]
        must_have_skills = request.POST["must_have_skills"]
        good_to_have_skills = request.POST["good_to_have_skills"]
        no_of_positions = int(request.POST.get("no_of_positions", 1))
        jd_status = request.POST.get("jd_status", "active")
        created_by = request.session.get("user_id", None)
        budget_ctc = request.POST.get("budget_ctc")
        experience_required = request.POST.get("experience_required", "")
        education_required = request.POST.get("education_required", "")
        location = request.POST.get("location", "")

        jd_description = DataValidators.sanitize_html(jd_description)

        required_fields = [jd_id, jd_summary, jd_description, must_have_skills, company_id, no_of_positions, budget_ctc]
        if not all(required_fields):
            messages.error(request, "All required fields must be filled.")
            return redirect('jd_create')

        try:
            conn = DataOperations.get_db_connection()
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO recruitment_jds
                (
                    jd_id, company_id, jd_summary, jd_description, 
                    must_have_skills, good_to_have_skills, no_of_positions, 
                    jd_status, created_by, budget_ctc, experience_required, 
                    education_required, location
                )
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                jd_id, company_id, jd_summary, jd_description,
                must_have_skills, good_to_have_skills, no_of_positions, 
                jd_status, created_by, budget_ctc, experience_required, 
                education_required, location
            ))
            conn.commit()

            message = f"Task {escape(jd_id)} {escape(jd_summary)} created successfully!"
            messages.success(request, message)
        except Exception as e:
            print("create_jd -> Error creating JD:", str(e))

            if conn:
                conn.rollback()
            if "Duplicate entry" in str(e):
                error = "A JD with this ID already exists."
            else:
                error = f"Failed to create JD: {escape(str(e))}"
            messages.error(request, error)
        finally:
            if 'cursor' in locals():
                cursor.close()
            if 'conn' in locals():
                conn.close()
        return redirect('jd_create')
    
    return JsonResponse({"error": "Invalid request"}, status=400)

@csrf_exempt
def jd_detail(request, jd_id):
    """
    View to get details of a specific job description.
    """
    print("jd_detail -> Request method:", request.method)
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    if request.method == "GET":
        cursor.execute("""
            SELECT j.*, c.company_name
            FROM recruitment_jds j
            LEFT JOIN customers c ON j.company_id = c.company_id
            WHERE j.jd_id=%s
        """, (jd_id,))
        jd = cursor.fetchone()
        # Also fetch companies for modal dropdown
        cursor.execute("SELECT company_id, company_name FROM customers ORDER BY company_name")
        companies = cursor.fetchall()
        jd['companies'] = companies
        conn.close()
        return JsonResponse(jd)
    elif request.method == "POST":

        company_id = request.POST.get("company_id")
        jd_summary = request.POST["jd_summary"]
        jd_description = request.POST["jd_description"]
        must_have_skills = request.POST["must_have_skills"]
        good_to_have_skills = request.POST["good_to_have_skills"]
        no_of_positions = int(request.POST.get("no_of_positions", 1))
        jd_status = request.POST.get("jd_status", "active")
        total_profiles = int(request.POST.get("total_profiles", 0))
        try:
            cursor.execute("""
                UPDATE recruitment_jds SET
                company_id=%s, jd_summary=%s, jd_description=%s, must_have_skills=%s, good_to_have_skills=%s,
                no_of_positions=%s, jd_status=%s, updated_at=NOW()
                WHERE jd_id=%s
            """, (company_id, jd_summary, jd_description, must_have_skills, good_to_have_skills, no_of_positions, total_profiles, jd_status, jd_id))
            # get emp_ids of all users who are part of team of given jd
            cursor.execute("""
                SELECT DISTINCT tm.emp_id
                FROM team_members tm
                JOIN recruitment_jds j ON tm.team_id = j.assigned_team_id
                WHERE j.jd_id = %s
            """, [jd_id])
            emp_ids = [row['emp_id'] for row in cursor.fetchall()]
            for emp_id in emp_ids:
                user_id = DataOperations.get_user_id_from_emp_id(emp_id)
                if user_id and DataOperations.get_user_settings(user_id).get('notifications_enabled', False):
                    MessageProviders.send_notification(user_id, "JD Update", f"The job description '{jd_id}' has been updated.", created_by="system", notification_type="JD")
            conn.commit()
        except Exception as e:
            if conn:
                conn.rollback()
            return JsonResponse({"error": f"Failed to update JD: {str(e)}"}, status=500)
        finally:
            if 'cursor' in locals():
                cursor.close()
            if 'conn' in locals():
                conn.close()
        return JsonResponse({"success": True})


def download_jd_pdf(request, jd_id):
    """
    View to download a job description as a PDF.
    """
    from .utils import PDFGenerator
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT 
            j.jd_id, j.jd_summary, j.jd_description, j.must_have_skills, 
            j.good_to_have_skills, j.no_of_positions, j.total_profiles,
            j.profiles_selected, j.profiles_completed, j.jd_status, t.team_name,
            c.company_name, j.location, j.experience_required, j.education_required
        FROM recruitment_jds j
        LEFT JOIN customers c ON c.company_id=j.company_id
        LEFT JOIN teams t on t.team_id=j.team_id
        WHERE j.jd_id=%s
    """, (jd_id,))
    jd = cursor.fetchone()
    cursor.close()
    conn.close()
    if not jd:
        return HttpResponse("JD not found", status=404)
    
    html_content = render_to_string('jd_pdf_template.html', {'jd': jd})
    pdf_file = PDFGenerator.generate_pdf_from_html(html_content)
    
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{jd_id}.pdf"'
    return response

def get_customer_list_optimized(request):
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # --- Pagination and Search Setup ---
    search = request.GET.get('search', '').strip()
    page = int(request.GET.get('page', 1))
    per_page = 5
    offset = (page - 1) * per_page
    
    # --- SQL Query Construction ---
    base_select = """
        SELECT 
            company_id, company_name, contact_person_name,
            contact_email, contact_phone, created_at, note 
        FROM 
            customers
    """
    
    base_count = "SELECT COUNT(*) as total FROM customers"
    
    where_clause = ""
    sql_params = []
    
    if search:
        # Search parameters
        search_query = "%" + search + "%"
        search_fields = ["company_name", "contact_person_name", "contact_email", "contact_phone"]
        
        # Build the WHERE clause dynamically
        where_clause = " WHERE " + " OR ".join([f"{field} LIKE %s" for field in search_fields])
        
        # Add search_query four times for the 4 placeholders in COUNT query
        sql_params.extend([search_query] * 4)

    # --- 1. Get Total Count ---
    cursor.execute(base_count + where_clause, sql_params)
    total = cursor.fetchone()['total']
    
    # --- 2. Calculate Pagination ---
    num_pages = (total + per_page - 1) // per_page  # Optimized ceiling division
    page_range = range(1, num_pages + 1)
    
    # --- 3. Get Paginated Customer Data ---
    order_limit_offset = " ORDER BY company_name LIMIT %s OFFSET %s"
    
    # Add LIMIT and OFFSET parameters
    sql_params.extend([per_page, offset])

    # Execute the final query
    cursor.execute(base_select + where_clause + order_limit_offset, sql_params)
    customer_list = cursor.fetchall()

    # --- Cleanup and Final Formatting ---
    cursor.close()
    conn.close()

    # Optimized list comprehension for date conversion
    customer_list = [
        dict(customer, created_at=customer['created_at'].date()) 
        for customer in customer_list
    ]

    # The final output structure remains the same as implied by the original code
    return {
        'customer_list': customer_list,
        'total': total,
        'num_pages': num_pages,
        'page_range': page_range,
        # ... any other context variables
    }

@role_required('Admin')
def create_customer(request):
    """
    View to create a new customer.
    """
    message = error = None
    if request.method == "POST":
        # POST: Handle customer creation
        conn = DataOperations.get_db_connection()
        cursor = conn.cursor(dictionary=True)
        company_name = request.POST.get("company_name", "").strip()
        contact_person_name = request.POST.get("contact_person_name", "").strip()
        contact_email = request.POST.get("contact_email", "").strip()
        contact_phone = request.POST.get("contact_phone", "").strip()
        # Basic validation
        if not company_name or not contact_person_name or not contact_email or not contact_phone:
            error = "All fields are required."
        elif '@' not in contact_email or len(contact_email) > 100 or not DataValidators.is_valid_email(contact_email):
            error = "Invalid email address."
        elif len(company_name) > 255 or len(contact_person_name) > 100:
            error = "Company or contact name too long."

        elif not DataValidators.is_valid_mobile_number(contact_phone):
            error = "Invalid phone number."
        else:
            try:
                cursor.execute("""
                    INSERT INTO customers (company_name, contact_person_name, contact_email, contact_phone)
                    VALUES (%s, %s, %s, %s)
                """, [company_name, contact_person_name, contact_email, contact_phone])
                message = f"Customer '{escape(company_name)}' created successfully!"
                conn.commit()
            except Exception as e:
                if conn:
                    conn.rollback()
                if 'Duplicate entry' in str(e):
                    error = "A customer with this company name or email/phone already exists."
            else:
                # If search is blank, show first 10 companies ordered by created_at
                cursor.execute("""
                    SELECT 
                        company_id, company_name, contact_person_name, 
                        contact_email, contact_phone, created_at, note 
                    FROM 
                        customers 
                    ORDER BY 
                        created_at DESC LIMIT 10
                """)
                customer_list = cursor.fetchall()
                customer_list = [dict(customer, created_at=customer['created_at'].date()) for customer in customer_list]

            finally:
                if message:
                    messages.success(request, message)
                if error:
                    messages.error(request, error)
                if 'cursor' in locals():
                    cursor.close()
                if 'conn' in locals():
                    conn.close()
    
    # GET: Show customer list with pagination and search
    result = get_customer_list_optimized(request)
    customer_list = result['customer_list']
    page = int(request.GET.get('page', 1))
    num_pages = result['num_pages']
    total = result['total']
    page_range = result['page_range']
    search = request.GET.get('search', '').strip()

    return render(request, "create_customer.html", {
        "customer_list": customer_list,
        "page": page,
        "num_pages": num_pages,
        "total": total,
        "page_range": page_range,
        "search": search
    })

@csrf_exempt
@role_required('Admin', is_api=True)
def update_customer(request, company_id):
    """
    View to update an existing customer.
    """
    if request.method == "POST":
        company_name = request.POST.get("company_name", "").strip()
        contact_person_name = request.POST.get("contact_person_name", "").strip()
        contact_email = request.POST.get("contact_email", "").strip()
        contact_phone = request.POST.get("contact_phone", "").strip()
        customer_note = request.POST.get("note", "").strip()
        # Basic validation
        if not company_name or not contact_person_name or not contact_email or not contact_phone:
            return JsonResponse({"error": "All fields are required."}, status=400)
        elif '@' not in contact_email or len(contact_email) > 100 or not DataValidators.is_valid_email(contact_email):
            return JsonResponse({"error": "Invalid email address."}, status=400)
        elif len(company_name) > 255 or len(contact_person_name) > 100:
            return JsonResponse({"error": "Company or contact name too long."}, status=400)
        elif len(contact_phone) > 20 or not DataValidators.is_valid_mobile_number(contact_phone):
            return JsonResponse({"error": "Invalid phone number."}, status=400)
        
        try:
            conn = DataOperations.get_db_connection()
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE customers SET
                company_name=%s, contact_person_name=%s, contact_email=%s, contact_phone=%s, note=%s, updated_at=NOW()
                WHERE company_id=%s
            """, [company_name, contact_person_name, contact_email, contact_phone, customer_note, company_id])
            conn.commit()
            return JsonResponse({"success": True})
        except Exception as e:
            if 'Duplicate entry' in str(e):
                return JsonResponse({"error": "A customer with this company name or email/phone already exists."}, status=400)
            else:
                return JsonResponse({"error": f"Failed to update customer: {str(e)}"}, status=500)
        finally:
            if 'cursor' in locals():
                cursor.close()
            if 'conn' in locals():
                conn.close()
    return JsonResponse({"error": "Invalid request"}, status=400)



def delete_customer(request, company_id):
    if request.method == "DELETE":
        try:
            conn = DataOperations.get_db_connection()
            cursor = conn.cursor()
            cursor.execute("DELETE FROM customers WHERE company_id=%s", [company_id])
            conn.commit()
            return JsonResponse({"success": True})
        except Exception as e:
            return JsonResponse({"error": f"Failed to delete customer: {str(e)}"}, status=500)
        finally:
            if 'cursor' in locals():
                cursor.close()
            if 'conn' in locals():
                conn.close()

@login_required
def customer_details(request):
    """
    View of details of customers for recruiters.

    Here hr_team_members can view the details of customers. and can edit note field only if it is not already filled.
    """
    message = error = None
    if request.method == "POST":
        customer_id = request.POST.get("customer_id")
        note = request.POST.get("note", "").strip()
        if customer_id and note:
            try:
                conn = DataOperations.get_db_connection()
                cursor = conn.cursor()
                cursor.execute("UPDATE customers SET note=%s WHERE company_id=%s", [note, customer_id])
                conn.commit()
                message = "Note updated successfully."
            except Exception as e:
                error = f"Failed to update note: {str(e)}"
            finally:
                if 'cursor' in locals():
                    cursor.close()
                if 'conn' in locals():
                    conn.close()
            # Redirect after POST to prevent resubmission
            from django.urls import reverse
            from django.http import HttpResponseRedirect
            url = reverse('customer_details')
            # Pass message/error via session (or use Django messages framework)
            if message:
                request.session['customer_details_message'] = message
            if error:
                request.session['customer_details_error'] = error
            return HttpResponseRedirect(url)
    # Always show customer list
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT company_id, company_name, contact_person_name, contact_email, contact_phone, created_at, note FROM customers ORDER BY created_at DESC")
    customers = cursor.fetchall()
    for customer in customers:
        customer['created_at'] = customer['created_at'].date()
    name = request.session.get('name', 'Guest')
    user_role_ = request.session.get('role', 'Guest')
    # Retrieve message/error from session if present
    message = request.session.pop('customer_details_message', None)
    error = request.session.pop('customer_details_error', None)
    context = {
        "customers": customers,
        "name": name,
        "user_role": user_role_,
        "message": message,
        "error": error
    }
    return render(request, "customer_details.html", context)


@login_required
def view_edit_jds(request):
    """
    View to edit existing job descriptions.

    """
    user_role_ = request.session.get('role', 'Guest')
    user_id = request.session.get('user_id', None)

    print("view_edit_jds -> Request method:", request.method)
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if user_role_ == 'Admin':
        cursor.execute("""
            SELECT 
                    jd.jd_id, jd.jd_summary, jd.jd_status, jd.no_of_positions, 
                    c.company_name, t.team_name, jd.created_at
            FROM recruitment_jds jd
            LEFT JOIN customers c ON jd.company_id = c.company_id
            LEFT JOIN teams t ON jd.team_id = t.team_id
            ORDER BY jd.created_at DESC
        """)
    elif user_role_ == 'Team_Lead':
        emp_id = DataOperations.get_emp_id_from_user_id(user_id)

        cursor.execute("""
            SELECT DISTINCT j.jd_id, j.jd_summary, j.jd_status, j.no_of_positions, c.company_name, t.team_name, j.created_at
            FROM recruitment_jds j
            LEFT JOIN teams t ON j.team_id = t.team_id
            LEFT JOIN team_members tm ON t.team_id = tm.team_id
            LEFT JOIN customers c ON j.company_id = c.company_id
            WHERE (t.lead_emp_id=%s OR tm.emp_id=%s)
            ORDER BY j.created_at DESC
        """, [emp_id, emp_id])
    elif user_role_ == 'User':
        emp_id = DataOperations.get_emp_id_from_user_id(user_id)
        cursor.execute("""
            SELECT DISTINCT j.jd_id, j.jd_summary, j.jd_status, j.no_of_positions, c.company_name, t.team_name, j.created_at
            FROM recruitment_jds j
            LEFT JOIN teams t ON j.team_id = t.team_id
            LEFT JOIN team_members tm ON t.team_id = tm.team_id
            LEFT JOIN customers c ON j.company_id = c.company_id
            WHERE tm.emp_id=%s
            ORDER BY j.created_at DESC
        """, [emp_id])
    else:
        cursor.execute("""
            SELECT jd_id, jd_summary, jd_status, no_of_positions, c.company_name, t.team_name, created_at
            FROM recruitment_jds jd
            LEFT JOIN customers c ON jd.company_id = c.company_id
            LEFT JOIN teams t ON jd.team_id = t.team_id
            WHERE 1=2
            ORDER BY jd.created_at DESC
        """)


    jds = [
        {
            'jd_id': row['jd_id'],
            'jd_summary': row['jd_summary'],
            'jd_status': row['jd_status'],
            'no_of_positions': row['no_of_positions'],
            'company_name': row['company_name'],
            'team_name': row['team_name'],
            'created_at': row['created_at'].strftime('%Y-%m-%d'),
        }
        for row in cursor.fetchall()
    ]
    # Fetch companies and teams for dropdowns
    cursor.execute("SELECT company_id, company_name FROM customers")
    companies = cursor.fetchall()
    if user_role_ == 'Admin':
        cursor.execute("SELECT team_id, team_name FROM teams")
    else:
        emp_id = DataOperations.get_emp_id_from_user_id(user_id)
        cursor.execute("""
            SELECT DISTINCT t.team_id, t.team_name
            FROM teams t
            LEFT JOIN team_members tm ON t.team_id = tm.team_id
            WHERE t.lead_emp_id=%s OR tm.emp_id=%s
        """, [emp_id, emp_id])

    teams = cursor.fetchall()
    name = request.session.get('name', 'Guest')
    return render(request, 'view_edit_jds.html', {
        'jds': jds,
        'name': name,
        'companies': companies,
        'teams': teams,
        'user_role': user_role_
    })

def get_all_jds(request):
    user_role_ = request.session.get('role', 'Guest')
    user_id = request.session.get('user_id', None)

    print("get_all_jds -> Request method:", request.method)
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if user_role_ == 'Admin':
        cursor.execute("""
            SELECT DISTINCT
                j.jd_id,
                j.jd_summary,
                j.jd_description,
                j.must_have_skills,
                j.good_to_have_skills,
                j.no_of_positions,
                j.jd_status,
                c.company_name,
                t.team_name,
                j.budget_ctc,
                j.location,
                j.experience_required,
                j.education_required,
                j.closure_date,
                j.total_profiles,
                j.profiles_completed,
                j.profiles_in_progress,
                j.profiles_rejected,
                j.profiles_selected,
                j.profiles_on_hold,
                j.updated_at
            FROM recruitment_jds j
            LEFT JOIN customers c ON j.company_id = c.company_id
            LEFT JOIN teams t ON j.team_id = t.team_id
            ORDER BY j.jd_status, j.updated_at DESC
        """)
    elif user_role_ == 'Team_Lead':
        emp_id = DataOperations.get_emp_id_from_user_id(user_id)

        cursor.execute("""
            SELECT DISTINCT
                j.jd_id,
                j.jd_summary,
                j.jd_description,
                j.must_have_skills,
                j.good_to_have_skills,
                j.no_of_positions,
                j.jd_status,
                c.company_name,
                t.team_name,
                j.budget_ctc,
                j.location,
                j.experience_required,
                j.education_required,
                j.closure_date,
                j.total_profiles,
                j.profiles_completed,
                j.profiles_in_progress,
                j.profiles_rejected,
                j.profiles_selected,
                j.profiles_on_hold,
                j.updated_at
            FROM recruitment_jds j
            LEFT JOIN teams t ON j.team_id = t.team_id
            LEFT JOIN team_members tm ON t.team_id = tm.team_id
            LEFT JOIN customers c ON j.company_id = c.company_id
            WHERE (t.lead_emp_id=%s OR tm.emp_id=%s)
            ORDER BY j.updated_at DESC
        """, [emp_id, emp_id])
    elif user_role_ == 'User':
        emp_id = DataOperations.get_emp_id_from_user_id(user_id)
        cursor.execute("""
            SELECT DISTINCT
                j.jd_id,
                j.jd_summary,
                j.jd_description,
                j.must_have_skills,
                j.good_to_have_skills,
                j.no_of_positions,
                j.jd_status,
                c.company_name,
                t.team_name,
                j.budget_ctc,
                j.location,
                j.experience_required,
                j.education_required,
                j.closure_date,
                j.total_profiles,
                j.profiles_completed,
                j.profiles_in_progress,
                j.profiles_rejected,
                j.profiles_selected,
                j.profiles_on_hold,
                j.updated_at
            FROM recruitment_jds j
            LEFT JOIN teams t ON j.team_id = t.team_id
            LEFT JOIN team_members tm ON t.team_id = tm.team_id
            LEFT JOIN customers c ON j.company_id = c.company_id
            WHERE tm.emp_id=%s AND j.jd_status='active'
            ORDER BY j.updated_at DESC
        """, [emp_id])
    else:
        cursor.execute("""
            SELECT DISTINCT
                j.jd_id,
                j.jd_summary,
                j.jd_description,
                j.must_have_skills,
                j.good_to_have_skills,
                j.no_of_positions,
                j.jd_status,
                c.company_name,
                t.team_name,
                j.budget_ctc,
                j.location,
                j.experience_required,
                j.education_required,
                j.closure_date,
                j.total_profiles,
                j.profiles_completed,
                j.profiles_in_progress,
                j.profiles_rejected,
                j.profiles_selected,
                j.profiles_on_hold,
                j.updated_at
            FROM recruitment_jds j
            LEFT JOIN customers c ON j.company_id = c.company_id
            LEFT JOIN teams t ON j.team_id = t.team_id
            WHERE 1=2
            ORDER BY j.updated_at DESC
        """)

    all_rows = cursor.fetchall()
    jds = []
    for row in all_rows:
        jd = {
            'jd_id': row['jd_id'],
            'jd_summary': row['jd_summary'],
            'jd_description': row['jd_description'],
            'must_have_skills': row['must_have_skills'],
            'good_to_have_skills': row['good_to_have_skills'],
            'no_of_positions': row['no_of_positions'],
            'total_profiles': row['total_profiles'],
            'profiles_completed': row['profiles_completed'],
            'profiles_in_progress': row['profiles_in_progress'],
            'profiles_rejected': row['profiles_rejected'],
            'profiles_selected': row['profiles_selected'],
            'profiles_on_hold': row['profiles_on_hold'],
            'jd_status': row['jd_status'],
            'company_name': row['company_name'],
            'team_name': row['team_name'],
            'budget_ctc': row['budget_ctc'],
            'location': row['location'],
            'experience_required': row['experience_required'],
            'education_required': row['education_required'],
            'closure_date': row['closure_date'].isoformat() if row['closure_date'] else ''
        }
        jds.append(jd)

    return JsonResponse({'jds': jds})


def get_jd(request, jd_id):
    """
    View to get details of a specific job description.
    """

    print("get_jd details-> Request method:", request.method)
    print("get_jd details-> JD ID:", jd_id)
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT
            j.jd_id,
            j.jd_summary,
            j.jd_description,
            j.must_have_skills,
            j.good_to_have_skills,
            j.no_of_positions,
            j.jd_status,
            j.company_id,
            c.company_name,
            j.team_id,
            t.team_name,
            j.budget_ctc,
            j.location,
            j.experience_required,
            j.education_required,
            j.closure_date,
            j.total_profiles,
            j.profiles_completed,
            j.profiles_in_progress,
            j.profiles_rejected,
            j.profiles_selected,
            j.profiles_on_hold
        FROM
            recruitment_jds j
        LEFT JOIN
            customers c ON j.company_id = c.company_id
        LEFT JOIN
            teams t ON j.team_id = t.team_id
        WHERE
            j.jd_id = %s;
    """, [jd_id])
    row = cursor.fetchone()
    if not row:
        return JsonResponse({'error': 'JD not found'}, status=404)
    # jd = {
    #     'jd_id': row['jd_id'],
    #     'jd_summary': row['jd_summary'],
    #     'jd_description': row['jd_description'],
    #     'must_have_skills': row['must_have_skills'],
    #     'good_to_have_skills': row['good_to_have_skills'],
    #     'no_of_positions': row['no_of_positions'],
    #     'total_profiles': row['total_profiles'],
    #     'profiles_completed': row['profiles_completed'],
    #     'profiles_in_progress': row['profiles_in_progress'],
    #     'profiles_rejected': row['profiles_rejected'],
    #     'profiles_selected': row['profiles_selected'],
    #     'profiles_on_hold': row['profiles_on_hold'],
    #     'jd_status': row['jd_status'],
    #     'company_id': row['company_id'],
    #     'company_name': row['company_name'],
    #     'team_name': row['team_name'],
    #     'budget_ctc': row['budget_ctc'],
    #     'location': row['location'],
    #     'experience_required': row['experience_required'],
    #     'education_required': row['education_required'],
    #     'closure_date': row['closure_date'].isoformat() if row['closure_date'] else ''
    # }
    return JsonResponse({'jd': row})

@csrf_exempt
@login_required
def update_jd(request, jd_id):
    """
    View to update an existing job description.
    """

    print("update_jd -> Request method:", request.method)
    if request.method == 'POST':
        data = json.loads(request.body)
        conn = DataOperations.get_db_connection()
        cursor = conn.cursor(dictionary=True)
        print("update_jd -> Data received:", data)

        team_id = data['team_id'] if data['team_id'] not in ('', None) else None
        company_id = data['company_id'] if data['company_id'] not in ('', None) else None

        try:
            # make firstly job description as free from risk as it contains html tags
            data['jd_description'] = DataValidators.sanitize_html(data['jd_description'])
            cursor.execute("""
                UPDATE recruitment_jds SET
                    jd_summary=%s,
                    jd_description=%s,
                    must_have_skills=%s,
                    good_to_have_skills=%s,
                    experience_required=%s,
                    education_required=%s,
                    budget_ctc=%s,
                    location=%s,
                    no_of_positions=%s,
                    jd_status=%s,
                    company_id=%s,
                    team_id=%s,
                    closure_date=%s
                WHERE jd_id=%s
            """, [
                data['jd_summary'],
                data['jd_description'],
                data['must_have_skills'],
                data['good_to_have_skills'],
                data['experience_required'],
                data['education_required'],
                data['budget_ctc'],
                data['location'],
                data['no_of_positions'],
                data['jd_status'],
                company_id,
                team_id,
                data['closure_date'] if data['closure_date'] else None,
                jd_id
            ])
            
            conn.commit()
        except Exception as e:
            if conn:
                conn.rollback()
            return JsonResponse({'error': f'Failed to update JD: {str(e)}'}, status=500)

        # Send notification to the members of team if jd is allocated to a team.
        try:
            cursor.execute("""
                SELECT user_id FROM users 
                    WHERE email IN (
                        SELECT email FROM hr_team_members 
                        WHERE emp_id IN ( 
                            SELECT emp_id FROM team_members WHERE team_id=%s
                        )
                    );
            """, [team_id])
            users = cursor.fetchall()
            for user in users:
                if DataOperations.get_user_settings(user['user_id']).get('notifications_enabled', False):
                    MessageProviders.send_notification(user['user_id'], "JD Update", f"JD '{jd_id}' has been updated.", created_by="system", notification_type="Job")
        except Exception as e:
            print("update_jd -> Error sending notifications:", str(e))
        finally:
            cursor.close()
            conn.close()
        return JsonResponse({'success': True})
    return JsonResponse({'error': 'Invalid method'}, status=405)

def assign_jd_data(request):
    """
    View to get data for assigning job descriptions.
    """

    user_role = request.session.get('role', 'Guest')
    user_id = request.session.get('user_id', None)
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    if user_role == 'Admin':
        cursor.execute("""
            SELECT jd_id, jd_summary, jd_status, no_of_positions, company_id
            FROM recruitment_jds
            ORDER BY created_at DESC
        """)
    elif user_role == 'Team_Lead':
        cursor.execute("""
            SELECT emp_id FROM hr_team_members
                    WHERE email=(SELECT email from users WHERE user_id=%s)
                    LIMIT 1;
        """, [user_id])
        row = cursor.fetchone()
        emp_id=row['emp_id']

        cursor.execute("""
            SELECT DISTINCT j.jd_id, j.jd_summary, j.jd_status, j.no_of_positions, j.company_id, j.created_at
            FROM recruitment_jds j
            LEFT JOIN teams t ON j.team_id = t.team_id
            LEFT JOIN team_members tm ON t.team_id = tm.team_id
            WHERE t.lead_emp_id=%s OR tm.emp_id=%s
            ORDER BY j.created_at DESC
        """, [emp_id, emp_id])
    elif user_role == 'User':
        cursor.execute("""
            SELECT emp_id FROM hr_team_members
                    WHERE email=(SELECT email from users WHERE user_id=%s)
                    LIMIT 1;
        """, [user_id])
        row = cursor.fetchone()
        emp_id=row['emp_id']

        cursor.execute("""
            SELECT DISTINCT j.jd_id, j.jd_summary, j.jd_status, j.no_of_positions, j.company_id, j.team_id, j.created_at
            FROM recruitment_jds j
            LEFT JOIN teams t ON j.team_id = t.team_id
            LEFT JOIN team_members tm ON t.team_id = tm.team_id
            WHERE tm.emp_id=%s
            ORDER BY j.created_at DESC
        """, [emp_id])
    else:
        cursor.execute("""
            SELECT jd_id, jd_summary, jd_status, no_of_positions, company_id
            FROM recruitment_jds
            ORDER BY created_at DESC
        """)

    jds = cursor.fetchall()
    cursor.execute("SELECT team_id, team_name FROM teams ORDER BY team_name")
    teams = cursor.fetchall()
    conn.close()
    return JsonResponse({"jds": jds, "teams": teams})

@csrf_exempt
@role_required('Admin', is_api=True)
def assign_jd(request):
    """
    View to assign a job description to a team.
    """
    if request.method != "POST":
        return JsonResponse({"error": "Invalid method"}, status=405)
    data = json.loads(request.body)
    jd_id = data.get("jd_id")
    team_id = data.get("team_id")
    if not jd_id or not team_id:
        return JsonResponse({"error": "JD and Team required"}, status=400)
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)  # <-- Use dictionary=True
    cursor.execute("UPDATE recruitment_jds SET team_id=%s WHERE jd_id=%s", [team_id, jd_id])
    conn.commit()
    cursor.execute("""
        SELECT j.jd_id, j.jd_summary, j.jd_status, j.no_of_positions, j.company_id, c.company_name
        FROM recruitment_jds j
        LEFT JOIN customers c ON j.company_id = c.company_id
        WHERE j.jd_id=%s
    """, [jd_id])
    jd = cursor.fetchone()
    cursor.execute("SELECT team_id, team_name FROM teams WHERE team_id=%s", [team_id])
    team = cursor.fetchone()
    cursor.execute("""
        SELECT m.emp_id, m.first_name, m.last_name, m.email
        FROM hr_team_members m
        INNER JOIN team_members tm ON m.emp_id = tm.emp_id
        WHERE tm.team_id=%s
    """, [team_id])
    members = cursor.fetchall()

    # Send notifications to all team members about the JD assignment
    for member in members:
        user_id = DataOperations.get_user_id_from_emp_id(member['emp_id'])
        if user_id and DataOperations.get_user_settings(user_id).get('notifications_enabled', False):
            MessageProviders.send_notification(user_id, "JD Assignment", f"A new JD has been assigned to your team: {jd['jd_summary']}", created_by="system", notification_type="Job")

    conn.close()
    return JsonResponse({"success": True, "jd": jd, "team": team, "members": members})

@role_required('Admin')
def assign_jd_page(request):
    """
    View to render the job description assignment page.
    """

    name = request.session.get('name', 'Guest')
    # Get JD assignments with pagination
    # try:
    #     page = int(request.GET.get('page', 1))
    #     if page < 1:
    #         page = 1
    # except ValueError:
    #     page = 1
        
    # page_size = 5
    # conn = DataOperations.get_db_connection()
    # cursor = conn.cursor(dictionary=True)
    
    # # Count total JDs for pagination
    # cursor.execute("""
    #     SELECT COUNT(*) as total
    #     FROM recruitment_jds rjd
    #     WHERE jd_status='active'
    # """)
    # total_jds = cursor.fetchone()['total']
    
    # # Calculate pagination values
    # num_pages = (total_jds + page_size - 1) // page_size
    # if page > num_pages and num_pages > 0:
    #     page = num_pages
    
    # offset = (page - 1) * page_size
    # page_range = range(1, num_pages + 1)
    
    # # Fetch paginated JDs with a single query (more efficient)
    # cursor.execute("""
    #     SELECT rjd.jd_id, rjd.jd_summary, rjd.no_of_positions, t.team_name, c.company_name
    #     FROM recruitment_jds rjd
    #     LEFT JOIN teams t ON rjd.team_id = t.team_id
    #     LEFT JOIN customers c ON rjd.company_id = c.company_id
    #     WHERE jd_status='active'
    #     ORDER BY rjd.updated_at DESC
    #     LIMIT %s OFFSET %s
    # """, [page_size, offset])
    
    # jds = cursor.fetchall()
    # cursor.close()
    # conn.close()

    # pagination = {
    #     'current_page': page,
    #     'num_pages': num_pages,
    #     'has_previous': page > 1,
    #     'has_next': page < num_pages,
    #     'previous_page': page - 1 if page > 1 else None,
    #     'next_page': page + 1 if page < num_pages else None,
    #     'page_range': page_range,
    #     'offset': offset  # This is needed for proper row numbering
    # }

    return render(request, "assign_jd.html", {
        'name': name, 
        # 'jds': jds, 
        # 'pagination': pagination
    })

@login_required
def get_jd_assignments_api(request):
    """
    API endpoint to get paginated JD assignments.
    """
    try:
        page = int(request.GET.get('page', 1))
        if page < 1:
            page = 1
    except ValueError:
        page = 1
        
    page_size = 5
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # Count total JDs for pagination
    cursor.execute("""
        SELECT COUNT(*) as total
        FROM recruitment_jds rjd
        WHERE jd_status='active'
    """)
    total_jds = cursor.fetchone()['total']
    
    # Calculate pagination values
    num_pages = (total_jds + page_size - 1) // page_size
    if page > num_pages and num_pages > 0:
        page = num_pages
    
    offset = (page - 1) * page_size
    
    # Fetch paginated JDs with a single query
    cursor.execute("""
        SELECT rjd.jd_id, rjd.jd_summary, rjd.no_of_positions, 
               COALESCE(t.team_name, 'Not Assigned') as team_name, 
               COALESCE(c.company_name, 'Unknown') as company_name
        FROM recruitment_jds rjd
        LEFT JOIN teams t ON rjd.team_id = t.team_id
        LEFT JOIN customers c ON rjd.company_id = c.company_id
        WHERE jd_status='active'
        ORDER BY rjd.updated_at DESC
        LIMIT %s OFFSET %s
    """, [page_size, offset])
    
    jds = cursor.fetchall()
    cursor.close()
    conn.close()

    # Add row numbers starting from offset
    for i, jd in enumerate(jds):
        jd['row_number'] = offset + i + 1

    pagination = {
        'current_page': page,
        'num_pages': num_pages,
        'has_previous': page > 1,
        'has_next': page < num_pages,
        'previous_page': page - 1 if page > 1 else None,
        'next_page': page + 1 if page < num_pages else None,
        'total_jds': total_jds,
        'offset': offset
    }

    return JsonResponse({
        'success': True,
        'jds': jds,
        'pagination': pagination
    })

@login_required
def employee_view_page(request):
    """
    View to render the employee details page.
    """
    name = request.session.get('name', 'Guest')
    return render(request, "employee_view.html",{'name': name})

def employee_view_data(request):
    """
    API endpoint to get data for a specific employee.
    """
    user_role = request.session.get('role', 'Guest')
    user_email = request.session.get('email', None)
    print("employee_view_data -> Request method:", request.method)
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # For admin there will be all members, for team lead there will be members of his teams, for user there will be only his details
    if user_role == 'Admin':
        cursor.execute("""
            SELECT emp_id, first_name, last_name, email, role, status
            FROM hr_team_members
            WHERE status='active'
            ORDER BY first_name, last_name
        """)
    elif user_role == 'Team_Lead':
        cursor.execute("""
                       SELECT
                            T1.emp_id,
                            T1.first_name,
                            T1.last_name,
                            T1.email,
                            T1.role,
                            T1.status
                        FROM
                            hr_team_members AS T1
                        JOIN
                            team_members AS T2
                        ON 
                            T1.emp_id = T2.emp_id
                        JOIN
                            teams AS T3
                        ON
                            T2.team_id = T3.team_id
                        JOIN
                            hr_team_members AS T4
                        ON
                            T3.lead_emp_id = T4.emp_id
                        WHERE
                            T4.email = %s AND T1.status = 'active'
                        ORDER BY T1.first_name, T1.last_name;
            """, [user_email])
    elif user_role == 'User':
        cursor.execute("""
                       SELECT emp_id, first_name, last_name, email, role, status
                        FROM hr_team_members
                        WHERE email = %s AND status='active'
                        ORDER BY first_name, last_name;
            """, [user_email])
    else:
        # TO DO: need to be checked
        cursor.execute("""
            SELECT emp_id, first_name, last_name, email, role, status
            FROM hr_team_members
            WHERE status='active'
            ORDER BY first_name, last_name
        """)
    
    members = cursor.fetchall()
    cursor.close()
    conn.close()
    return JsonResponse({"members": members})

def employee_view_report(request):
    """
    API endpoint to get report data for a specific employee.
    """
    print("employee_view_report -> Request method:", request.method)
    emp_id = request.GET.get("emp_id")
    if not emp_id:
        return JsonResponse({"error": "emp_id required"}, status=400)
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    # Member details
    cursor.execute("""
        SELECT emp_id, first_name, last_name, email, role, status
        FROM hr_team_members WHERE emp_id=%s
    """, [emp_id])
    member = cursor.fetchone()
    # JDs assigned to member (via team membership)
    cursor.execute("""
        SELECT j.jd_id, j.jd_summary, j.jd_status, t.team_name, c.company_name
        FROM recruitment_jds j
        LEFT JOIN teams t ON j.team_id = t.team_id
        LEFT JOIN customers c ON j.company_id = c.company_id
        WHERE j.team_id IN (
            SELECT team_id FROM team_members WHERE emp_id=%s
        )
        ORDER BY j.jd_status, j.created_at DESC
    """, [emp_id])
    jds = cursor.fetchall()
    # Teams member is part of, and JDs for each team
    cursor.execute("""
        SELECT t.team_id, t.team_name
        FROM teams t
        INNER JOIN team_members tm ON t.team_id = tm.team_id
        WHERE tm.emp_id=%s
        ORDER BY t.team_name
    """, [emp_id])
    teams = cursor.fetchall()
    for team in teams:
        cursor.execute("""
            SELECT j.jd_id, j.jd_summary, j.jd_status, c.company_name
            FROM recruitment_jds j
            LEFT JOIN customers c ON j.company_id = c.company_id
            WHERE j.team_id=%s
            ORDER BY j.jd_status, j.created_at DESC
        """, [team['team_id']])
        team['jds'] = cursor.fetchall()

    cursor.close()
    conn.close()
    return JsonResponse({"member": member, "jds": jds, "teams": teams})


def upload_resume_page(request):
    """
    View to render the resume upload page.
    """

    # get the user role and emp_id
    user_role = request.session.get('role', 'Guest')
    user_id = request.session.get('user_id', None)
    
    # Get all JDs for dropdown
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    if user_role in ['Team_Lead', 'User']:

        cursor.execute("""
            SELECT emp_id FROM hr_team_members
                    WHERE email=(SELECT email from users WHERE user_id=%s)
                    LIMIT 1;
        """, [user_id])
        row = cursor.fetchone()
        emp_id=row['emp_id']
        cursor.execute("""
            SELECT DISTINCT jd.jd_id, jd.jd_summary, c.company_name
            FROM recruitment_jds jd
            LEFT JOIN teams t ON jd.team_id = t.team_id
            LEFT JOIN team_members tm ON t.team_id = tm.team_id
            LEFT JOIN customers c ON jd.company_id = c.company_id
            WHERE t.lead_emp_id=%s OR tm.emp_id=%s
            ORDER BY jd.jd_id DESC
        """, [emp_id, emp_id])

    else:
        cursor.execute("""
            SELECT jd.jd_id, jd.jd_summary, c.company_name
            FROM recruitment_jds jd
            LEFT JOIN customers c ON jd.company_id = c.company_id
            ORDER BY jd.jd_id DESC
        """)
    jds = cursor.fetchall()
    cursor.close()
    conn.close()
    name = request.session.get('name', 'Guest')
    return render(request, 'upload_resume.html', {'jds': jds,'name': name})

@csrf_exempt
def upload_resume(request):
    """
    API endpoint to upload a resume.
    """
    if request.method == 'POST':
        try:
            # Validate JD ID and file
            jd_id = request.POST.get('jd_id')
            resume_file = request.FILES.get('resume_file')
            if not jd_id:
                return JsonResponse({'success': False, 'error': 'JD ID is required.'}, status=400)
            if not resume_file:
                return JsonResponse({'success': False, 'error': 'Resume file is required.'}, status=400)

            # Check if JD exists
            conn = DataOperations.get_db_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT company_id FROM recruitment_jds WHERE jd_id=%s", (jd_id,))
            jd_row = cursor.fetchone()
            if not jd_row:
                return JsonResponse({'success': False, 'error': 'Invalid JD ID.'}, status=404)
            customer_id = jd_row['company_id']

            # Validate file type
            allowed_extensions = ['pdf', 'doc', 'docx']
            file_extension = resume_file.name.split('.')[-1].lower()
            if file_extension not in allowed_extensions:
                return JsonResponse({'success': False, 'error': 'Invalid file type. Only PDF, DOC, and DOCX are allowed.'}, status=400)

            # Save file to the appropriate folder
            static_dir = settings.STATICFILES_DIRS[0] if hasattr(settings, 'STATICFILES_DIRS') else os.path.join(settings.BASE_DIR, 'static')
            base_folder = os.path.join(static_dir, 'resumes', jd_id, 'to_be_screened')
            os.makedirs(base_folder, exist_ok=True)
            import uuid
            orig_name, ext = os.path.splitext(resume_file.name)
            unique_str = uuid.uuid4().hex[:8]  # short unique string
            # Format: JDID__OriginalName__Unique.ext
            file_name = f"{jd_id}__{orig_name}__{unique_str}{ext}"
            file_path = os.path.join(base_folder, file_name)
            with open(file_path, 'wb+') as destination:
                for chunk in resume_file.chunks():
                    destination.write(chunk)

            # Save metadata in the database
            cursor.execute("""
                INSERT INTO resumes (jd_id, file_name, file_path, status, customer_id)
                VALUES (%s, %s, %s, %s, %s)
            """, (jd_id, file_name, file_path, 'toBeScreened', customer_id))

            resume_id = cursor.lastrowid
            # inclease the total_profiles count for the JD
            cursor.execute("""
                UPDATE recruitment_jds
                SET total_profiles = total_profiles + 1
                WHERE jd_id = %s
            """, (jd_id,))
            conn.commit()
            return JsonResponse({'success': True, 'resume_id': resume_id})

        except mysql.connector.Error as db_error:
            print("upload_resume -> Database error:", str(db_error))
            conn.rollback()
            return JsonResponse({'success': False, 'error': f'Database error: {str(db_error)}'}, status=500)
        except Exception as e:
            print("upload_resume -> Unexpected error:", str(e))
            conn.rollback()
            return JsonResponse({'success': False, 'error': f'Unexpected error: {str(e)}'}, status=500)
        finally:
            if 'cursor' in locals():
                cursor.close()
            if 'conn' in locals():
                conn.close()

    return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=405)

def recent_resumes(request):
    """
    View to list recent resumes.
    """
    user_role = request.session.get('role', 'Guest')
    user_id = request.session.get('user_id', None)

    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    if user_role in ['Team_Lead', 'User']:
        cursor.execute("""
            SELECT emp_id FROM hr_team_members
                    WHERE email=(SELECT email from users WHERE user_id=%s)
                    LIMIT 1;
        """, [user_id])
        row = cursor.fetchone()
        emp_id=row['emp_id']
        cursor.execute("""
            SELECT DISTINCT r.resume_id, r.file_name, r.jd_id, jd.jd_summary, r.uploaded_on, c.company_name, r.status 
            FROM resumes r
            LEFT JOIN recruitment_jds jd ON r.jd_id=jd.jd_id
            LEFT JOIN customers c ON r.customer_id=c.company_id
            LEFT JOIN teams t ON jd.team_id=t.team_id
            LEFT JOIN team_members tm ON jd.team_id=tm.team_id
            WHERE tm.emp_id=%s OR t.lead_emp_id=%s
            ORDER BY r.resume_id DESC
            LIMIT 20;
        """, [emp_id, emp_id])
    else:
        cursor.execute("""
            SELECT r.resume_id, r.file_name, r.jd_id, jd.jd_summary, r.uploaded_on,
                c.company_name, r.status
            FROM resumes r
            LEFT JOIN recruitment_jds jd ON r.jd_id = jd.jd_id
            LEFT JOIN customers c ON r.customer_id = c.company_id
            ORDER BY r.uploaded_on DESC
            LIMIT 20
        """)
    from .utils import get_display_filename
    resumes = []
    for row in cursor.fetchall():
        # Build file URL for static serving
        file_url = f"/static/resumes/{row['jd_id']}/{row['status']}/{row['file_name']}"
        display_name = get_display_filename(row['file_name'], row['jd_id'])
        resumes.append({
            'resume_id': row['resume_id'],
            'file_name': display_name,
            'jd_id': row['jd_id'],
            'jd_summary': row['jd_summary'],
            'uploaded_on': row['uploaded_on'].strftime('%Y-%m-%d %H:%M'),
            'customer': row['company_name'] or '',
            'status': row['status'],
            'file_url': file_url
        })
    cursor.close()
    conn.close()
    return JsonResponse({'resumes': resumes})

def download_resume(request, resume_id):
    """
    View to download a specific resume.
    """
    # Connect to DB and fetch file path and name
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT file_path, file_name FROM resumes WHERE resume_id=%s", (resume_id,))
    row = cursor.fetchone()
    cursor.close()
    conn.close()
    if not row:
        raise Http404("Resume not found")
    file_path = row['file_path']
    file_name = row['file_name']
    if not os.path.exists(file_path):
        raise Http404("File not found")
    return FileResponse(open(file_path, 'rb'), as_attachment=True, filename=file_name)



def view_parse_resumes_page(request):
    """
    View to render the resume parsing page.
    """
    name = request.session.get('name', 'Guest')
    return render(request, 'view_parse_resumes.html', {'name': name})

@csrf_exempt
def view_parse_resumes(request):
    """
    API endpoint to parse resumes for a specific job description.
    """
    jd_id = request.GET.get('jd_id')
    if not jd_id:
        return JsonResponse({'success': False, 'error': 'JD ID required'}, status=400)
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM resumes WHERE jd_id=%s", (jd_id,))
    resumes = cursor.fetchall()
    parsed_resumes = []
    for r in resumes:
        file_path = r['file_path']
        if not os.path.exists(file_path):
            continue
        try:
            text = textract.process(file_path).decode('utf-8')
            # Simple parsing logic (use regex for better extraction)
            import re
            name = re.search(r'Name[:\- ]*(.*)', text)
            email = re.search(r'[\w\.-]+@[\w\.-]+', text)
            phone = re.search(r'(\+?\d{10,13})', text)
            experience = re.search(r'Experience[:\- ]*(.*)', text)
            summary = text[:300]  # First 300 chars as summary
            from .utils import get_display_filename
            display_name = get_display_filename(r['file_name'], r['jd_id'])
            parsed_resumes.append({
                'resume_id': r['resume_id'],
                    'file_name': display_name,
                'name': name.group(1) if name else '',
                'email': email.group(0) if email else '',
                'phone': phone.group(1) if phone else '',
                'experience': experience.group(1) if experience else '',
                'summary': summary,
                'status': r['status'],
                'file_url': f"/download_resume/{r['resume_id']}/"
            })
        except Exception as e:
            from .utils import get_display_filename
            display_name = get_display_filename(r['file_name'], r['jd_id'])
            parsed_resumes.append({
                'resume_id': r['resume_id'],
                    'file_name': display_name,
                'error': str(e),
                'status': r['status'],
                'file_url': f"/download_resume/{r['resume_id']}/"
            })
    cursor.close()
    conn.close()
    return JsonResponse({'success': True, 'resumes': parsed_resumes})

@csrf_exempt
def update_resume_status(request):
    """
    API endpoint to update the status of a resume.
    """
    print("update_resume_status -> Request method:", request.method)
    if request.method == 'POST':
        resume_id = request.POST.get('resume_id')
        status = request.POST.get('status')
        if not resume_id or status not in ['selected', 'rejected', 'toBeScreened']:
            return JsonResponse({'success': False, 'error': 'Invalid input'}, status=400)
        conn = DataOperations.get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("UPDATE resumes SET status=%s WHERE resume_id=%s", (status, resume_id))
            conn.commit()
            return JsonResponse({'success': True})
        except Exception as e:
            conn.rollback()
            return JsonResponse({'success': False, 'error': f'Failed to update status: {str(e)}'}, status=500)
        finally:
            DataOperations.close_db_connection(conn, cursor)
    return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)


def export_resumes_excel(request):
    """
    API endpoint to export resumes to an Excel file.
    """
    print("export_resumes_excel -> Request method:", request.method)
    jd_id = request.GET.get('jd_id')
    if not jd_id:
        return HttpResponse("JD ID required", status=400)
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT r.file_name, r.status, r.uploaded_on, jd.jd_summary, c.company_name
        FROM resumes r
        LEFT JOIN recruitment_jds jd ON r.jd_id = jd.jd_id
        LEFT JOIN customers c ON r.customer_id = c.company_id
        WHERE r.jd_id = %s
        ORDER BY r.uploaded_on DESC
    """, (jd_id,))
    rows = cursor.fetchall()
    cursor.close()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumes"
    headers = ["File Name", "Status", "Uploaded On", "JD Summary", "Company"]
    ws.append(headers)
    for row in rows:
        ws.append([
            row['file_name'],
            row['status'],
            row['uploaded_on'].strftime('%Y-%m-%d %H:%M') if row['uploaded_on'] else '',
            row['jd_summary'],
            row['company_name']
        ])
    for col in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="resumes_{jd_id}.xlsx"'
    wb.save(response)
    return response

@csrf_exempt
def parse_resumes(request):
    """
    API endpoint to parse resumes for a specific job description.
    """
    jd_id = request.GET.get('jd_id')
    if not jd_id:
        return JsonResponse({'success': False, 'error': 'JD ID required'}, status=400)
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM resumes WHERE jd_id=%s", (jd_id,))
    resumes = cursor.fetchall()
    parser = ResumeParser()
    parsed_resumes = []
    for r in resumes:
        file_path = r['file_path']
        if not os.path.exists(file_path):
            continue
        try:
            result = parser.parse_resume(file_path)
            from .utils import get_display_filename
            display_name = get_display_filename(r['file_name'], r['jd_id'])
            parsed_resumes.append({
                'resume_id': r['resume_id'],
                    'file_name': display_name,
                'name': result.get('Name', ''),
                'email': result.get('Email', ''),
                'phone': result.get('Contact Number', ''),
                'experience': result.get('Work Experience (Years)', ''),
                'status': r['status'],
                'file_url': f"/download_resume/{r['resume_id']}/"
            })
        except Exception as e:
            from .utils import get_display_filename
            display_name = get_display_filename(r['file_name'], r['jd_id'])
            parsed_resumes.append({
                'resume_id': r['resume_id'],
                    'file_name': display_name,
                'error': str(e),
                'status': r['status'],
                'file_url': f"/download_resume/{r['resume_id']}/"
            })
    cursor.close()
    conn.close()
    print("parse_resumes -> Parsed resumes:", parsed_resumes)
    if not parsed_resumes:
        return JsonResponse({'success': False, 'error': 'No resumes found for this JD'}, status=404)
    return JsonResponse({'success': True, 'resumes': parsed_resumes})

@csrf_exempt
def save_candidate_details(request):
           
    """
    API endpoint to save candidate details.
    """
    # Check for duplicate mobile number or email for the same jd_id
    from .utils import compare_mobile_numbers
    # Fetch all candidates for this jd_id except current resume_id (for update)
    if request.method == 'POST':
        data = json.loads(request.body)
        conn = DataOperations.get_db_connection()
        cursor = conn.cursor(dictionary=True)
        # if True:
        try:
            name = data.get('name')
            phone = data.get('phone')
            email = data.get('email')
            skills = data.get('skills')
            education = data.get('education')
            experience = data.get('experience')
            previous_job_profile = data.get('prev_job_profile')
            current_ctc = data.get('current_ctc')
            expected_ctc = data.get('expected_ctc')
            notice_period = data.get('notice_period')
            location = data.get('location')
            resume_id = data.get('resume_id')
            jd_id = data.get('jd_id', None)
            screened_on = data.get('screened_on')
            screen_status = data.get('screen_status')
            screened_remarks = data.get('screened_remarks')
            screening_team = data.get('screening_team')
            hr_member_id = data.get('hr_member_id')
            shared_on = data.get('shared_on')
            recruiters_comment = data.get('recruiter_comments')

            # Normalize shared_on to None if empty or falsy
            shared_on = shared_on or None

            current_ctc = None if current_ctc in ('', None) else current_ctc
            expected_ctc = None if expected_ctc in ('', None) else expected_ctc
            screened_on = None if screened_on in ('', None) else screened_on
            notice_period = None if notice_period in ('', None) else notice_period


            # Validate required fields
            required_fields = [resume_id, name, screen_status, screening_team, hr_member_id, jd_id]
            if not all(required_fields):
                DataOperations.close_db_connection(conn, cursor)
                return JsonResponse({'success': False, 'error': 'Missing required fields'}, status=400)

            # If sharing date is provided, screened_on must also be provided
            if shared_on:
                if not screened_on:
                    DataOperations.close_db_connection(conn, cursor)
                    return JsonResponse({'success': False, 'error': 'Screened on date is required if shared on date is provided.'}, status=400)
                # Sharing date cannot be before screened date
                if screened_on and shared_on < screened_on:
                    DataOperations.close_db_connection(conn, cursor)
                    return JsonResponse({'success': False, 'error': 'Sharing date cannot be before screened date.'}, status=400)
                # Cannot share if status is rejected or to be screened
                if screen_status in ('rejected', 'toBeScreened'):
                    DataOperations.close_db_connection(conn, cursor)
                    return JsonResponse({'success': False, 'error': 'Cannot share profile if status is rejected or to be screened.'}, status=400)

            # Always fetch after SELECT
            cursor.execute("SELECT jd_id, candidate_id,  screen_status, l1_result, l2_result, l3_result, screened_on, shared_on FROM candidates WHERE resume_id=%s", [data.get('resume_id')])
            cd_data = cursor.fetchall()  # Fetch all results, even if you only need one

            if cd_data:

                # screen date and shared date could not be set to past dates if already set
                # Convert string dates to date objects for comparison
                from datetime import datetime, date
                existing_screened_on = cd_data[0]['screened_on']
                existing_shared_on = cd_data[0]['shared_on']
                screened_on_date = None
                shared_on_date = None
                if screened_on:
                    try:
                        screened_on_date = datetime.strptime(screened_on, "%Y-%m-%d").date()
                    except Exception:
                        pass
                if shared_on:
                    try:
                        shared_on_date = datetime.strptime(shared_on, "%Y-%m-%d").date()
                    except Exception:
                        pass
                # Compare only if both are date objects
                if existing_screened_on and screened_on_date and screened_on_date < existing_screened_on:
                    DataOperations.close_db_connection(conn, cursor)
                    return JsonResponse({'success': False, 'error': 'Screened on date cannot be earlier than the existing screened on date.'}, status=400)

                if existing_shared_on and shared_on_date and shared_on_date < existing_shared_on:
                    DataOperations.close_db_connection(conn, cursor)
                    return JsonResponse({'success': False, 'error': 'Sharing date cannot be earlier than the existing sharing date.'}, status=400)

                # count parts
                candidate_prev_data = cd_data[0]
                candidate_new_data = {
                    "jd_id": jd_id,
                    "screen_status": screen_status,
                    "l1_result": candidate_prev_data.get("l1_result"),
                    "l2_result": candidate_prev_data.get("l2_result"),
                    "l3_result": candidate_prev_data.get("l3_result"),
                    "screened_on": screened_on,
                }

               
                # Check here there might be other candidates with the same phone or email for the same JD
                cursor.execute("SELECT resume_id, phone, email FROM candidates WHERE jd_id=%s AND resume_id!=%s", (jd_id, resume_id))
                existing_candidates = cursor.fetchall()
                for ec in existing_candidates:
                    if phone and compare_mobile_numbers(phone, ec['phone']):
                        DataOperations.close_db_connection(conn, cursor)
                        return JsonResponse({'success': False, 'error': f'Duplicate mobile number found for the same JD (Resume ID: {ec["resume_id"]})'}, status=400)
                    if email and email.lower() == (ec['email'] or '').lower():
                        DataOperations.close_db_connection(conn, cursor)
                        return JsonResponse({'success': False, 'error': f'Duplicate email found for the same JD (Resume ID: {ec["resume_id"]})'}, status=400)

                cursor.execute("""
                    UPDATE candidates
                    SET name=%s, phone=%s, email=%s, skills=%s, education=%s, experience=%s,
                        previous_job_profile=%s, current_ctc=%s, expected_ctc=%s, notice_period=%s, 
                        location=%s, screened_on=%s, screen_status=%s, screened_remarks=%s,
                        team_id=%s, hr_member_id=%s, updated_at=NOW(), shared_on=%s, recruiter_comments=%s
                    WHERE resume_id=%s
                """, [
                    name, phone, email, skills, education, 
                    experience, previous_job_profile, current_ctc, 
                    expected_ctc, notice_period, location, screened_on,
                    screen_status, screened_remarks, screening_team, hr_member_id,
                    shared_on, recruiters_comment, resume_id
                ])

                check = DataOperations.update_recruitment_jds(cursor, candidate_prev_data, candidate_new_data)
                if not check:
                    conn.rollback()
                    DataOperations.close_db_connection(conn, cursor)
                    return JsonResponse({'success': False, 'error': 'Error updating recruitment_jds counts'}, status=500)
            else:
                # Check here there might be other candidates with the same phone or email for the same JD
                cursor.execute("SELECT resume_id, phone, email FROM candidates WHERE jd_id=%s", (jd_id,))
                existing_candidates = cursor.fetchall()
                for ec in existing_candidates:
                    if phone and compare_mobile_numbers(phone, ec['phone']):
                        DataOperations.close_db_connection(conn, cursor)
                        return JsonResponse({'success': False, 'error': f'Duplicate mobile number found for the same JD (Resume ID: {ec["resume_id"]})'}, status=400)
                    if email and email.lower() == (ec['email'] or '').lower():
                        DataOperations.close_db_connection(conn, cursor)
                        return JsonResponse({'success': False, 'error': f'Duplicate email found for the same JD (Resume ID: {ec["resume_id"]})'}, status=400)
                    
                cursor.execute("""
                    INSERT INTO candidates (
                                jd_id, resume_id, name, phone, email, skills,
                                education, experience, previous_job_profile, 
                                current_ctc, expected_ctc, notice_period, 
                                location, screened_on, screen_status, screened_remarks,
                                recruiter_comments, team_id, hr_member_id, shared_on
                            )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, [
                    jd_id, resume_id, name, phone,
                    email, skills, education, experience,
                    previous_job_profile, current_ctc, 
                    expected_ctc, notice_period, location, 
                    screened_on, screen_status, screened_remarks, recruiters_comment,
                    screening_team, hr_member_id, shared_on
                ])
                # increase the total_profiles count for the JD
                
                screen_addition = ', profiles_in_progress = profiles_in_progress + 1' if screen_status=='selected' else ''
                cursor.execute(f"""
                    UPDATE recruitment_jds SET total_profiles = total_profiles + 1 {screen_addition}
                    WHERE jd_id = (SELECT jd_id FROM candidates WHERE resume_id = %s)
                """, (resume_id,))

            conn.commit()
            response = {'success': True}

            return JsonResponse(response)
        except Exception as e:
            conn.rollback()
            print("save_candidate_details -> Error:", str(e))
            response = {'success': False, 'error': str(e)}
            return JsonResponse(response, status=500)
        # finally:
        #     DataOperations.close_db_connection(conn, cursor)
    return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)

@csrf_exempt
def update_candidate_screen_status(request):
    """
    API endpoint to update the screening status of a candidate.
    """
    if request.method == 'POST':
        resume_id = request.POST.get('resume_id')
        status = request.POST.get('status')
        conn = DataOperations.get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT jd_id, screen_status, l1_result, l2_result, l3_result, screened_on FROM candidates WHERE resume_id=%s", (resume_id,))
        candidate_prev_data = cursor.fetchone()
        if not candidate_prev_data:
            return JsonResponse({'success': False, 'error': 'Candidate not found'}, status=404)
        candidate_new_data = {
            "jd_id": candidate_prev_data['jd_id'],
            "screen_status": status,
            "l1_result": candidate_prev_data.get("l1_result"),
            "l2_result": candidate_prev_data.get("l2_result"),
            "l3_result": candidate_prev_data.get("l3_result"),
            "screened_on": candidate_prev_data.get("screened_on"),
        }

        

        cursor.execute("""
            UPDATE candidates SET screen_status=%s WHERE resume_id=%s
        """, (status, resume_id))

        check = DataOperations.update_recruitment_jds(cursor, candidate_prev_data, candidate_new_data)
        if not check:
            conn.rollback()
            return JsonResponse({'success': False, 'error': 'Error updating recruitment_jds counts'}, status=500)

        conn.commit()
        cursor.close()
        conn.close()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)

def get_jd_team_members(request):
    """
    API endpoint to get team members for a specific job description.
    """
    jd_id = request.GET.get('jd_id')
    if not jd_id:
        return JsonResponse({'success': False, 'error': 'JD ID required'}, status=400)
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT team_id FROM recruitment_jds WHERE jd_id=%s", (jd_id,))
    jd_row = cursor.fetchone()
    if not jd_row or not jd_row['team_id']:
        cursor.close()
        conn.close()
        return JsonResponse({'success': True, 'team_id': None, 'team_name': '', 'members': []})
    team_id = jd_row['team_id']
    cursor.execute("SELECT team_name FROM teams WHERE team_id=%s", (team_id,))
    team_row = cursor.fetchone()
    team_name = team_row['team_name'] if team_row else ''
    cursor.execute("""
        SELECT m.emp_id, m.first_name, m.last_name, m.email
        FROM hr_team_members m
        INNER JOIN team_members tm ON m.emp_id = tm.emp_id
        WHERE tm.team_id = %s AND m.status='active'
        ORDER BY m.first_name, m.last_name
    """, (team_id,))
    members = cursor.fetchall()
    cursor.close()
    conn.close()
    return JsonResponse({'success': True, 'team_id': team_id, 'team_name': team_name, 'members': members})

from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse


@csrf_exempt
@login_required
def get_candidate_details(request):
    """
    API endpoint to get details of a specific candidate.
    """
    print("get_candidate_details -> Request method:", request.method)
    resume_id = request.GET.get('resume_id')
    if not resume_id:
        return JsonResponse({'success': False, 'error': 'resume_id required'}, status=400)
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM candidates WHERE resume_id=%s ORDER BY candidate_id DESC LIMIT 1", (resume_id,))
    candidate = cursor.fetchone()
    cursor.close()
    conn.close()
    return JsonResponse({'success': True, 'candidate': candidate})

# @login_required
# def search_jds(request):
#     user_id = request.session.get('user_id', None)
#     user_role = request.session.get('role', 'Guest')
#     conn = DataOperations.get_db_connection()
#     cursor = conn.cursor(dictionary=True)

#     emp_id = DataOperations.get_emp_id_from_user_id(user_id)
#     jds = []

#     # Implement search functionality
#     search_query = request.GET.get('search', '').strip()

#     SQL_QUERY = """
#         SELECT 
#             j.jd_id, j.jd_summary, c.company_name 
#         FROM recruitment_jds j
#         LEFT JOIN customers c ON j.company_id = c.company_id
#     """
#     params = []
#     if user_role in ['Team_Lead', 'User']:
#         SQL_QUERY += """
#             LEFT JOIN teams t ON j.team_id = t.team_id
#             LEFT JOIN team_members tm ON t.team_id = tm.team_id
#             WHERE t.lead_emp_id=%s OR tm.emp_id=%s
#         """
#         params.extend([emp_id, emp_id])

#     if search_query:
#         query_filter = "j.jd_id LIKE %s OR j.jd_summary LIKE %s OR j.jd_description LIKE %s OR c.company_name LIKE %s"
#         query = f"""
#             {SQL_QUERY}
#             WHERE {query_filter} AND j.jd_status='active'
#             ORDER BY j.jd_id DESC
#             LIMIT 10
#         """
#         params = [*params, f'%{search_query}%', f'%{search_query}%', f'%{search_query}%', f'%{search_query}%']
#         cursor.execute(query, (*params,))
#     else:
#         query = f"""
#             {SQL_QUERY}
#             WHERE j.jd_status='active'
#             ORDER BY j.updated_at DESC
#             LIMIT 10
#         """
#         cursor.execute(query, params)

#     jds = cursor.fetchall()
#     cursor.close()
#     conn.close()
#     return JsonResponse({'success': True, 'jds': jds})

@login_required
def search_jds(request):
    user_id = request.session.get('user_id', None)
    user_role = request.session.get('role', 'Guest')
    search_query = request.GET.get('search', '').strip()
    
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)

    emp_id = DataOperations.get_emp_id_from_user_id(user_id)
    jds = []
    
    # --- 1. Base Query Structure ---
    # Start with the SELECT and FROM clauses
    SQL_SELECT_FROM = """
        SELECT 
            j.jd_id, j.jd_summary, c.company_name 
        FROM recruitment_jds j
        LEFT JOIN customers c ON j.company_id = c.company_id
    """
    
    # List to hold all WHERE conditions and parameters
    where_conditions = []
    params = []
    joins = []
    
    # --- 2. Role-Based Access Control (ACL) Filter ---
    if user_role in ['Team_Lead', 'User']:
        cursor.execute("""
            SELECT 
                team_id FROM team_members
            WHERE emp_id=%s
        """, (emp_id,))
        teams = cursor.fetchall()
        team_ids = [team['team_id'] for team in teams]
        team_ids_placeholders = f"({', '.join(['%s'] * len(team_ids))})" if team_ids else "(NULL)"  # Handle case with no teams

        # Add necessary JOINS for role-based filtering
        joins.append("LEFT JOIN teams t ON j.team_id = t.team_id")
        joins.append("LEFT JOIN team_members tm ON t.team_id = tm.team_id")
        
        # Add the WHERE condition for access control
        where_conditions.append(f"(t.lead_emp_id=%s OR tm.emp_id=%s) AND j.team_id IN {team_ids_placeholders}")
        params.extend([emp_id, emp_id, *team_ids])

    # --- 3. Mandatory JD Status Filter ---
    where_conditions.append("j.jd_status='active'")
    # Note: No %s placeholder needed here as 'active' is a fixed string.

    # --- 4. Optional Search Query Filter ---
    if search_query:
        search_filter = "(j.jd_id LIKE %s OR j.jd_summary LIKE %s OR j.jd_description LIKE %s OR c.company_name LIKE %s)"
        where_conditions.append(search_filter)
        
        search_term = f'%{search_query}%'
        params.extend([search_term, search_term, search_term, search_term])

    # --- 5. Final Query Construction ---
    
    # Combine JOINS and WHERE clauses
    SQL_QUERY = SQL_SELECT_FROM + " " + " ".join(joins)
    
    # Join all conditions with AND and prepend the WHERE keyword
    if where_conditions:
        SQL_QUERY += " WHERE " + " AND ".join(where_conditions)
    
    # Add ORDER BY and LIMIT
    if search_query:
        # Order by jd_id when searching, as per original logic
        SQL_QUERY += " ORDER BY j.jd_id DESC LIMIT 10"
    else:
        # Order by updated_at when not searching, as per original logic
        SQL_QUERY += " ORDER BY j.updated_at DESC LIMIT 10"

    # --- 6. Execution and Return ---
    cursor.execute(SQL_QUERY, params)
    jds = cursor.fetchall()

    cursor.close()
    conn.close()
    return JsonResponse({'success': True, 'jds': jds})

@login_required
def candidate_pipeline_page(request):
    """
    View to render the candidate handling page.
    """
    user_id = request.session.get('user_id', None)
    user_role = request.session.get('role', 'Guest')
    candidates_info = []
    

    return render(request, 'candidate_handle.html', {
        'user_role': user_role,
        'candidates_info': candidates_info,
    })

@login_required
def api_candidates_pipeline(request):
    """
    API endpoint to get filtered candidates for the pipeline.
    """
    user_id = request.session.get('user_id', None)
    user_role = request.session.get('role', 'Guest')
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # get emp_id
    emp_id = DataOperations.get_emp_id_from_user_id(user_id)
    
    # Get filter parameters
    jd_id = request.GET.get('jd_id', '').strip()
    search_query = request.GET.get('search', '').strip()
    
    # Build base query for candidates
    base_query = """
        SELECT 
            c.candidate_id,
            c.name,
            c.email,
            c.phone,
            c.resume_id,
            c.jd_id,
            c.screen_status,
            c.l1_result,
            c.l2_result,
            c.l3_result,
            c.offer_status,
            j.jd_summary,
            comp.company_name,
            r.file_name as resume_filename
        FROM candidates c
        LEFT JOIN recruitment_jds j ON c.jd_id = j.jd_id
        LEFT JOIN customers comp ON j.company_id = comp.company_id
        LEFT JOIN resumes r ON c.resume_id = r.resume_id
    """
    
    params = []
    where_conditions = []
    
    # Role-based filtering
    if user_role in ['Team_Lead', 'User']:
        base_query += """
            LEFT JOIN teams t ON j.team_id = t.team_id
            LEFT JOIN team_members tm ON t.team_id = tm.team_id
        """
        where_conditions.append("(t.lead_emp_id = %s OR tm.emp_id = %s) AND c.hr_member_id = %s")
        params.extend([emp_id, emp_id, emp_id])
    
    # JD filtering
    if jd_id and jd_id != 'all':
        where_conditions.append("c.jd_id = %s")
        params.append(jd_id)
    
    # Search filtering
    if search_query:
        where_conditions.append("""
            (c.name LIKE %s OR 
            c.email LIKE %s OR 
            j.jd_summary LIKE %s OR 
            comp.company_name LIKE %s OR
            j.jd_id LIKE %s
        )
        """)
        search_param = f'%{search_query}%'
        params.extend([search_param, search_param, search_param, search_param, search_param])
    
    where_conditions.append("j.jd_status='active'")
    # Combine query with WHERE conditions
    if where_conditions:
        base_query += " WHERE " + " AND ".join(where_conditions)
    
    base_query += " ORDER BY c.updated_at DESC LIMIT 50"
    
    try:
        cursor.execute(base_query, params)
        candidates = cursor.fetchall()
        
        # Process candidates data to add initials and other UI data
        candidate_info = []
        for candidate in candidates:
            # Generate initials from name
            name_parts = candidate['name'].strip().split()
            if len(name_parts) >= 2:
                initials = f"{name_parts[0][0]}{name_parts[-1][0]}".upper()
            elif len(name_parts) == 1:
                initials = name_parts[0][:2].upper()
            else:
                initials = "??"
            
            # Create job summary display text
            job_summary = candidate['jd_summary'] or candidate['jd_title'] or f"JD-{candidate['jd_id']}"
            
            candidate_info.append({
                'candidate_id': candidate['candidate_id'],
                'name': candidate['name'],
                'email': candidate['email'],
                'phone': candidate['phone'],
                'initials': initials,
                'job_summary': job_summary,
                'company_name': candidate['company_name'] or 'No Company',
                'screen_status': candidate['screen_status'],
                'l1_result': candidate['l1_result'],
                'l2_result': candidate['l2_result'],
                'l3_result': candidate['l3_result'],
                'offer_status': candidate['offer_status'],
                'jd_id': candidate['jd_id'],
                'resume_filename': candidate['resume_filename']
            })
        
        cursor.close()
        conn.close()
        
        return JsonResponse({
            'success': True,
            'candidates': candidate_info,
            'total_count': len(candidate_info)
        })
        
    except Exception as e:
        print(f"api_candidates_pipeline -> Error: {str(e)}")
        cursor.close()
        conn.close()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

# Action executions in take action popup

@csrf_exempt
@login_required
def candidate_action_handler(request):
    """
    Main handler for candidate actions from the frontend.
    Routes to appropriate action functions based on action type.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Only POST method allowed'}, status=405)
    
    try:
        # Get session data
        user_id = request.session.get('user_id', None)
        user_email = request.session.get('email', None)
        user_role = request.session.get('role', "Guest")
        
        # Parse request data
        data = json.loads(request.body)
        action = data.get('actionType')
        candidate_id = data.get('candidate_id')
        comment = data.get('actionComments', '').strip()
        notify_status = data.get('notifyCandidate', False)

        conn = DataOperations.get_db_connection()
        cursor = conn.cursor(dictionary=True)
            
        # Validate required fields
        if not action or not candidate_id:
            return JsonResponse({'success': False, 'error': 'Action type and candidate ID are required'}, status=400)

        cursor.execute("""
            SELECT candidate_id, hr_member_id 
            FROM candidates
            WHERE candidate_id=%s
        """, (candidate_id,))
        candidate = cursor.fetchone()

        DataOperations.close_db_connection(conn, cursor)

        if not candidate:
            return JsonResponse({'success': False, 'error': 'Candidate not found'}, status=404)
        
        if user_role in ["Team_Lead", "User"] and not user_id:
            emp_id = DataOperations.get_emp_id_from_user_id(user_id)
            if not emp_id:
                return JsonResponse({'success': False, 'error': 'Employee ID not found for the user'}, status=403)
            
            if emp_id != candidate['hr_member_id']:
                return JsonResponse({'success': False, 'error': "You are not authorized to access this candidate"})
        
        # Route to appropriate action based on action type
        if action == 'advance_stage':
            # TODO: Call advance_to_next_stage method
            return advanced_to_next_stage(request, candidate_id, comment, notify_status)
            
        elif action == 'reject':
            rejection_reason = data.get('rejectionReason', '')
            # TODO: Call reject_candidate method with rejection reason
            return reject_candidate(request, candidate_id, rejection_reason, comment, notify_status)
            
        elif action == 'put_on_hold':
            hold_reason = data.get('holdReason', '')
            # TODO: Call put_candidate_on_hold method with hold reason
            return put_candidate_on_hold(request, candidate_id, hold_reason, comment, notify_status)
        elif action == 'back_to_previous_stage':
            return back_to_previous_stage(request, candidate_id, comment)
        elif action == 'schedule_interview':
            interview_data = {
                'interview_date': data.get('interviewDate'),
                'interview_type': data.get('interviewType'),
                'interview_level': data.get('interviewLevel'),
                'interviewer': data.get('interviewer'),
                'interviewer_email': data.get('interviewerEmail'),
                'interview_link': data.get('interviewLink')
            }
            # TODO: Call schedule_interview_for_candidate method with interview data
            return schedule_interview_for_candidate(request, candidate_id, interview_data)
            
        elif action == 'send_offer':
            # TODO: Call send_offer_to_candidate method
            return send_offer_to_candidate(request, candidate_id)
            
        elif action == 'withdraw_offer':
            # TODO: Call withdraw_candidate_offer method
            return withdraw_candidate_offer(request, candidate_id)
            
        elif action == 'mark_hired':
            # TODO: Call mark_candidate_as_hired method
            return mark_candidate_as_hired(request, candidate_id)
            
        elif action == 'mark_resigned':
            # TODO: Call mark_candidate_as_resigned method
            return mark_candidate_as_resigned(request, candidate_id)
            
        else:
            return JsonResponse({'success': False, 'error': f'Unknown action type: {action}'}, status=400)
            
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        print(f"candidate_action_handler -> Error: {str(e)}")
        return JsonResponse({'success': False, 'error': 'Internal server error'}, status=500) 


@login_required
def advanced_to_next_stage(request, candidate_id: int, comment: str, notify_status: bool):
    """
    Advances a candidate's stage status to 'selected' at the current level (Screening, L1, L2, or L3).
    This function should only be called if the candidate is currently 'toBeScreened' at their last passed stage.
    """
    
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    today = date.today().isoformat()
    
    try:
        # 1. Determine current status and check for existing rejections/selections
        # Fetch all relevant status columns in one query
        cursor.execute("""
            SELECT 
                screen_status, l1_result, l2_result, l3_result, 
                name, email, jd_id
            FROM candidates
            WHERE candidate_id=%s
        """, (candidate_id,))
        candidate_data = cursor.fetchone()

        if not candidate_data:
            return JsonResponse({'success': False, 'error': 'Candidate data retrieval failed'}, status=500)

        # 2. Validate if advancement is possible

        # Check for rejection at any stage (prevents advancing a rejected candidate)
        rejection_statuses = ['rejected', 'onHold']
        if any(candidate_data[f'{level}_result'] in rejection_statuses for level in ['l1', 'l2', 'l3']):
            return JsonResponse({'success': False, 'error': 'Cannot advance: Candidate is already rejected or on hold at a previous level.'}, status=400)
        if candidate_data['screen_status'] in rejection_statuses:
             return JsonResponse({'success': False, 'error': 'Cannot advance: Candidate is already rejected or on hold at the screening level.'}, status=400)

        # Determine the stage to update
        update_column = None
        date_column = None
        comment_column = None
        new_status = 'selected'
        stage_name = ""

        # Logic to find the FIRST status that is NOT 'selected'
        if candidate_data['screen_status'] != 'selected':
            update_column = 'screen_status'
            date_column = 'screened_on'
            comment_column = 'screened_remarks'
            stage_name = "Screening"
        elif candidate_data['l1_result'] != 'selected':
            update_column = 'l1_result'
            date_column = 'l1_date'
            comment_column = 'l1_comments'
            stage_name = "L1 Interview"
        elif candidate_data['l2_result'] != 'selected':
            update_column = 'l2_result'
            date_column = 'l2_date'
            comment_column = 'l2_comments'
            stage_name = "L2 Interview"
        elif candidate_data['l3_result'] != 'selected':
            update_column = 'l3_result'
            date_column = 'l3_date'
            comment_column = 'l3_comments'
            stage_name = "L3 Interview"
        else:
            # Candidate is already selected at L3
            return JsonResponse({'success': False, 'error': 'Candidate is already fully selected (L3) and cannot be advanced further.'}, status=400)
            
        
        # 3. Update candidate status, date, and comments for the determined stage

        # We must verify the current status before advancing
        current_status = candidate_data.get(update_column)
        if current_status != 'toBeScreened':
             return JsonResponse({'success': False, 'error': f'Candidate is currently set to "{current_status}" at the {stage_name} stage and cannot be automatically advanced. Must be "toBeScreened"'}, status=400)


        # Dynamic SQL update based on the determined stage
        update_query = f"""
            UPDATE candidates 
            SET 
                {update_column}=%s, 
                {date_column}=%s, 
                {comment_column}=%s,
                updated_at=CURRENT_TIMESTAMP
            WHERE candidate_id=%s
        """
        update_params = [new_status, today, comment, candidate_id]
        
        cursor.execute(update_query, tuple(update_params))
        conn.commit()

        # 4. Notify candidate if required (TODO)
        if notify_status:
            # TODO: Implement robust email notification logic here.
            # Example: send_notification_email(candidate_data['email'], stage_name, new_status)
            pass

        # Log activity
        # TODO: Add logic to insert a record into candidate_activities table.
        # activity_comment = f"Advanced to next stage: {stage_name} marked as {new_status}"
        # DataOperations.log_candidate_activity(candidate_id, user_id, 'ADVANCE', activity_comment)

        return JsonResponse({
            'success': True, 
            'message': f'Candidate {candidate_id} successfully advanced. {stage_name} marked as "{new_status}".',
            'new_status_field': update_column,
            'new_status_value': new_status
        })

    except Exception as e:
        conn.rollback()
        print(f"advanced_to_next_stage -> Error: {str(e)}")
        return JsonResponse({'success': False, 'error': 'Database or internal server error during advancement'}, status=500)
        
    finally:
        DataOperations.close_db_connection(conn, cursor)


# def advanced_to_next_stage(request, candidate_id, comment=None, notify_candidate=False):
#     """

#     """
#     conn = DataOperations.get_db_connection()
#     cursor = conn.cursor(dictionary=True)
#     # TODO: Implement advancing to the next stage logic
#     # - Determine current stage
#     cursor.execute("SELECT * FROM candidates WHERE candidate_id=%s", (candidate_id,))
#     candidate = cursor.fetchone()
#     if not candidate:
#         DataOperations.close_db_connection(conn, cursor)
#         return JsonResponse({'success': False, 'error': 'Candidate not found'}, status=404)
#     current_stage = {
#         'screen_status': candidate['screen_status'],
#         'l1_result': candidate['l1_result'],
#         'l2_result': candidate['l2_result'],
#         'l3_result': candidate['l3_result'],
#         'offer_status': candidate['offer_status']
#     }
    

#     # - Validate if advancement is possible
#     anywhere_rejected = current_stage['screen_status'] == 'rejected' or \
#         current_stage['l1_result'] == 'rejected' or \
#         current_stage['l2_result'] == 'rejected' or \
#         current_stage['l3_result'] == 'rejected'

#     if anywhere_rejected:
#         DataOperations.close_db_connection(conn, cursor)
#         return JsonResponse({'success': False, 'error': 'Cannot advance candidate who has been rejected in any stage'}, status=400)
    
#     if current_stage['l3_result'] == 'selected':
#         DataOperations.close_db_connection(conn, cursor)
#         return JsonResponse({'success': False, 'error': 'Cannot advance candidate who is at last stage'}, status=400)

#     # - Get last completed stage
#     last_completed_stage = None
#     next_stage = None
#     if current_stage['l3_result'] in ['selected', 'rejected']:
#         last_completed_stage = 'l3'
#         next_stage = None

#     elif current_stage['l2_result'] in ['selected', 'rejected']:
#         last_completed_stage = 'l2'
#         next_stage = 'l3'
#     elif current_stage['l1_result'] in ['selected', 'rejected']:
#         last_completed_stage = 'l1'
#         next_stage = 'l2'
#     elif current_stage['screen_status'] == 'selected':
#         last_completed_stage = 'screen'
#         next_stage = 'l1'
#     else:
#         last_completed_stage=None
#         next_stage = 'screen'

#     # - Update candidate status
#     if next_stage == 'screen':
#         screen_status = 'selected'
#         screen_remarks = comment
#         screened_on = datetime.now().date().format('%Y-%m-%d')
#         cursor.execute("""
#             UPDATE candidates
#             SET screen_status=%s,
#                 screened_remarks=%s,
#                 screened_on=%s,
#                 updated_at=NOW()
#             WHERE candidate_id=%s
#         """, (screen_status, screen_remarks, screened_on, candidate_id))
    
#     elif next_stage=='l1':
#         l1_result = 'selected'
#         l1_remarks = comment
#         l1_interviewed_on = datetime.now().date().format('%Y-%m-%d')
#         cursor.execute("""
#             UPDATE candidates
#             SET l1_result=%s,
#                 l2_comments=%s,
#                 l1_date=%s,
#                 updated_at=NOW()
#             WHERE candidate_id=%s
#         """, (l1_result, l1_remarks, l1_interviewed_on, candidate_id))

#     # - Add comments if provided
#     # - Update candidate Mustor
#     # - Notify candidate if required
#     return JsonResponse({'success': True, 'message': 'Advanced to next stage successfully.'})

@login_required
def reject_candidate(request, candidate_id: int, rejection_reason: str = None, comment: str = None, notify_candidate: bool = False):
    """
    Rejects a candidate at their current stage (Screening, L1, L2, or L3).
    The stage rejected is the first one that is NOT 'selected'.
    """
    
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    today = date.today().isoformat()
    
    # Combined comment/reason for the DB remarks column
    full_comment = f"REASON: {rejection_reason.strip()}"
    if comment and comment.strip():
        full_comment += f" | REMARKS: {comment.strip()}"

    try:
        # 1. Determine current status (and fetch data for notification/logging)
        cursor.execute("""
            SELECT 
                screen_status, l1_result, l2_result, l3_result, 
                name, email, jd_id
            FROM candidates
            WHERE candidate_id=%s
        """, (candidate_id,))
        candidate_data = cursor.fetchone()

        if not candidate_data:
            return JsonResponse({'success': False, 'error': 'Candidate data retrieval failed or candidate not found'}, status=404)

        # 2. Validate if rejection is not possible

        # Check 1: Fully Selected
        if candidate_data['l3_result'] == 'selected':
            return JsonResponse({'success': False, 'error': 'Cannot reject: Candidate is already fully selected (L3) and moved to the final stage.'}, status=400)

        # Check 2: Already Rejected/On Hold (at any stage, as per user request)
        rejection_statuses = ['rejected']
        if any(candidate_data[f'{level}_result'] in rejection_statuses for level in ['l1', 'l2', 'l3']):
            return JsonResponse({'success': False, 'error': 'Cannot reject: Candidate is already formally rejected at a previous level.'}, status=400)
        if candidate_data['screen_status'] in rejection_statuses:
             return JsonResponse({'success': False, 'error': 'Cannot reject: Candidate is already formally rejected at the screening level.'}, status=400)

        
        # Determine the stage to update (The first stage that is NOT 'selected')
        update_column = None
        date_column = None
        remarks_column = None
        stage_name = ""
        
        # Determine the earliest possible stage that hasn't been successfully passed
        if candidate_data['screen_status'] != 'selected':
            update_column = 'screen_status'
            date_column = 'screened_on'
            remarks_column = 'screened_remarks'
            stage_name = "Screening"
        elif candidate_data['l1_result'] != 'selected':
            update_column = 'l1_result'
            date_column = 'l1_date'
            remarks_column = 'l1_comments'
            stage_name = "L1 Interview"
        elif candidate_data['l2_result'] != 'selected':
            update_column = 'l2_result'
            date_column = 'l2_date'
            remarks_column = 'l2_comments'
            stage_name = "L2 Interview"
        elif candidate_data['l3_result'] != 'selected':
            update_column = 'l3_result'
            date_column = 'l3_date'
            remarks_column = 'l3_comments'
            stage_name = "L3 Interview"
        
        # This condition is technically redundant due to Check 1, but kept for clarity
        if not update_column:
             return JsonResponse({'success': False, 'error': 'Candidate status is ambiguous or fully selected.'}, status=400)


        # 4. Validate member rejection reason
        if not rejection_reason or not rejection_reason.strip():
            return JsonResponse({'success': False, 'error': 'Rejection reason must be provided to reject the candidate.'}, status=400)

        
        # 5. Update candidate status, date, and remarks
        new_status = 'rejected'
        
        update_query = f"""
            UPDATE candidates 
            SET 
                {update_column}=%s, 
                {date_column}=%s, 
                {remarks_column}=%s,
                updated_at=CURRENT_TIMESTAMP
            WHERE candidate_id=%s
        """
        update_params = [new_status, today, full_comment, candidate_id]
        
        cursor.execute(update_query, tuple(update_params))
        conn.commit()

        # 6. Notify candidate if required (TODO)
        if notify_candidate:
            # TODO: Implement robust email notification logic here using
            # candidate_data['name'] and candidate_data['email'].
            # Example: send_notification_email(candidate_data['email'], stage_name, new_status, full_comment)
            pass

        # Log activity
        # TODO: Add logic to insert a record into candidate_activities table.
        # activity_comment = f"Rejected at {stage_name}. Reason: {rejection_reason}"
        # DataOperations.log_candidate_activity(candidate_id, request.session.get('user_id'), 'REJECTED', activity_comment)

        return JsonResponse({
            'success': True, 
            'message': f'Candidate {candidate_id} successfully rejected. {stage_name} marked as "{new_status}".',
            'rejected_stage': stage_name
        })

    except Exception as e:
        conn.rollback()
        print(f"reject_candidate -> Error: {str(e)}")
        return JsonResponse({'success': False, 'error': 'Database or internal server error during rejection process'}, status=500)
        
    finally:
        DataOperations.close_db_connection(conn, cursor)


# def reject_candidate(request, candidate_id, rejection_reason=None, comment=None, notify_candidate=False):
#     """

#     """
#     # TODO: Implement rejection logic
#     # - Check current candidate status
#     # - Validate rejection reason
#     # - Update candidate status to rejected
#     # - Record rejection reason and comments
#     # - Notify candidate if required

#     return JsonResponse({'success': True, 'message': 'Candidate rejected successfully.'})

@login_required
def put_candidate_on_hold(request, candidate_id: int, hold_reason: str = None, comment: str = None, notify_candidate: bool = False):
    """
    Puts a candidate on hold at their current stage (Screening, L1, L2, or L3).
    The status is updated to 'onHold' at the earliest stage that is NOT 'selected'.
    """
    
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    today = date.today().isoformat()
    
    # Combined comment/reason for the DB remarks column
    full_comment = f"HOLD REASON: {hold_reason.strip()}"
    if comment and comment.strip():
        full_comment += f" | REMARKS: {comment.strip()}"

    try:
        # 1. Determine current status (and fetch data for logging)
        cursor.execute("""
            SELECT 
                screen_status, l1_result, l2_result, l3_result, 
                name, email, jd_id
            FROM candidates
            WHERE candidate_id=%s
        """, (candidate_id,))
        candidate_data = cursor.fetchone()

        if not candidate_data:
            return JsonResponse({'success': False, 'error': 'Candidate data retrieval failed or candidate not found'}, status=404)

        # 2. Validate if status change is possible
        
        # Check 1: Fully Selected (cannot put on hold if they've passed all interviews)
        if candidate_data['l3_result'] == 'selected':
            return JsonResponse({'success': False, 'error': 'Cannot put on hold: Candidate is already fully selected (L3).'}, status=400)

        # Check 2: Already Rejected (rejected is usually final; should not be changed to 'onHold')
        rejection_statuses = ['rejected']
        if any(candidate_data[f'{level}_result'] in rejection_statuses for level in ['l1', 'l2', 'l3']):
            return JsonResponse({'success': False, 'error': 'Cannot put on hold: Candidate is already formally rejected at a previous level.'}, status=400)
        if candidate_data['screen_status'] in rejection_statuses:
             return JsonResponse({'success': False, 'error': 'Cannot put on hold: Candidate is already formally rejected at the screening level.'}, status=400)

        
        # Determine the stage to update (The first stage that is NOT 'selected')
        update_column = None
        date_column = None
        remarks_column = None
        stage_name = ""
        
        # Determine the earliest possible stage that hasn't been successfully passed
        if candidate_data['screen_status'] != 'selected':
            update_column = 'screen_status'
            date_column = 'screened_on'
            remarks_column = 'screened_remarks'
            stage_name = "Screening"
        elif candidate_data['l1_result'] != 'selected':
            update_column = 'l1_result'
            date_column = 'l1_date'
            remarks_column = 'l1_comments'
            stage_name = "L1 Interview"
        elif candidate_data['l2_result'] != 'selected':
            update_column = 'l2_result'
            date_column = 'l2_date'
            remarks_column = 'l2_comments'
            stage_name = "L2 Interview"
        elif candidate_data['l3_result'] != 'selected':
            update_column = 'l3_result'
            date_column = 'l3_date'
            remarks_column = 'l3_comments'
            stage_name = "L3 Interview"
        
        if not update_column:
             return JsonResponse({'success': False, 'error': 'Candidate status is ambiguous or fully selected.'}, status=400)


        # 4. Validate the hold reason (must not be blank)
        if not hold_reason or not hold_reason.strip():
            return JsonResponse({'success': False, 'error': 'Hold reason must be provided to put the candidate on hold.'}, status=400)

        
        # 5. Update candidate status, date, and remarks
        new_status = 'onHold'
        
        update_query = f"""
            UPDATE candidates 
            SET 
                {update_column}=%s, 
                {date_column}=%s, 
                {remarks_column}=%s,
                updated_at=CURRENT_TIMESTAMP
            WHERE candidate_id=%s
        """
        update_params = [new_status, today, full_comment, candidate_id]
        
        cursor.execute(update_query, tuple(update_params))
        conn.commit()

        # 6. Notify candidate if required (TODO)
        if notify_candidate:
            # TODO: Implement robust email notification logic here using
            # candidate_data['name'] and candidate_data['email'].
            # Example: send_notification_email(candidate_data['email'], stage_name, new_status, full_comment)
            pass

        # Log activity
        # TODO: Add logic to insert a record into candidate_activities table.
        # activity_comment = f"Put on hold at {stage_name}. Reason: {hold_reason}"
        # DataOperations.log_candidate_activity(candidate_id, request.session.get('user_id'), 'ON_HOLD', activity_comment)

        return JsonResponse({
            'success': True, 
            'message': f'Candidate {candidate_id} successfully put on hold. {stage_name} marked as "{new_status}".',
            'held_stage': stage_name
        })

    except Exception as e:
        conn.rollback()
        print(f"put_candidate_on_hold -> Error: {str(e)}")
        return JsonResponse({'success': False, 'error': 'Database or internal server error during hold process'}, status=500)
        
    finally:
        DataOperations.close_db_connection(conn, cursor)


# def put_candidate_on_hold(request, candidate_id, hold_reason=None, comment=None, notify_candidate=False):
    """

    """
    # TODO: Implement putting candidate on hold logic
    # - Check current candidate status
    # - Validate hold reason
    # - Update candidate status to on hold
    # - Record hold reason and comments
    # - Notify candidate if required

    return JsonResponse({'success': True, 'message': 'Candidate put on hold successfully.'})

@login_required
def back_to_previous_stage(request, candidate_id: int, comment: str):
    """
    Reverts the candidate's last recorded stage (anything other than 'toBeScreened') 
    back to 'toBeScreened'. Useful for re-activating rejected or on-hold candidates.
    """
    
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    user_id = request.session.get('user_id', 'SYSTEM')

    try:
        # 1. Fetch current statuses and recruiter comments
        cursor.execute("""
            SELECT 
                screen_status, l1_result, l2_result, l3_result, recruiter_comments
            FROM candidates
            WHERE candidate_id=%s
        """, (candidate_id,))
        candidate_data = cursor.fetchone()

        if not candidate_data:
            return JsonResponse({'success': False, 'error': 'Candidate not found'}, status=404)

        # 2. Determine the LAST stage that is NOT 'toBeScreened'
        update_column = None
        date_column = None
        comments_column = None
        stage_name = ""
        
        stages = [
            ('l3', 'l3_result', 'l3_date', 'l3_comments', "L3 Interview"),
            ('l2', 'l2_result', 'l2_date', 'l2_comments', "L2 Interview"),
            ('l1', 'l1_result', 'l1_date', 'l1_comments', "L1 Interview"),
            ('screen', 'screen_status', 'screened_on', 'screened_remarks', "Screening"),
        ]
        
        # Iterate in reverse order (L3 -> Screening) to find the LAST successful/recorded status
        for key, status_col, date_col, comments_col, name in stages:
            status_value = candidate_data.get(status_col)
            if status_value and status_value != 'toBeScreened':
                update_column = status_col
                date_column = date_col
                comments_column = comments_col
                stage_name = name
                # Found the last active stage, break the loop
                break
        
        if not update_column:
            return JsonResponse({'success': False, 'error': 'Cannot revert: Candidate is already at the initial "toBeScreened" state or status is ambiguous.'}, status=400)

        
        # Prepare the update
        new_status = 'toBeScreened'
        
        # Append the reason for rollback to the general recruiter_comments field
        new_recruiter_comment = f"\n[ROLLBACK on {date.today().isoformat()} by {user_id}]: Resetting {stage_name} from '{candidate_data.get(update_column)}' to '{new_status}'. Reason: {comment.strip()}"
        
        # 3. Update candidate status (resetting the status, date, and comments for that specific stage)
        update_query = f"""
            UPDATE candidates 
            SET 
                {update_column}=%s, 
                {date_column}=NULL, 
                {comments_column}=NULL,
                recruiter_comments=CONCAT(COALESCE(recruiter_comments, ''), %s),
                updated_at=CURRENT_TIMESTAMP
            WHERE candidate_id=%s
        """
        update_params = [new_status, new_recruiter_comment, candidate_id]
        
        cursor.execute(update_query, tuple(update_params))
        conn.commit()

        # We are explicitly not notifying the candidate as per instruction.

        # Log activity
        # TODO: Add logic to insert a record into candidate_activities table.
        # activity_comment = f"Reverted {stage_name} status from '{candidate_data.get(update_column)}' to '{new_status}'. Reason: {comment}"
        # DataOperations.log_candidate_activity(candidate_id, user_id, 'REVERT', activity_comment)

        return JsonResponse({
            'success': True, 
            'message': f'Candidate {candidate_id} successfully reverted. {stage_name} reset to "{new_status}".',
            'reverted_stage': stage_name
        })

    except Exception as e:
        conn.rollback()
        print(f"back_to_previous_stage -> Error: {str(e)}")
        return JsonResponse({'success': False, 'error': 'Database or internal server error during stage reversion process'}, status=500)
        
    finally:
        DataOperations.close_db_connection(conn, cursor)

@login_required
def schedule_interview_for_candidate(request, candidate_id, interview_data, **kwargs):
    """
    Schedule an interview for a specific candidate with provided interview details.
    """
    # Extract interview details from the provided data
    # TODO: Implement interview scheduling logic
    # - Validate interview data
    # - Save interview details to database
    # - Send email notifications if required
    # - Update candidate status
    return JsonResponse({'success': True, 'message': 'Interview scheduled successfully.'})

@login_required
def send_offer_to_candidate(request, candidate_id):
    """
    Send offer to a candidate.
    """
    # TODO: Implement offer sending logic
    # - Generate offer letter
    # - Update candidate offer status
    # - Send offer email to candidate
    # - Notify relevant team members
    return JsonResponse({'success': True, 'message': 'Offer sent successfully.'})

@login_required
def withdraw_candidate_offer(request, candidate_id: int, comment: str = None, notify_candidate: bool = False):
    """
    Withdraw an existing offer for a candidate.
    Sets offer_status to 'declined' and records the reason in recruiter_comments.
    
    :param notify_candidate: If True, triggers a notification to the candidate.
    """
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    user_id = request.session.get('user_id', 'SYSTEM')

    # Construct the final comment for recruiter_comments
    final_comment = f"\n[OFFER WITHDRAWAL on {date.today().isoformat()} by {user_id}]: Cause: Offer Withdrawn by us | Remarks: {comment.strip()}" if comment and comment.strip() else f"\n[OFFER WITHDRAWAL on {date.today().isoformat()} by {user_id}]: Cause: Offer Withdrawn by us | Remarks: No additional comments"

    try:
        # 1. Fetch current status and validate if the candidate is in the offering stage
        cursor.execute("""
            SELECT 
                l3_result, offer_status, name, email
            FROM candidates
            WHERE candidate_id=%s
        """, (candidate_id,))
        candidate_data = cursor.fetchone()

        if not candidate_data:
            return JsonResponse({'success': False, 'error': 'Candidate not found'}, status=404)

        # Basic validation: Check if they passed L3 (were 'selected')
        if candidate_data['l3_result'] != 'selected':
             return JsonResponse({'success': False, 'error': f'Cannot withdraw offer: Candidate has not successfully passed L3 (current L3 status: {candidate_data["l3_result"]}).'}, status=400)
             
        # Optional validation: Check if offer status is already set to declined/withdrawn
        # Note: 'declined' is used for both candidate-declined and recruiter-withdrawn offers
        if candidate_data['offer_status'] in ['declined', 'hired']:
             return JsonResponse({'success': False, 'error': f'Offer is already in a final state: {candidate_data["offer_status"]}.'}, status=400)

        # 2. Update candidate status
        new_offer_status = 'declined'
        
        update_query = """
            UPDATE candidates 
            SET 
                offer_status=%s, 
                recruiter_comments=CONCAT(COALESCE(recruiter_comments, ''), %s),
                updated_at=CURRENT_TIMESTAMP
            WHERE candidate_id=%s
        """
        update_params = [new_offer_status, final_comment, candidate_id]
        
        cursor.execute(update_query, tuple(update_params))
        conn.commit()

        # 3. Handle notification
        if notify_candidate:
            # TODO: Implement robust email notification logic here using
            # candidate_data['name'] and candidate_data['email'].
            # Example: send_withdrawal_email(candidate_data['email'])
            pass
        
        # Log activity
        # TODO: Add logic to insert a record into candidate_activities table.
        # activity_comment = f"Offer withdrawn by recruiter. Reason: {comment}"
        # DataOperations.log_candidate_activity(candidate_id, user_id, 'OFFER_WITHDRAWN', activity_comment)

        return JsonResponse({
            'success': True, 
            'message': f'Offer for Candidate {candidate_id} successfully withdrawn.',
            'new_offer_status': new_offer_status
        })

    except Exception as e:
        conn.rollback()
        print(f"withdraw_candidate_offer -> Error: {str(e)}")
        return JsonResponse({'success': False, 'error': 'Database or internal server error during offer withdrawal process'}, status=500)
        
    finally:
        DataOperations.close_db_connection(conn, cursor)

@login_required
def mark_candidate_as_hired(request, candidate_id):
    """
    Mark a candidate as hired.
    """
    # TODO: Implement hired marking logic
    # - Update candidate status to hired
    # - Move candidate to hired candidates table
    # - Send welcome email
    # - Notify HR and relevant teams
    return JsonResponse({'success': True, 'message': 'Candidate marked as hired successfully.'})

@login_required
def mark_candidate_as_resigned(request, candidate_id):
    """
    Mark a candidate as resigned.
    """
    # TODO: Implement resignation marking logic
    # - Update candidate status to resigned
    # - Record resignation date and reason
    # - Notify relevant teams
    # - Update reporting systems
    return JsonResponse({'success': True, 'message': 'Candidate marked as resigned successfully.'})

def schedule_interviews_page(request):
    """
    View to render the schedule interviews page.
    """
    return render(request, 'schedule_interviews.html')

def get_candidates_for_jd(request):
    """
    API endpoint to get candidates for a specific job description.
    """
    print("get_candidates_for_jd -> Request method:", request.method)
    jd_id = request.GET.get('jd_id')
    if not jd_id:
        return JsonResponse({'success': False, 'error': 'JD ID required'}, status=400)

    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        # First check if there are any candidates at all for this JD
        cursor.execute("SELECT COUNT(*) as total_count FROM candidates WHERE jd_id = %s", (jd_id,))
        total_count = cursor.fetchone()['total_count']

        # Then check how many are selected
        cursor.execute("SELECT COUNT(*) as selected_count FROM candidates WHERE jd_id = %s AND screen_status = 'selected'", (jd_id,))
        selected_count = cursor.fetchone()['selected_count']

        # Now get only the selected candidates
        cursor.execute("""
            SELECT c.*, r.status as resume_status
            FROM candidates c
            JOIN resumes r ON c.resume_id = r.resume_id
            WHERE c.jd_id = %s AND c.screen_status = 'selected'
            ORDER BY c.name
        """, (jd_id,))
        candidates = cursor.fetchall()

        # Prepare meaningful message based on counts
        message = ""
        if total_count == 0:
            message = "No selected candidates found for this JD."
        elif selected_count == 0:
            message = f"Found {total_count} candidates for this JD, but none have been Screened OK for interviews yet."

        return JsonResponse({
            'success': True,
            'candidates': candidates,
            'total_count': total_count,
            'selected_count': selected_count,
            'message': message
        })
    except Exception as e:
        print(f"get_candidates_for_jd -> Error: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    finally:
        cursor.close()
        conn.close()

@csrf_exempt
def schedule_interview(request):
    """
    API endpoint to schedule an interview.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)
        
    session = request.session
    # get user_id and email from session
    user_email = session.get('email')

    if not user_email:
        return redirect('login')

    try:
        data = json.loads(request.body)
        candidate_id = data.get('candidate_id')
        level = data.get('level')
        date = data.get('date')
        time = data.get('time')
        interviewer_name = data.get('interviewer_name')
        interviewer_email = data.get('interviewer_email')

        # Validate required fields
        if not all([candidate_id, level, date, time, interviewer_name, interviewer_email]):
            return JsonResponse({'success': False, 'error': 'All fields are required'}, status=400)

        # Get candidate info for calendar invite
        conn = DataOperations.get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT 
                       c.candidate_id, c.name, 
                       c.email, c.jd_id, 
                       r.jd_summary
            FROM candidates c
            JOIN recruitment_jds r ON c.jd_id = r.jd_id
            WHERE c.candidate_id = %s
        """, (candidate_id,))
        candidate = cursor.fetchone()

        if not candidate:
            return JsonResponse({'success': False, 'error': 'Candidate not found'}, status=404)
        

        # Update candidate record with interview details
        cursor.execute(f"""
            UPDATE candidates 
            SET 
                {level}_date = %s,
                {level}_result = 'toBeScreened',
                {level}_interviewer_name = %s,
                {level}_interviewer_email = %s,
                updated_at = NOW()
            WHERE candidate_id = %s
        """, (date, interviewer_name, interviewer_email, candidate_id))

        conn.commit()
        user = request.user
        token = f"{candidate_id}-{level}-{int(datetime.now().timestamp())}"  # Simple token
        if send_interview_result_email(
            hr_email=user_email, # if request.user.is_authenticated else 'hr@yourdomain.com',
            interviewer_email=interviewer_email,
            candidate_id = candidate_id,
            interviewer_name=interviewer_name,
            candidate=candidate,
            level=level,
            token=token
        ):
            
            # TO DO: If necessary, need to send notification to team leads.
            team_lead_user_id = DataOperations.get_team_lead_user_id_from_team_id(candidate['team_id'])
            if team_lead_user_id and DataOperations.get_user_settings(team_lead_user_id).get('notifications_enabled', False):
                MessageProviders.send_notification(team_lead_user_id, "Interview Scheduled", f"Interview for {candidate['name']} ({level.upper()}) has been scheduled.", created_by=user_email, notification_type="Candidate")

            return JsonResponse({
                'success': True,
                'message': 'Interview scheduled successfully'
            })
        else:
            return JsonResponse({'success': False, 'error': 'Failed to send interview result email'}, status=500)

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'conn' in locals():
            conn.close()

def send_interview_result_email(hr_email, interviewer_email, candidate_id, interviewer_name, candidate, level, token):
    """
    Send email to interviewer with candidate details and interview result link.
    """
    print("send_interview_result_email -> Sending email to interviewer:", interviewer_email)
    print("send_interview_result_email -> Candidate details:", candidate, "candidiate id:", candidate_id)
    print("HR email:", hr_email)
    base_url = "http://127.0.0.1:8000"
    subject = f"Action Required: Record Interview Result for {candidate['name']} ({level.upper()})"
    result_url = f"{base_url}/record_interview_result/?candidate_id={candidate['candidate_id']}&level={level}&token={token}"
    html_content = render_to_string('interview_result_request.html', {
        'interviewer_name': interviewer_name,
        'candidate': candidate,
        'level': level.upper(),
        'result_url': result_url,
        'hr_email': hr_email,
    })

    # Fetch email config for hr_email
    from .utils import decrypt_password
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT email, email_smtp_host, email_smtp_port, email_host_password FROM email_config WHERE email=%s", (hr_email,))
    row = cursor.fetchone()
    cursor.close()
    conn.close()
    if not row:
        print(f"send_interview_result_email -> No email config found for {hr_email}")
        return False
    app_password = decrypt_password(row['email_host_password'])
    from_email = row['email']
    smtp_host = row['email_smtp_host']
    smtp_port = row['email_smtp_port']

    # Use send_email utility
    result = MessageProviders.send_email(
        from_email=from_email,
        app_password=app_password,
        to_email=interviewer_email,
        subject=subject,
        html_body=html_content,
        smtp_host=smtp_host,
        smtp_port=smtp_port
    )

    if result:
        print("send_interview_result_email -> Email sent successfully")
    else:
        print("send_interview_result_email -> Failed to send email")
    return result

def record_interview_result_page(request):
    """
    View to render the record interview result page.
    """
    candidate_id = request.GET.get('candidate_id')
    level = request.GET.get('level')
    token = request.GET.get('token')
    # Optionally validate token here
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM candidates WHERE candidate_id=%s", (candidate_id,))
    candidate = cursor.fetchone()
    cursor.close()
    conn.close()
    if not candidate:
        return render(request, 'record_interview_result.html', {'error': 'Candidate not found'})
    return render(request, 'record_interview_result.html', {
        'candidate': candidate,
        'level': level,
        'token': token
    })


@csrf_exempt
def submit_interview_result(request):
    """
    API endpoint to submit interview results.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)

    data = json.loads(request.body)
    candidate_id = data.get('candidate_id')
    level = data.get('level')
    result = data.get('result')
    comments = data.get('comments')
    token = data.get('token')

    if not all([candidate_id, level, result]):
        return JsonResponse({'success': False, 'error': 'Missing fields'}, status=400)

    conn = None
    cursor = None
    try:
        conn = DataOperations.get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Step 1: Get the current (previous) data of the candidate
        cursor.execute("SELECT jd_id, screen_status, l1_result, l2_result, l3_result FROM candidates WHERE candidate_id = %s", (candidate_id,))
        previous_candidate_data = cursor.fetchone()
        
        if not previous_candidate_data:
            return JsonResponse({'success': False, 'error': 'Candidate not found'}, status=404)

        # Step 2: Update the candidate's interview result
        cursor.execute(f"""
            UPDATE candidates
            SET {level}_result = %s, {level}_comments = %s, updated_at = NOW()
            WHERE candidate_id = %s
        """, (result, comments, candidate_id))
        

        # Step 3: Get the new (current) data of the candidate
        current_candidate_data = {
            "jd_id": previous_candidate_data['jd_id'],
            "screen_status": previous_candidate_data['screen_status'],
            "l1_result": previous_candidate_data['l1_result'] if level != 'l1' else result,
            "l2_result": previous_candidate_data['l2_result'] if level != 'l2' else result,
            "l3_result": previous_candidate_data['l3_result'] if level != 'l3' else result,
        }
        check = None
        if current_candidate_data:
            # Step 4: Call the function to update recruitment_jds based on the change
            check = DataOperations.update_recruitment_jds(cursor, previous_candidate_data, current_candidate_data)
            if not check:
                conn.rollback()
                return JsonResponse({'success': False, 'error': 'Error updating recruitment_jds counts'}, status=500)
        if check:
            conn.commit()
        # TO DO: If necessary, need to send notification to team leads.
        cursor.execute("SELECT team_id FROM candidates WHERE candidate_id=%s", (candidate_id,))
        team_row = cursor.fetchone()
        if team_row:
            team_id = team_row['team_id']
            team_lead_user_id = DataOperations.get_team_lead_user_id_from_team_id(team_id)
            if team_lead_user_id and DataOperations.get_user_settings(team_lead_user_id).get('notifications_enabled', False):
                MessageProviders.send_notification(team_lead_user_id, "Interview Result Submitted", f"Interview result for candidate ID {candidate_id} ({level.upper()}) has been submitted.", notification_type="Candidate")

    except Exception as e:
        if conn:
            conn.rollback() # Rollback in case of an error
        print(f"Error in submit_interview_result: {e}")
        return JsonResponse({'success': False, 'error': 'Internal Server Error'}, status=500)
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

    return JsonResponse({'success': True})


def manage_candidate_status_page(request):
    """
    View to render the manage candidate status page.
    """
    return render(request, "manage_candidate_status.html")

def manage_candidate_status_data(request):
    """
    API endpoint to get candidate status data.
    """

    user_role = request.session.get('role', 'Guest')
    user_id = request.session.get('user_id', None)
    search = request.GET.get("search", "")
    page = int(request.GET.get("page", 1))
    limit = 10
    offset = (page - 1) * limit
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Build base query and params
    params = []
    if user_role == 'Admin':
        query = """
            SELECT c.*, jd.jd_summary, jd.jd_id
            FROM candidates c
            LEFT JOIN recruitment_jds jd ON c.jd_id = jd.jd_id
            WHERE c.screen_status = 'selected'
        """
        count_query = "SELECT COUNT(*) as total FROM candidates c LEFT JOIN recruitment_jds jd ON c.jd_id = jd.jd_id WHERE c.screen_status = 'selected'"
        if search:
            query += " AND (c.name LIKE %s OR c.email LIKE %s OR jd.jd_summary LIKE %s)"
            count_query += " AND (c.name LIKE %s OR c.email LIKE %s OR jd.jd_summary LIKE %s)"
            params = [f"%{search}%", f"%{search}%", f"%{search}%"]
    else:
        # Team_Lead or User: restrict to JDs associated with their teams
        # Get emp_id for user
        emp_id = None
        if request.session.get('email'):
            cursor.execute("SELECT emp_id FROM hr_team_members WHERE email=%s", (request.session.get('email'),))
            emp_row = cursor.fetchone()
            emp_id = emp_row['emp_id'] if emp_row else None
        team_ids = []
        if emp_id:
            cursor.execute("SELECT team_id FROM team_members WHERE emp_id=%s", (emp_id,))
            team_ids = [row['team_id'] for row in cursor.fetchall()]
        if team_ids:
            team_ids_str = ','.join(str(tid) for tid in team_ids)
            query = f"""
                SELECT c.*, jd.jd_summary, jd.jd_id
                FROM candidates c
                LEFT JOIN recruitment_jds jd ON c.jd_id = jd.jd_id
                WHERE c.screen_status = 'selected' AND jd.team_id IN ({team_ids_str})
            """
            count_query = f"SELECT COUNT(*) as total FROM candidates c LEFT JOIN recruitment_jds jd ON c.jd_id = jd.jd_id WHERE c.screen_status = 'selected' AND jd.team_id IN ({team_ids_str})"
            if search:
                query += " AND (c.name LIKE %s OR c.email LIKE %s OR jd.jd_summary LIKE %s)"
                count_query += " AND (c.name LIKE %s OR c.email LIKE %s OR jd.jd_summary LIKE %s)"
                params = [f"%{search}%", f"%{search}%", f"%{search}%"]
        else:
            # No teams, return empty
            cursor.close()
            conn.close()
            return JsonResponse({"candidates": [], "page": page, "num_pages": 0})

    query += " ORDER BY c.updated_at DESC LIMIT %s OFFSET %s"
    params += [limit, offset]
    cursor.execute(query, params)
    candidates = cursor.fetchall()
    # For count, only use search params if present
    count_params = params[:3] if search else []
    cursor.execute(count_query, count_params)
    total = cursor.fetchone()['total']
    num_pages = (total // limit) + (1 if total % limit else 0)
    cursor.close()
    conn.close()
    return JsonResponse({
        "candidates": candidates,
        "page": page,
        "num_pages": num_pages
    })

@csrf_exempt
def update_candidate_status(request):
    """
    API endpoint to update the status of a candidate.
    """
    print("update_candidate_status -> Request method:", request.method)
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Invalid method"}, status=405)
    data = json.loads(request.body)
    candidate_id = data.get("candidate_id")
    l1_result = data.get("l1_result")
    l2_result = data.get("l2_result")
    l3_result = data.get("l3_result")
    if not candidate_id:
        return JsonResponse({"success": False, "error": "Candidate ID required"}, status=400)
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("""
            SELECT jd_id, screen_status, l1_result, l2_result, l3_result
            FROM candidates
            WHERE candidate_id=%s
        """, [candidate_id])
        previous_candidate_data = cursor.fetchone()
        if not previous_candidate_data:
            return JsonResponse({"success": False, "error": "Candidate not found"}, status=404)
    
        cursor.execute("""
            UPDATE candidates
            SET l1_result=%s, l2_result=%s, l3_result=%s, updated_at=NOW()
            WHERE candidate_id=%s
        """, [l1_result, l2_result, l3_result, candidate_id])

        candidata_new_data = {
            "jd_id": previous_candidate_data['jd_id'],
            "screen_status": previous_candidate_data['screen_status'],
            "l1_result": l1_result,
            "l2_result": l2_result,
            "l3_result": l3_result
        }
        check = DataOperations.update_recruitment_jds(cursor, previous_candidate_data, candidata_new_data)
        if not check:
            conn.rollback()
            return JsonResponse({"success": False, "error": "Error updating recruitment_jds counts after candidate status update"}, status=500)
        conn.commit()

        # TO DO: Send notification to team lead if any candidate is finallized.(means, selected for all levels)
        cursor.execute("SELECT name, team_id, hr_member_id FROM candidates WHERE candidate_id=%s", [candidate_id])
        candidate_data = cursor.fetchone()
        print("Candidate data for notification:", candidate_data)
        if candidate_data:
            team_id = candidate_data['team_id']
            hr_member_id = candidate_data['hr_member_id']
            hr_user_id = DataOperations.get_user_id_from_emp_id(hr_member_id)
            if hr_user_id and DataOperations.get_user_settings(hr_user_id).get('notifications_enabled', False):
                MessageProviders.send_notification(
                    user_id=hr_user_id,
                    title="Candidate Status Updated",
                    message=f"Candidate {candidate_data['name']}'s status has been updated."
                )
            cursor.execute("SELECT lead_emp_id FROM teams WHERE team_id=%s", [team_id])
            team_row = cursor.fetchone()
            if team_row:
                lead_emp_id = team_row['lead_emp_id']
                lead_user_id = DataOperations.get_user_id_from_emp_id(lead_emp_id)
                if lead_user_id and lead_user_id != hr_user_id and DataOperations.get_user_settings(lead_user_id).get('notifications_enabled', False):
                    MessageProviders.send_notification(
                        user_id=lead_user_id,
                        title="Candidate Status Updated",
                        message=f"Candidate {candidate_data['name']}'s status has been updated.",
                        notification_type="status_update"
                    )
        return JsonResponse({"success": True})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)
    finally:
        cursor.close()
        conn.close()


def view_finalized_candidates(request):
    """
    View to render the finalized candidates page.
    """
    print("view_finalized_candidates -> Request method:", request.method)
    name = request.session.get('name', 'Guest')
    return render(request, 'view_finalized_candidates.html', {'name': name})

def api_jds(request):
    """
    API endpoint to get job descriptions.
    """
    user_id = request.session.get('user_id', None)
    user_role = request.session.get('role', 'Guest')

    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    if user_role in ['Team_Lead', 'User']:
        # Get emp_id for user
        cursor.execute("""
            SELECT emp_id FROM hr_team_members
                    WHERE email=(SELECT email from users WHERE user_id=%s)
                    LIMIT 1;
        """, [user_id])
        row = cursor.fetchone()
        emp_id=row['emp_id']

        cursor.execute("""
            SELECT DISTINCT j.jd_id, j.jd_summary
            FROM recruitment_jds j
            LEFT JOIN teams t ON j.team_id = t.team_id
            LEFT JOIN team_members tm ON t.team_id = tm.team_id
            WHERE t.lead_emp_id=%s OR tm.emp_id=%s
            AND j.jd_status='active'
            ORDER BY j.jd_summary
        """, [emp_id, emp_id])
    else:
        cursor.execute("SELECT jd_id, jd_summary FROM recruitment_jds WHERE jd_status='active'")
    jds = cursor.fetchall()
    cursor.close()
    conn.close()
    return JsonResponse({'jds': jds})

def api_finalized_candidates(request):
    """
    API endpoint to get finalized candidates for a specific job description.
    """
    jd_id = request.GET.get('jd_id')
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT candidate_id, name, email, phone, experience
        FROM candidates
        WHERE jd_id=%s AND l3_result='selected'
        ORDER BY updated_at DESC
    """, (jd_id,))
    candidates = cursor.fetchall()
    cursor.close()
    conn.close()
    return JsonResponse({'candidates': candidates})

def api_candidate_details(request):
    """
    API endpoint to get details of a specific candidate.
    """
    candidate_id = request.GET.get('candidate_id')
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM candidates WHERE candidate_id=%s", (candidate_id,))
    candidate = cursor.fetchone()
    cursor.close()
    conn.close()
    return JsonResponse({'details': candidate})


@login_required
def logout_page(request):
    """
    View to render the logout page.
    """
    session_id = request.session.get('session_id')
    if session_id:
        conn = DataOperations.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM user_sessions WHERE session_id=%s", (session_id,))
        conn.commit()
        cursor.close()
        conn.close()
    logout(request)  # This logs out the user
    return render(request, 'logout.html')


def candidate_profile(request):
    """
    View to render the candidate profile page.
    """
    name = request.session.get('name', 'Guest')
    user_role_ = request.session.get('role', 'Guest')
    """Render the Candidate Profile page."""
    return render(request, 'candidate_profile.html', {'name': name, 'user_role': user_role_})


@login_required
def get_members_from_team_id(request, team_id):

    if request.method != "GET":
        return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)
    if not team_id:
        return JsonResponse({'success': False, 'error': 'Team ID required'}, status=400)
    if team_id != 'all' and not team_id.isdigit():
        return JsonResponse({'success': False, 'error': 'Invalid Team ID'}, status=400)

    name = request.session.get('name', 'Guest')
    user_role = request.session.get('role', 'Guest')
    user_id = request.session.get("user_id", None)

    emp_id = DataOperations.get_emp_id_from_user_id(user_id)
    if not emp_id:
        return JsonResponse({'success': False, 'error': 'Employee ID not found for user'}, status=404)

    # members will have data as [{emp_id, first_name, last_name, email}, ...]
    members = []

    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    query = """
            SELECT DISTINCT hr.emp_id, hr.first_name, hr.last_name, hr.email
            FROM hr_team_members hr
            INNER JOIN team_members tm ON tm.emp_id=hr.emp_id
        """
    filter = ""
    params = [emp_id]

    # Build filter based on role and team_id
    if team_id != 'all':
        if user_role == 'Admin':
            filter = "WHERE tm.team_id=%s"
            params = (team_id,)
        elif user_role == 'Team_Lead':
            # if user is team lead of team provided, select all members
            leading_teams = DataOperations.is_user_team_lead(user_id)
            if leading_teams and int(team_id) in leading_teams:
                filter = " AND tm.team_id=%s"
                params.append(team_id)
            else:
                # if not team lead of the team provided, select only self if he is part of any team member
                filter = " AND hr.emp_id=%s AND hr.emp_id IN (SELECT emp_id FROM team_members)"
                params.append(emp_id)
        elif user_role == 'User':
            # select only self if he is part of any team member
            filter = " AND hr.emp_id=%s AND hr.emp_id IN (SELECT emp_id FROM team_members)"
            params.append(emp_id)

    # Execute query based on role
    if user_role == 'Admin':
        params = params if filter else []
        cursor.execute(f"""
            {query}
            {filter}
            ORDER BY hr.first_name, hr.last_name;
        """, params)
    elif user_role == 'Team_Lead':
        # check if user is team lead of any team
        leading_teams = DataOperations.is_user_team_lead(user_id)
        if leading_teams:
            cursor.execute(f"""
                {query}
                INNER JOIN teams t ON t.team_id=tm.team_id
                WHERE t.lead_emp_id=%s
                {filter if filter else 'AND hr.emp_id IN (SELECT emp_id FROM team_members)'}
                ;
            """, params)
        else:
            # if not team lead of any team, select only self if he is part of any team member
            cursor.execute(f"""
                {query}
                WHERE hr.emp_id=%s
                {filter if filter else 'AND hr.emp_id IN (SELECT emp_id FROM team_members)'}
            """, params)
    elif user_role == 'User':
        # select only self if he is part of any team member
        cursor.execute(f"""
            {query}
            WHERE hr.emp_id=%s
            {filter if filter else 'AND hr.emp_id IN (SELECT emp_id FROM team_members)'}
        """, params)
    members = cursor.fetchall()
    cursor.close()
    conn.close()
    return JsonResponse({'success': True, 'members': members,})


def candidate_suggestions_for_muster(request, query):
    """
    API endpoint to get candidate suggestions for muster page based on a query.
    """
    if request.method != "GET":
        return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)
    query = query.strip()
    if not query:
        return JsonResponse({'success': False, 'error': 'Query parameter is required'}, status=400)

    filter_data = {
        'team_id': request.GET.get('team_id', 'all'),
        'jd_id': request.GET.get('jd_id', 'all'),
        'member_id': request.GET.get('member_id', 'all'),
    }

    # Where clauses
    where_clauses = []
    params = []
    if filter_data['team_id'] != 'all' and filter_data['team_id'].isdigit():
        where_clauses.append("c.team_id=%s")
        params.append(filter_data['team_id'])
    if filter_data['jd_id'] != 'all' and filter_data['jd_id'].isdigit():
        where_clauses.append("c.jd_id=%s")
        params.append(filter_data['jd_id'])
    if filter_data['member_id'] != 'all' and filter_data['member_id'].isdigit():
        where_clauses.append("c.hr_member_id=%s")
        params.append(filter_data['member_id'])
    
    if where_clauses:
        where_statement = " AND " + " AND ".join(where_clauses)
    else:
        where_statement = ""

    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute(f"""
        SELECT candidate_id, name, email
        FROM candidates
        WHERE (name LIKE %s OR email LIKE %s OR phone LIKE %s) {where_statement}
        ORDER BY name
        LIMIT 5
    """, (f"%{query}%", f"%{query}%", f"%{query}%", *params))
    candidates = cursor.fetchall()
    cursor.close()
    conn.close()
    return JsonResponse({'success': True, 'candidates': candidates})

@login_required
def get_search_candidates(request):
    if request.method != "GET":
        return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)
    query = request.GET.get('query', '').strip()

    filters = {
        'team_id': request.GET.get('team_id', 'all'),
        'jd_id': request.GET.get('jd_id', 'all'),
        'member_id': request.GET.get('member_id', 'all'),
        'from_date': request.GET.get('from_date', ''),
        'to_date': request.GET.get('to_date', ''),
        'search_query': query, 
    }

    from .utils import DataValidators

    # Only validate dates if they are provided, don't set defaults
    if filters['from_date'] and not DataValidators.is_valid_date(filters['from_date']):
        return JsonResponse({'success': False, 'error': 'Invalid from_date format. Use YYYY-MM-DD.'}, status=400)
    
    if filters['to_date'] and not DataValidators.is_valid_date(filters['to_date']):
        return JsonResponse({'success': False, 'error': 'Invalid to_date format. Use YYYY-MM-DD.'}, status=400)

    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    role = request.session.get('role', 'Guest')
    user_id = request.session.get('user_id', None)
    emp_id = DataOperations.get_emp_id_from_user_id(user_id)

    if not emp_id:
        return JsonResponse({'success': False, 'error': 'Employee ID not found for user'}, status=404)

    # Build where clauses based on filters and role
    where_clauses = []
    params = []

    # if team_id is all, get all team_ids depending upon role of user
    all_teams = []
    if role == 'Admin':
        cursor.execute("SELECT team_id FROM teams")
    elif role == 'Team_Lead':
        cursor.execute("SELECT team_id FROM teams WHERE lead_emp_id=%s UNION SELECT team_id FROM team_members WHERE emp_id=%s", (emp_id, emp_id))
    elif role == 'User':
        cursor.execute("SELECT team_id FROM team_members WHERE emp_id=%s", (emp_id,))

    all_teams = [row['team_id'] for row in cursor.fetchall()]

    all_jds = []
    if role == 'Admin':
        cursor.execute("SELECT jd_id FROM recruitment_jds WHERE jd_status='active'")
    elif role in ['Team_Lead', 'User']:
        if all_teams:
            format_strings = ','.join(['%s'] * len(all_teams))
            cursor.execute(f"SELECT jd_id FROM recruitment_jds WHERE jd_status='active' AND team_id IN ({format_strings})", tuple(all_teams))
    all_jds = [row['jd_id'] for row in cursor.fetchall()]

    all_members = []
    if role == 'Admin':
        cursor.execute("SELECT emp_id FROM hr_team_members WHERE emp_id IN (SELECT emp_id FROM team_members)")
    elif role in ['Team_Lead']:
        if all_teams:
            format_strings = ','.join(['%s'] * len(all_teams))
            cursor.execute(f"""
                SELECT DISTINCT hm.emp_id
                FROM hr_team_members hm
                JOIN team_members tm ON hm.emp_id = tm.emp_id
                WHERE tm.team_id IN ({format_strings})
            """, tuple(all_teams))

    all_members = [row['emp_id'] for row in cursor.fetchall()]
    all_members = all_members if all_members else [emp_id]  # Ensure at least self is included for User role


    if filters['team_id'] != 'all':
        if filters['team_id'].isdigit() and int(filters['team_id']) in all_teams:
            where_clauses.append("c.team_id=%s")
            params.append(filters['team_id'])
        else:
            return JsonResponse({'success': True, 'candidates': [], 'page': 1, 'num_pages': 0})
    else:
        if all_teams:
            format_strings = ','.join(['%s'] * len(all_teams))
            where_clauses.append(f"c.team_id IN ({format_strings})")
            params.extend(all_teams)
        else:
            return JsonResponse({'success': True, 'candidates': [], 'page': 1, 'num_pages': 0})
    # get candidate_data

    if filters['jd_id'] != 'all':
        if filters['jd_id'].isdigit() and int(filters['jd_id']) in all_jds:
            where_clauses.append("c.jd_id=%s")
            params.append(filters['jd_id'])
        else:
            return JsonResponse({'success': True, 'candidates': [], 'page': 1, 'num_pages': 0})
    else:
        if all_jds:
            format_strings = ','.join(['%s'] * len(all_jds))
            where_clauses.append(f"c.jd_id IN ({format_strings})")
            params.extend(all_jds)
        else:
            return JsonResponse({'success': True, 'candidates': [], 'page': 1, 'num_pages': 0})

    if filters['member_id'] != 'all':
        if filters['member_id'].isdigit() and int(filters['member_id']) in all_members:
            where_clauses.append("c.hr_member_id=%s")
            params.append(filters['member_id'])
        else:
            return JsonResponse({'success': True, 'candidates': [], 'page': 1, 'num_pages': 0})
    else:
        if all_members:
            format_strings = ','.join(['%s'] * len(all_members))
            where_clauses.append(f"c.hr_member_id IN ({format_strings})")
            params.extend(all_members)
        else:
            return JsonResponse({'success': True, 'candidates': [], 'page': 1, 'num_pages': 0})
    
    if filters['search_query']:
        where_clauses.append("(c.name LIKE %s OR c.email LIKE %s OR c.phone LIKE %s)")
        params.extend([f"%{filters['search_query']}%"] * 3)
    
    # Only apply date filters if both dates are provided
    if filters['from_date'] and filters['to_date']:
        where_clauses.append("ca.activity_date BETWEEN %s AND %s")
        params.append(filters['from_date'])
        params.append(filters['to_date']) 

    where_statement = " AND ".join(where_clauses) if where_clauses else "1=1"

    # get page number
    page = int(request.GET.get('page', 1))
    limit = 10
    offset = (page - 1) * limit

    cursor.execute(f"""
        SELECT DISTINCT c.candidate_id, c.name, c.email, c.phone, c.experience, jd.jd_summary, c.updated_at
        FROM candidates c
        LEFT JOIN recruitment_jds jd ON c.jd_id = jd.jd_id
        LEFT JOIN candidate_activities ca ON c.candidate_id = ca.candidate_id
        WHERE {where_statement}
        ORDER BY c.updated_at DESC
        LIMIT %s OFFSET %s
    """, (*params, limit, offset))
    candidates = cursor.fetchall()
    total_count = len(candidates)
    num_pages = (total_count // limit) + (1 if total_count % limit else 0)
    DataOperations.close_db_connection(conn, cursor)
    return JsonResponse({'success': True, 'candidates': candidates, 'page': page, 'num_pages': num_pages})
        
    
@login_required
def get_candidate_muster(request, candidate_id):
    """
    API endpoint to get muster details of a specific candidate.
    """
    if request.method != "GET":
        return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)

    if not candidate_id or not str(candidate_id).isdigit():
        return JsonResponse({'success': False, 'error': 'Invalid candidate ID'}, status=400)
    candidate_id = int(candidate_id)

    role = request.session.get('role', 'Guest')
    user_id = request.session.get('user_id', None)
    emp_id = DataOperations.get_emp_id_from_user_id(user_id)
    if not emp_id:
        return JsonResponse({'success': False, 'error': 'Employee ID not found for user'}, status=404)

    # Get date filters
    FILTERS = {
        'from_date': request.GET.get('from_date', ''),
        'to_date': request.GET.get('to_date', '')
    }
    
    if not FILTERS['from_date']:
        # select start date of current month
        FILTERS['from_date'] = datetime.now().replace(month=datetime.now().month-3, day=1).strftime('%Y-%m-%d')

    if not FILTERS['to_date']:
        # select end date of current month
        FILTERS['to_date'] = (datetime.now().replace(day=1, month=datetime.now().month+1) - timedelta(days=1)).strftime('%Y-%m-%d')

    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)


    if role in ['Team_Lead', 'User']:
        cursor.execute("""
            select candidate_id, team_id from candidates
            where candidate_id=%s and team_id in (select team_id from team_members where emp_id=%s)
        """, (candidate_id, emp_id))
        candidate_row = cursor.fetchone()
        if not candidate_row:
            DataOperations.close_db_connection(conn, cursor)
            return JsonResponse({'success': False, 'error': 'Permission denied to access this candidate'}, status=403)
    
    cursor.execute("""
        SELECT
            ca.activity_id,
            c.candidate_id,
            c.name AS candidate_name,
            rj.jd_summary,
            ca.activity_date,
            ca.notes,
            ca.activity_type,
            ca.note_title,
            CONCAT(htm.first_name, ' ', htm.last_name) AS hr_member_name
        FROM
            candidate_activities AS ca
        LEFT JOIN
            candidates AS c ON ca.candidate_id = c.candidate_id
        LEFT JOIN
            recruitment_jds AS rj ON c.jd_id = rj.jd_id
        LEFT JOIN
            hr_team_members AS htm ON ca.emp_id = htm.emp_id
        LEFT JOIN
            team_members AS tm ON htm.emp_id = tm.emp_id AND c.team_id = tm.team_id
        WHERE
            c.candidate_id=%s AND ca.activity_date BETWEEN %s AND %s
        ORDER BY ca.updated_at DESC
    """, (candidate_id, FILTERS['from_date'], FILTERS['to_date']))
    activity_records = cursor.fetchall()
    cursor.close()
    conn.close()
    return JsonResponse({'success': True, 'activity_records': activity_records})
    
@login_required
def add_candidate_activity(request):
    """
    API endpoint to add a new activity for a candidate.
    """
    if request.method != "POST":
        return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)

    data = json.loads(request.body)
    candidate_id = data.get('candidate_id')
    activity_type = data.get('activity_type')
    activity_title =  data.get('activity_title', 'Untitled Activity').strip()
    notes = data.get('notes', '').strip()
    priority = data.get('priority', 'medium').strip().lower()

    if not candidate_id or not candidate_id.isdigit():
        return JsonResponse({'success': False, 'error': 'Invalid candidate ID'}, status=400)
    candidate_id = int(candidate_id)

    
    

    if not activity_type or activity_type not in ('interview_feedback','screening_notes','hr_notes','technical_assessment','offer_details','onboarding','rejection','general','other'):
        return JsonResponse({'success': False, 'error': 'Invalid activity type'}, status=400)

    user_id = request.session.get('user_id', None)
    emp_id = DataOperations.get_emp_id_from_user_id(user_id)
    if not emp_id:
        return JsonResponse({'success': False, 'error': 'Employee ID not found for user'}, status=404)

    role = request.session.get('role', 'Guest')
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if role in ['Team_Lead', 'User']:
        cursor.execute("""
            select candidate_id, team_id from candidates
            where candidate_id=%s and team_id in (select team_id from team_members where emp_id=%s)
        """, (candidate_id, emp_id))
        candidate_row = cursor.fetchone()
        if not candidate_row:
            DataOperations.close_db_connection(conn, cursor)
            return JsonResponse({'success': False, 'error': 'Permission denied to access this candidate'}, status=403)

    try:
        cursor.execute("""
            INSERT INTO candidate_activities (candidate_id, emp_id, activity_type, note_title, notes)
            VALUES (%s, %s, %s, %s, %s)
        """, (candidate_id, emp_id, activity_type, activity_title, notes))
        conn.commit()
        return JsonResponse({'success': True}, status=201)
    except Exception as e:
        conn.rollback()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@role_required(['Admin'], is_api=True)
def update_candidate_activity(request):
    """
    API endpoint to update an existing activity for a candidate.
    """
    if request.method != "POST":
        return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)

    data = json.loads(request.body)
    activity_id = data.get('activity_id')
    notes = data.get('notes', '').strip()

    if not activity_id or not str(activity_id).isdigit():
        return JsonResponse({'success': False, 'error': 'Invalid activity ID'}, status=400)
    activity_id = int(activity_id)

    if not notes:
        return JsonResponse({'success': False, 'error': 'Notes cannot be empty'}, status=400)

    user_id = request.session.get('user_id', None)
    emp_id = DataOperations.get_emp_id_from_user_id(user_id)
    if not emp_id:
        return JsonResponse({'success': False, 'error': 'Employee ID not found for user'}, status=404)

    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("""
            UPDATE candidate_activities
            SET notes=%s, updated_at=NOW()
            WHERE activity_id=%s
        """, (notes, activity_id))
        if cursor.rowcount == 0:
            conn.rollback()
            return JsonResponse({'success': False, 'error': 'Activity not found or no changes made'}, status=404)
        conn.commit()
        return JsonResponse({'success': True})
    except Exception as e:
        conn.rollback()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    finally:
        DataOperations.close_db_connection(conn, cursor)

@role_required('Admin', is_api=True)
def delete_candidate_activity(request):
    """
    API endpoint to delete an existing activity for a candidate.
    """
    if request.method != "POST":
        return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)

    data = json.loads(request.body)
    activity_id = data.get('activity_id')

    if not activity_id or not str(activity_id).isdigit():
        return JsonResponse({'success': False, 'error': 'Invalid activity ID'}, status=400)
    activity_id = int(activity_id)

    user_id = request.session.get('user_id', None)
    emp_id = DataOperations.get_emp_id_from_user_id(user_id)
    if not emp_id:
        return JsonResponse({'success': False, 'error': 'Employee ID not found for user'}, status=404)

    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("DELETE FROM candidate_activities WHERE activity_id=%s", (activity_id,))
        if cursor.rowcount == 0:
            conn.rollback()
            return JsonResponse({'success': False, 'error': 'Activity not found'}, status=404)
        conn.commit()
        return JsonResponse({'success': True})
    except Exception as e:
        conn.rollback()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    finally:
        DataOperations.close_db_connection(conn, cursor)

@login_required
def candidate_muster_page(request):
    """
    View to render the candidate muster page.
    """
    name = request.session.get('name', 'Guest')
    user_id = request.session.get('user_id', None)
    user_role_ = request.session.get('role', 'Guest')

    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    query = "SELECT team_id, team_name FROM teams"
    params = []

    if user_role_ in ['Team_Lead', 'User']:
        cursor.execute("SELECT emp_id FROM hr_team_members WHERE email=(SELECT email FROM users WHERE user_id=%s) LIMIT 1", [user_id])
        emp_row = cursor.fetchone()
        if emp_row:
            query += " WHERE lead_emp_id=%s OR team_id IN (SELECT team_id FROM team_members WHERE emp_id=%s)"
            params = [emp_row['emp_id'], emp_row['emp_id']]
    
    query += " ORDER BY team_name"
    if params:
        cursor.execute(query, params)
    else:
        cursor.execute(query)
    teams = cursor.fetchall()

    jds = []

    if user_role_ in ['Team_Lead', 'User']:
        cursor.execute("""
            SELECT DISTINCT j.jd_id, j.jd_summary
            FROM recruitment_jds j
            LEFT JOIN teams t ON j.team_id = t.team_id
            LEFT JOIN team_members tm ON t.team_id = tm.team_id
            WHERE (t.lead_emp_id=%s OR tm.emp_id=%s)
            AND j.jd_status='active'
            ORDER BY j.jd_summary
        """, [emp_row['emp_id'], emp_row['emp_id']])
    else:
        cursor.execute("SELECT jd_id, jd_summary FROM recruitment_jds WHERE jd_status='active' ORDER BY jd_summary")

    jds = cursor.fetchall()
    cursor.close()
    conn.close()

    return render(request, 'candidate_muster.html', {'name': name, 'teams': teams, 'user_role': user_role_, 'jds': jds})


@csrf_exempt
def get_candidate_details_profile(request):
    """Fetch candidate details based on name or email using raw SQL."""
    if request.method == 'GET':
        search_query = request.GET.get('query', '').strip()
        print("get_candidate_details -> Search query:", search_query)
        if search_query:
            user_role = request.session.get('role', 'Guest')
            user_id = request.session.get('user_id', None)
            conn = DataOperations.get_db_connection()
            cursor = conn.cursor(dictionary=True)
            params = [f"%{search_query}%", f"%{search_query}%"]
            if user_role == 'Admin':
                # Admin can search all candidates
                query = """
                    SELECT candidate_id, name, email, phone, skills, experience, screened_remarks,
                           l1_comments, l2_comments, l3_comments, screen_status
                    FROM candidates
                    WHERE email LIKE %s OR name LIKE %s
                    LIMIT 1
                """
            elif user_role in ['Team_Lead', 'User']:
                # Restrict to candidates in user's team
                # Get emp_id for user
                cursor.execute("SELECT emp_id FROM hr_team_members WHERE email=(SELECT email FROM users WHERE user_id=%s) LIMIT 1", [user_id])
                emp_row = cursor.fetchone()
                team_id = None
                if emp_row:
                    cursor.execute("SELECT team_id FROM team_members WHERE emp_id=%s LIMIT 1", [emp_row['emp_id']])
                    team_row = cursor.fetchone()
                    if team_row:
                        team_id = team_row['team_id']
                if team_id:
                    query = """
                        SELECT candidate_id, name, email, phone, skills, experience, screened_remarks,
                               l1_comments, l2_comments, l3_comments, screen_status
                        FROM candidates
                        WHERE (email LIKE %s OR name LIKE %s) AND team_id = %s
                        LIMIT 1
                    """
                    params.append(team_id)
                else:
                    # No team found, return not found
                    cursor.close()
                    conn.close()
                    return JsonResponse({'success': False, 'message': 'No team found for user.'})
            else:
                # Other roles not allowed
                cursor.close()
                conn.close()
                return JsonResponse({'success': False, 'message': 'Permission denied.'})

            cursor.execute(query, params)
            candidate = cursor.fetchone()
            print("get_candidate_details -> Candidate found:", candidate)
            if candidate:
                return JsonResponse({
                    'success': True,
                    'data': {
                        'candidate_id': candidate['candidate_id'],
                        'name': candidate['name'],
                        'email': candidate['email'],
                        'phone': candidate['phone'],
                        'skills': candidate['skills'],
                        'experience': candidate['experience'],
                        'screened_remarks': candidate['screened_remarks'],
                        'l1_comments': candidate['l1_comments'],
                        'l2_comments': candidate['l2_comments'],
                        'l3_comments': candidate['l3_comments'],
                        'status': candidate['screen_status'],
                    }
                })
            cursor.close()
            conn.close()
        return JsonResponse({'success': False, 'message': 'Candidate not found.'})

def save_candidate_details_profile(request):
    """Save updated candidate details using raw SQL."""
    if request.method == 'POST':
        candidate_id = request.POST.get('candidate_id')
        screened_remarks = request.POST.get('screened_remarks', '')
        l1_comments = request.POST.get('l1_comments', '')
        l2_comments = request.POST.get('l2_comments', '')
        l3_comments = request.POST.get('l3_comments', '')
        status = request.POST.get('status', '')

        conn = DataOperations.get_db_connection()
        cursor = conn.cursor(dictionary=True)
        # get candidate_prev_data
        cursor.execute("SELECT jd_id, screen_status, l1_result, l2_result, l3_result FROM candidates WHERE candidate_id = %s", (candidate_id,))
        candidate_prev_data = cursor.fetchone()
        if not candidate_prev_data:
            return JsonResponse({'success': False, 'message': 'Candidate not found.'}, status=404)
        candidate_new_data = {
            "jd_id": candidate_prev_data['jd_id'],
            "screen_status": status,
            "l1_result": candidate_prev_data['l1_result'],
            "l2_result": candidate_prev_data['l2_result'],
            "l3_result": candidate_prev_data['l3_result']
        }
        cursor.execute("""
            UPDATE candidates
            SET screened_remarks = %s, l1_comments = %s, l2_comments = %s, l3_comments = %s, screen_status = %s
            WHERE candidate_id = %s
        """, [screened_remarks, l1_comments, l2_comments, l3_comments, status, candidate_id])

        check = DataOperations.update_recruitment_jds(cursor, candidate_prev_data, candidate_new_data)
        if not check:
            conn.rollback()
            return JsonResponse({'success': False, 'message': 'Error updating recruitment_jds counts after candidate status update'}, status=500)
        conn.commit()
        cursor.close()
        conn.close()
        return JsonResponse({'success': True, 'message': 'Candidate details updated successfully.'})
    return JsonResponse({'success': False, 'message': 'Invalid request method.'})


@csrf_exempt
def candidate_suggestions(request):
    """
    API endpoint to get candidate suggestions based on a search query.
    """
    user_id = request.session.get('user_id', None)
    user_role = request.session.get('role', 'Guest')

    if request.method == 'GET':
        query = request.GET.get('q', '').strip()
        if len(query) < 3:
            return JsonResponse({'results': []})
        conn = DataOperations.get_db_connection()
        cursor = conn.cursor(dictionary=True)
        params = [f"%{query}%", f"%{query}%"]
        if user_role == 'Admin':
            sql = """
                SELECT candidate_id, name, email
                FROM candidates
                WHERE name LIKE %s OR email LIKE %s
                LIMIT 8
            """
        elif user_role in ['Team_Lead', 'User']:
            # Get emp_id for user
            cursor.execute("SELECT emp_id FROM hr_team_members WHERE email=(SELECT email FROM users WHERE user_id=%s) LIMIT 1", [user_id])
            emp_row = cursor.fetchone()
            team_id = None
            if emp_row:
                cursor.execute("SELECT team_id FROM team_members WHERE emp_id=%s LIMIT 1", [emp_row['emp_id']])
                team_row = cursor.fetchone()
                if team_row:
                    team_id = team_row['team_id']
            if team_id:
                sql = """
                    SELECT candidate_id, name, email
                    FROM candidates
                    WHERE (name LIKE %s OR email LIKE %s) AND team_id = %s
                    LIMIT 8
                """
                params.append(team_id)
            else:
                cursor.close()
                conn.close()
                return JsonResponse({'results': []})
        else:
            cursor.close()
            conn.close()
            return JsonResponse({'results': []})
        cursor.execute(sql, params)
        results = cursor.fetchall()
        cursor.close()
        conn.close()
        return JsonResponse({'results': results})
    return JsonResponse({'results': []})


def dashboard_data(request):
    """
    API endpoint to get dashboard data for the logged-in user.
    """
    email = request.session['email'] if 'email' in request.session else None
    print("dashboard_data -> User email:", email)
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Get emp_id for logged-in HR member
    cursor.execute("SELECT emp_id FROM hr_team_members WHERE email=%s", (email,))
    emp_row = cursor.fetchone()
    emp_id = emp_row['emp_id'] if emp_row else None
    print("dashboard_data -> Employee ID:", emp_id)
    # Pending/Active JDs assigned to user via team membership
    # Python
    # Logic Fault @1627
    cursor.execute("""
        SELECT
            r.jd_id,
            r.jd_summary,
            r.jd_status,
            cu.company_name,
            COUNT(
                CASE
                    WHEN (c.l3_result IS NULL OR c.l3_result != 'selected')
                    AND c.screen_status != 'rejected'
                    AND c.l1_result != 'rejected'
                    AND c.l2_result != 'rejected'
                    AND c.l3_result != 'rejected'
                    AND c.hr_member_id = %s THEN 1
                    ELSE NULL
                END
            ) AS not_finalized_count
        FROM
            recruitment_jds r
        JOIN
            customers cu ON r.company_id = cu.company_id
        LEFT JOIN
            candidates c ON r.jd_id = c.jd_id
        WHERE
            r.jd_status = 'active'
            AND r.team_id IN (
                SELECT team_id FROM team_members WHERE emp_id = %s
            )
        GROUP BY
            r.jd_id, r.jd_summary, r.jd_status, cu.company_name
        ORDER BY
            r.jd_id DESC;
    """, (emp_id, emp_id))

    pending_jds = cursor.fetchall()
    print("dashboard_data -> Pending JDs:", pending_jds)

    # Monthly closed JDs and candidates (last 6 months)
    monthly_report = []
    for i in range(6, 0, -1):
        month_start = (datetime.now().replace(day=1) - timedelta(days=30*i)).replace(day=1)
        month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        cursor.execute("""
            SELECT COUNT(DISTINCT r.jd_id) AS closed_jds
            FROM recruitment_jds r
            JOIN candidates c ON r.jd_id = c.jd_id
            WHERE c.hr_member_id = %s AND r.jd_status = 'closed'
            AND r.closure_date BETWEEN %s AND %s
        """, (emp_id, month_start.date(), month_end.date()))
        closed_jds = cursor.fetchone()['closed_jds'] or 0

        cursor.execute("""
            SELECT COUNT(*) AS candidates
            FROM candidates
            WHERE hr_member_id = %s AND screen_status = 'selected'
            AND l1_result = 'selected'
            AND l3_result = 'selected'
            AND created_at BETWEEN %s AND %s
        """, (emp_id, month_start.date(), month_end.date()))
        candidates = cursor.fetchone()['candidates'] or 0

        monthly_report.append({
            "month": month_start.strftime("%b %Y"),
            "closed_jds": closed_jds,
            "candidates": candidates
        })

    # Pie chart: active JDs by customer
    cursor.execute("""
        SELECT cu.company_name, COUNT(DISTINCT r.jd_id) AS jd_count
        FROM recruitment_jds r
        JOIN customers cu ON r.company_id = cu.company_id
        JOIN candidates c ON r.jd_id = c.jd_id
        WHERE c.hr_member_id = %s AND r.jd_status = 'active'
        GROUP BY cu.company_name
    """, (emp_id,))
    pie_rows = cursor.fetchall()

    customer_pie = {
        "labels": [row['company_name'] for row in pie_rows],
        "data": [row['jd_count'] for row in pie_rows]
    }

    # Bar chart: closed JDs in last 6 months
    bar_labels = [row['month'] for row in monthly_report]
    bar_data = [row['closed_jds'] for row in monthly_report]
    closed_jds_bar = {
        "labels": bar_labels,
        "data": bar_data
    }
    # In Progress Candidates
    cursor.execute("""
            SELECT 
                c.candidate_id,
                c.name,
                c.jd_id,
                c.l1_result,
                c.l2_result,
                c.l3_result,
                cu.company_name,
                t.team_name
            FROM candidates c
            LEFT JOIN recruitment_jds r ON c.jd_id = r.jd_id
            LEFT JOIN customers cu ON r.company_id = cu.company_id
            LEFT JOIN teams t ON c.team_id = t.team_id
            WHERE c.hr_member_id = %s
              AND r.jd_status = 'active'
              AND (c.screen_status IN ('toBeScreened', 'onHold', 'selected'))
              AND (c.l3_result IS NULL OR c.l3_result != 'selected')
            ORDER BY c.updated_at DESC
            LIMIT 30
        """, (emp_id,))
    in_progress_candidates = cursor.fetchall()

    cursor.close()
    conn.close()

    return JsonResponse({
        "pending_jds": pending_jds,
        "monthly_report": monthly_report,
        "customer_pie": customer_pie,
        "closed_jds_bar": closed_jds_bar,
        "in_progress_candidates": in_progress_candidates
    })

def offer_letter_page(request):
    """
    View to render the offer letter page.
    """
    name = request.session.get('name', 'Guest')
    return render(request, 'offer_letter.html', {'name': name})

@csrf_exempt
def generate_offer_letter(request):
    """
    API endpoint to generate an offer letter.
    """
    if request.method == 'POST':
        data = json.loads(request.body)
        candidate_id = data.get('candidate_id')
        basic = float(data.get('basic', 0))
        hra = float(data.get('hra', 0))
        special_allowance = float(data.get('special_allowance', 0))
        pf = float(data.get('pf', 0))
        gratuity = float(data.get('gratuity', 0))
        bonus = float(data.get('bonus', 0))
        other = float(data.get('other', 0))
        total_ctc = sum([basic, hra, special_allowance, pf, gratuity, bonus, other])

        # Save to DB
        conn = DataOperations.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO offer_letters (candidate_id, basic, hra, special_allowance, pf, gratuity, bonus, other, total_ctc)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, [candidate_id, basic, hra, special_allowance, pf, gratuity, bonus, other, total_ctc])
        conn.commit()
        

        # Generate offer letter HTML
        offer_html = f"""
        <div class='offer-letter'>
            <h2>Offer Letter</h2>
            <p>Candidate ID: {candidate_id}</p>
            <table class='salary-stack'>
                <tr><th>Component</th><th>Amount (INR)</th></tr>
                <tr><td>Basic</td><td>{basic:.2f}</td></tr>
                <tr><td>HRA</td><td>{hra:.2f}</td></tr>
                <tr><td>Special Allowance</td><td>{special_allowance:.2f}</td></tr>
                <tr><td>PF</td><td>{pf:.2f}</td></tr>
                <tr><td>Gratuity</td><td>{gratuity:.2f}</td></tr>
                <tr><td>Bonus</td><td>{bonus:.2f}</td></tr>
                <tr><td>Other</td><td>{other:.2f}</td></tr>
                <tr class='total'><td>Total CTC</td><td>{total_ctc:.2f}</td></tr>
            </table>
            <p>Congratulations! Please find your salary stack above.</p>
        </div>
        """
        
        # TO DO: send notification  to team lead about offere letter generated to candidate and Customer
        # get hr_member_id and team_id from candidates table
        cursor.execute("SELECT name, team_id, hr_member_id FROM candidates WHERE candidate_id=%s", [candidate_id])
        candidate_data = cursor.fetchone()
        print("Candidate data for notification:", candidate_data)
        if candidate_data:
            # get user_id from table users from hr_member_id
            hr_member_id = candidate_data['hr_member_id']
            hr_user_id = DataOperations.get_user_id_from_emp_id(hr_member_id)
            # get lead user id in similar way, lead emp id will be available in teams table
            team_id = candidate_data['team_id']
            cursor.execute("SELECT lead_emp_id FROM teams WHERE team_id=%s", [team_id])
            team_row = cursor.fetchone()
            if team_row:
                lead_emp_id = team_row['lead_emp_id']
                lead_user_id = DataOperations.get_user_id_from_emp_id(lead_emp_id)
                message = f"Offer letter generated for candidate {candidate_data['name']} with CTC INR {total_ctc:.2f}."
                if hr_user_id and DataOperations.get_user_settings(hr_user_id).get('notifications_enabled', False):
                    MessageProviders.send_notification(
                        user_id=hr_user_id,
                        title="Offer Letter Generated",
                        message=message
                    )
                if lead_user_id and lead_user_id != hr_user_id and DataOperations.get_user_settings(lead_user_id).get('notifications_enabled', False):
                    MessageProviders.send_notification(
                        user_id=lead_user_id,
                        title="Offer Letter Generated",
                        message=message
                    )
        cursor.close()
        conn.close()
        return JsonResponse({'success': True, 'offer_html': offer_html})
    return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)

@role_required(['Admin', 'Team_Lead'], is_api=True)
def teams_list(request):
    """
    API endpoint to get the list of teams.
    """
    print("teams_list -> Request method:", request.method)
    user_role = request.session.get('role', 'Guest')
    user_id = request.session.get('user_id', None)
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    query = "SELECT team_id, team_name FROM teams"
    params = []
    if user_role == 'Admin':
        pass
    elif user_role == 'Team_Lead':
        query += " WHERE lead_emp_id = (SELECT emp_id FROM hr_team_members WHERE email=(SELECT email FROM users WHERE user_id=%s) LIMIT 1)"
        params.append(user_id)
    else:
        query += "1=0"

    cursor.execute(query, params)
    teams = [{"id": row['team_id'], "name": row['team_name']} for row in cursor.fetchall()]
    cursor.close()
    conn.close()
    print("teams_list -> Teams fetched:", teams)
    return JsonResponse({"teams": teams})

def teams_filters(request):
    """
    API endpoint to get filters for teams.
    """
    print("teams_filters -> Request method:", request.method)
    role = request.session.get('role', 'Guest')
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    if role == 'Admin':
        cursor.execute("SELECT emp_id, first_name, last_name FROM hr_team_members WHERE status='active'")
    if role == 'Team_Lead':
        user_id = request.session.get('user_id', None)
        cursor.execute("""
            SELECT emp_id, first_name, last_name 
            FROM hr_team_members 
            WHERE status='active' AND emp_id IN (
                SELECT emp_id FROM team_members 
                WHERE team_id IN (
                    SELECT team_id FROM teams 
                    WHERE lead_emp_id = (SELECT emp_id FROM hr_team_members WHERE email=(SELECT email FROM users WHERE user_id=%s) LIMIT 1)
                )
            )
        """, (user_id,))
    members = [{"id": row['emp_id'], "name": f"{row['first_name']} {row['last_name']}"} for row in cursor.fetchall()]
    cursor.execute("SELECT company_id, company_name FROM customers")
    # Python
    customers = [{"id": row["company_id"], "name": row["company_name"]} for row in cursor.fetchall()]
    cursor.close()
    conn.close()
    return JsonResponse({"members": members, "customers": customers})

@role_required(['Admin', 'Team_Lead'])
def team_reports_page(request):
    """
    View to render the team reports page.
    """
    name = request.session.get('name', 'Guest')
    return render(request, 'team_reports.html', {'name': name})

@csrf_exempt
@role_required(['Admin', 'Team_Lead'], is_api=True)
def team_report(request):
    """
    API endpoint to get team report data.
    """
    print("team_reports_api -> Request method:", request.method)
    if request.method != "POST":
        return JsonResponse({}, status=400)
    try:
        params = json.loads(request.body)
        conne = DataOperations.get_db_connection()
        cursor = conne.cursor(dictionary=True)

        user_role = request.session.get('role', 'Guest')
        user_id = request.session.get('user_id', None)
        
        is_team_lead = (user_role == 'Team_Lead')
        team_name = params.get("team_search", "")

        # Use helper to get team lead data and perform access checks
        lead_team_ids, error_response = _get_team_lead_data(cursor, is_team_lead, user_id, team_name)
        if error_response:
            return error_response

        team_filter_clause = _get_team_filter_clause(is_team_lead, team_name, lead_team_ids)
        
        # Define date range
        start_date_value = params.get("start_date", "")
        end_date_value = params.get("end_date", "")
        today = date.today()
        start_date = start_date_value if start_date_value else today.replace(day=1).strftime('%Y-%m-%d')
        end_date = end_date_value if end_date_value else (today + timedelta(days=1)).strftime('%Y-%m-%d')
        
        # Fetch all data sections
        team_overview = _fetch_team_overview(cursor, team_filter_clause)
        recruitment_metrics = _fetch_recruitment_metrics(cursor, params, team_filter_clause, start_date, end_date)
        candidate_pipeline = _fetch_candidate_pipeline(cursor, team_filter_clause, start_date, end_date)
        member_contribution = _fetch_member_contribution(cursor, team_filter_clause)
        customer_distribution = _fetch_customer_distribution(cursor, team_filter_clause)
        performance_analytics = _get_performance_analytics(candidate_pipeline, start_date, end_date, cursor, team_filter_clause)
        
        return JsonResponse({
            "team_overview": team_overview,
            "recruitment_metrics": recruitment_metrics,
            "candidate_pipeline": candidate_pipeline,
            "member_contribution": member_contribution,
            "customer_distribution": customer_distribution,
            'performance_analytics': performance_analytics,
        })
    except Exception as e:
        print(f"Error in team_report: {e}")
        return JsonResponse({"error": "Internal Server Error"}, status=500)
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'conne' in locals() and conne:
            conne.close()

def _get_team_lead_data(cursor, is_team_lead, user_id, team_name):
    """Fetches team lead data and performs access checks."""
    if not is_team_lead:
        return [], None
    
    try:
        cursor.execute("SELECT emp_id FROM hr_team_members WHERE email = (SELECT email FROM users WHERE user_id = %s)", (user_id,))
        emp_row = cursor.fetchone()
        emp_id = emp_row['emp_id'] if emp_row else None
        
        if not emp_id:
            return None, JsonResponse({"error": "User data not found"}, status=404)

        cursor.execute("SELECT team_id FROM teams WHERE lead_emp_id = %s", (emp_id,))
        lead_team_ids = [row['team_id'] for row in cursor.fetchall()]

        if team_name:
            cursor.execute("SELECT team_id FROM teams WHERE team_name = %s AND lead_emp_id = %s", (team_name, emp_id))
            if not cursor.fetchone():
                return None, JsonResponse({"error": "Access Denied"}, status=403)
        return lead_team_ids, None
    except Exception as e:
        print(f"Error fetching team lead data: {e}")
        return None, JsonResponse({"error": "Internal Server Error"}, status=500)

def _get_team_filter_clause(is_team_lead, team_name, lead_team_ids):
    """Constructs the SQL WHERE clause for team filtering."""
    if is_team_lead and not team_name:
        return f"t.team_id IN ({','.join(map(str, lead_team_ids))})" if lead_team_ids else "1=2"
    elif is_team_lead and team_name:
        return f"LOWER(t.team_name) = LOWER('{team_name}') AND t.team_id IN ({','.join(map(str, lead_team_ids))})"
    elif team_name:
        return f"LOWER(t.team_name) = LOWER('{team_name}')"
    else:
        return "1=1"

def _fetch_team_overview(cursor, team_filter_clause):
    """Fetches and formats team overview data."""
    overview_sql = f"""
        SELECT 
            t.team_name, 
            CONCAT(m.first_name, ' ', m.last_name) AS team_lead,
            GROUP_CONCAT(CONCAT(tm2.first_name, ' ', tm2.last_name) SEPARATOR ', ') AS members
        FROM teams t
        LEFT JOIN hr_team_members m ON t.lead_emp_id = m.emp_id
        LEFT JOIN team_members tm ON t.team_id = tm.team_id
        LEFT JOIN hr_team_members tm2 ON tm.emp_id = tm2.emp_id
        WHERE {team_filter_clause}
        GROUP BY t.team_id
    """
    cursor.execute(overview_sql)
    return [{
        "team_name": row['team_name'],
        "team_lead": row['team_lead'] if row['team_lead'] else "",
        "members": row['members'].split(", ") if row['members'] else []
    } for row in cursor.fetchall()]

def _fetch_recruitment_metrics(cursor, params, team_filter_clause, start_date, end_date):
    """Fetches, formats, and calculates totals for recruitment metrics."""
    req_metrics_query = """
        SELECT
            r.team_id as team_id,
            t.team_name AS team_name,
            t.lead_emp_id AS team_lead_id,
            CONCAT(htm.first_name, ' ', htm.last_name) AS team_lead_name,
            COUNT(r.jd_id) AS total_jds,
            SUM(CASE WHEN r.jd_status = 'active' THEN 1 ELSE 0 END) AS in_progress,
            SUM(CASE WHEN r.jd_status = 'closed' THEN 1 ELSE 0 END) AS closed,
            AVG(CASE WHEN r.jd_status = 'closed' THEN DATEDIFF(r.closure_date, r.created_at) ELSE NULL END) AS avg_closure_time
        FROM recruitment_jds r
        JOIN teams t ON r.team_id = t.team_id
        JOIN hr_team_members htm on htm.emp_id = t.lead_emp_id
        LEFT JOIN customers c ON r.company_id = c.company_id
        WHERE 1=1 AND (r.closure_date BETWEEN %s AND %s OR r.jd_status <> 'closed')
    """
    query_params = [start_date, end_date]

    if team_filter_clause != "1=1":
        req_metrics_query += f" AND ({team_filter_clause})"

    if params.get("team_search"):
        req_metrics_query += F" AND LOWER(t.team_name) = LOWER(%s)"
        query_params.append(params["team_search"])
    if params.get("jd_status"):
        req_metrics_query += " AND r.jd_status = %s"
        query_params.append(params["jd_status"])
    if params.get("customer"):
        req_metrics_query += " AND c.company_name = %s"
        query_params.append(params["customer"])

    req_metrics_query += " GROUP BY r.team_id ORDER BY t.team_name"

    cursor.execute(req_metrics_query, tuple(query_params))
    recruitment_metrics = [{
        "team_id": row['team_id'],
        "team_name": row['team_name'],
        "team_lead": row['team_lead_name'] if row['team_lead_name'] else "",
        "total_jds": int(row['total_jds'] or 0),
        "in_progress": int(row['in_progress'] or 0),
        "closed": int(row['closed'] or 0),
        "avg_closure_time": float(row['avg_closure_time'] or 0)
    } for row in cursor.fetchall()]

    return _calculate_totals(recruitment_metrics, 'recruitment')

def _fetch_candidate_pipeline(cursor, team_filter_clause, start_date, end_date):
    """Fetches, formats, and calculates totals for the candidate pipeline."""
    pipeline_sql = f"""
        SELECT
            t.team_name,
            COUNT(DISTINCT c.candidate_id) AS total_candidates,
            SUM(CASE WHEN c.screen_status='toBeScreened' THEN 1 ELSE 0 END) AS sourced,
            SUM(CASE WHEN c.screen_status='selected' AND c.l1_result NOT IN ('selected', 'rejected') AND c.l2_result NOT IN ('selected', 'rejected') AND c.l3_result NOT IN ('selected', 'rejected') THEN 1 ELSE 0 END) AS screened,
            SUM(CASE WHEN c.l1_result='selected' AND c.l2_result NOT IN ('selected', 'rejected') AND c.l3_result NOT IN ('selected', 'rejected') THEN 1 ELSE 0 END) AS l1,
            SUM(CASE WHEN c.l2_result='selected' AND c.l3_result NOT IN ('selected', 'rejected') THEN 1 ELSE 0 END) AS l2,
            SUM(CASE WHEN c.l3_result='selected' THEN 1 ELSE 0 END) AS l3,
            SUM(CASE WHEN c.offer_status='released' THEN 1 ELSE 0 END) AS offered,
            SUM(CASE WHEN c.offer_status='accepted' THEN 1 ELSE 0 END) AS accepted,
            SUM(CASE WHEN c.screen_status='rejected' OR c.l1_result='rejected' OR c.l2_result='rejected' OR c.l3_result='rejected' OR c.offer_status='declined' THEN 1 ELSE 0 END) AS rejected,
            SUM(CASE WHEN c.offer_status NOT IN ('not_initiated', 'in_progress') THEN 1 ELSE 0 END) AS position_offered
        FROM candidates c
        JOIN teams t ON c.team_id = t.team_id
        WHERE (c.updated_at BETWEEN %s AND %s) AND ({team_filter_clause})
        GROUP BY t.team_id
        ORDER BY t.team_name
    """
    cursor.execute(pipeline_sql, (start_date, end_date))
    candidate_pipeline = [{
        "team_name": row['team_name'],
        "total_candidates": int(row['total_candidates'] or 0),
        "sourced": int(row['sourced'] or 0),
        "screened": int(row['screened'] or 0),
        "l1": int(row['l1'] or 0),
        "l2": int(row['l2'] or 0),
        "l3": int(row['l3'] or 0),
        "offered": int(row['offered'] or 0),
        "accepted": int(row['accepted'] or 0),
        "rejected": int(row['rejected'] or 0),
        "position_offered": int(row['position_offered'] or 0)
    } for row in cursor.fetchall()]

    return _calculate_totals(candidate_pipeline, 'pipeline')

def _calculate_totals(data, data_type):
    """Calculates and appends a 'Total' row to a list of dictionaries."""
    if not data:
        return data

    totals = {}
    if data_type == 'recruitment':
        fields = ['total_jds', 'in_progress', 'closed']
        for field in fields:
            totals[field] = sum(item[field] for item in data)
        avg_closure_time = round(sum(item['avg_closure_time'] for item in data) / len(data), 2) if data else 0
        totals['avg_closure_time'] = avg_closure_time
        totals.update({"team_id": "Total", "team_name": "Total", "team_lead": ""})
    else: # data_type == 'pipeline'
        fields = ['total_candidates', 'sourced', 'screened', 'l1', 'l2', 'l3', 'offered', 'accepted', 'rejected']
        for field in fields:
            totals[field] = sum(item[field] for item in data)
        totals["team_name"] = "Total"

    data.append(totals)
    return data

def _get_performance_analytics(candidate_pipeline, start_date, end_date, cursor, team_filter_clause):
    """Calculates performance rates and monthly trends."""
    team_conversion_rates = []
    team_success_rates = []
    overall_offered = 0
    overall_accepted = 0
    overall_total = 0
    
    # Exclude the "Total" row from calculations
    for team_row in candidate_pipeline[:-1]:
        total = int(team_row.get('total_candidates', 0))
        offered = int(team_row.get('position_offered', 0))
        accepted = int(team_row.get('accepted', 0))
        
        conv_rate = round((offered / total) * 100, 2) if total else 0
        succ_rate = round((accepted / offered) * 100, 2) if offered else 0
        
        team_conversion_rates.append({'team_name': team_row['team_name'], 'conversion_rate': conv_rate})
        team_success_rates.append({'team_name': team_row['team_name'], 'success_rate': succ_rate})
        
        overall_total += total
        overall_offered += offered
        overall_accepted += accepted
    
    overall_conv_rate = round((overall_offered / overall_total) * 100, 2) if overall_total else 0
    overall_succ_rate = round((overall_accepted / overall_offered) * 100, 2) if overall_offered else 0
    
    monthly_trends = _get_trends_data(start_date, end_date, cursor, team_filter_clause)
    
    return {
        'team_conversion_rates': team_conversion_rates,
        'team_success_rates': team_success_rates,
        'overall_conversion_rate': overall_conv_rate,
        'overall_success_rate': overall_succ_rate,
        'monthly_trends': monthly_trends
    }

def _get_trends_data(start_date_str, end_date_str, cursor, team_filter_clause):
    """Generates monthly/yearly trends for closed JDs."""
    start_dt = datetime.strptime(start_date_str, '%Y-%m-%d')
    end_dt = datetime.strptime(end_date_str, '%Y-%m-%d')
    diff_months = (end_dt.year - start_dt.year) * 12 + (end_dt.month - start_dt.month) + 1
    
    monthly_trends = {'labels': [], 'datasets': []}
    team_colors = ['#36a2eb', '#4caf50', '#ff9800', '#e91e63', '#9c27b0', '#009688', '#795548', '#607d8b', '#f44336', '#8bc34a']

    cursor.execute(f"SELECT t.team_id, t.team_name FROM teams t WHERE {team_filter_clause}")
    teams = cursor.fetchall()
    
    time_buckets = []
    if diff_months <= 3:
        for i in range(6, 0, -1):
            bucket_start = (end_dt.replace(day=1) - relativedelta(months=i))
            bucket_end = (end_dt.replace(day=1) - relativedelta(months=i-1))
            time_buckets.append((bucket_start, bucket_end))
        monthly_trends['labels'] = [b[0].strftime('%b %Y') for b in time_buckets]
    elif diff_months > 10:
        for y in range(start_dt.year, end_dt.year + 1):
            year_start = datetime(y, 1, 1)
            year_end = datetime(y + 1, 1, 1) if y < end_dt.year else end_dt + timedelta(days=1)
            time_buckets.append((year_start, year_end))
        monthly_trends['labels'] = [b[0].strftime('%Y') for b in time_buckets]
    else:
        cur = start_dt.replace(day=1)
        while cur <= end_dt:
            next_month = (cur + relativedelta(months=1))
            time_buckets.append((cur, next_month))
            cur = next_month
        monthly_trends['labels'] = [b[0].strftime('%b %Y') for b in time_buckets]

    for idx, team in enumerate(teams):
        team_id, team_name = team['team_id'], team['team_name']
        color = team_colors[idx % len(team_colors)]
        data_points = []
        for bucket_start, bucket_end in time_buckets:
            cursor.execute("""
                SELECT COUNT(*) as cnt FROM recruitment_jds r
                WHERE r.jd_status='closed' AND r.team_id=%s AND r.closure_date >= %s AND r.closure_date < %s
            """, (team_id, bucket_start.date(), bucket_end.date()))
            count = int(cursor.fetchone()['cnt'] or 0)
            data_points.append(count)
        monthly_trends['datasets'].append({
            'label': team_name,
            'data': data_points,
            'borderColor': color,
            'backgroundColor': color,
            'fill': False,
            'tension': 0.3
        })
    return monthly_trends

def _fetch_member_contribution(cursor, team_filter_clause):
    """Fetches and formats member contribution data."""
    member_sql = f"""
        SELECT m.first_name, m.last_name,
            COUNT(DISTINCT j.jd_id) AS jds_handled, 
            COUNT(c.candidate_id) AS candidates_processed,
            SUM(CASE WHEN c.offer_status='released' THEN 1 ELSE 0 END) AS offers_made
        FROM hr_team_members m
        LEFT JOIN team_members tm ON m.emp_id = tm.emp_id
        LEFT JOIN teams t ON tm.team_id = t.team_id
        LEFT JOIN recruitment_jds j ON t.team_id = j.team_id
        LEFT JOIN candidates c ON j.jd_id = c.jd_id AND c.hr_member_id = m.emp_id
        WHERE {team_filter_clause}
        GROUP BY m.emp_id
    """
    cursor.execute(member_sql)
    return [{
        "member": f"{row['first_name']} {row['last_name']}",
        "jds_handled": int(row['jds_handled'] or 0),
        "candidates_processed": int(row['candidates_processed'] or 0),
        "offers_made": int(row['offers_made'] or 0),
        "top_performer": False
    } for row in cursor.fetchall()]

def _fetch_customer_distribution(cursor, team_filter_clause):
    """Fetches and formats customer distribution data."""
    customer_sql = f"""
        SELECT c.company_name, COUNT(j.jd_id) AS jds_handled, SUM(CASE WHEN cand.screen_status='selected' THEN 1 ELSE 0 END) AS candidates_placed
        FROM customers c
        LEFT JOIN recruitment_jds j ON c.company_id = j.company_id
        LEFT JOIN teams t ON j.team_id = t.team_id
        LEFT JOIN candidates cand ON j.jd_id = cand.jd_id
        WHERE {team_filter_clause}
        GROUP BY c.company_id
    """
    cursor.execute(customer_sql)
    return [{
        "customer": row['company_name'],
        "jds_handled": int(row['jds_handled'] or 0),
        "candidates_placed": int(row['candidates_placed'] or 0)
    } for row in cursor.fetchall()]

def team_reports_export(request):
    """
    API endpoint to export team reports.
    """
    team_search = request.GET.get('team_search', '')
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM teams WHERE team_name LIKE %s", (f"%{team_search}%",))
    teams = cursor.fetchall()
    team_ids = [t['team_id'] for t in teams]
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="team_reports.csv"'
    writer = csv.writer(response)
    writer.writerow(['Team Name', 'Team Lead', 'Members', 'Total JDs', 'In Progress', 'Closed', 'Avg Closure Time'])
    for t in teams:
        cursor.execute("SELECT first_name, last_name FROM hr_team_members WHERE emp_id IN (SELECT emp_id FROM team_members WHERE team_id=%s AND role LIKE 'lead%%')", (t['team_id'],))
        lead = cursor.fetchone()
        cursor.execute("SELECT COUNT(*) as cnt FROM team_members WHERE team_id=%s", (t['team_id'],))
        members = cursor.fetchone()['cnt']
        cursor.execute("SELECT COUNT(*) as cnt FROM recruitment_jds WHERE team_id=%s", (t['team_id'],))
        total_jds = cursor.fetchone()['cnt']
        cursor.execute("SELECT COUNT(*) as cnt FROM recruitment_jds WHERE team_id=%s AND jd_status='active'", (t['team_id'],))
        in_progress = cursor.fetchone()['cnt']
        cursor.execute("SELECT COUNT(*) as cnt FROM recruitment_jds WHERE team_id=%s AND jd_status='closed'", (t['team_id'],))
        closed = cursor.fetchone()['cnt']
        writer.writerow([
            t['team_name'],
            f"{lead['first_name']} {lead['last_name']}" if lead else '',
            members,
            total_jds,
            in_progress,
            closed,
            ''  # Avg closure time can be added if needed
        ])
    cursor.close()
    conn.close()
    return response



@role_required(['Admin', 'Team_Lead'])
def task_progress_reports_page(request):
    """
    View to render the task progress reports page.
    """
    name = request.session.get('name', 'Guest')
    return render(request, "task_progress_reports.html", {"name": name})

def team_report_filters(request):
    """
    API endpoint to get filters for team reports.
    """
    # print("team_report_filters -> Request method:", request.method)
    role = request.session.get('role', 'Guest')
    user_id = request.session.get('user_id', None)
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    if role == 'Team_Lead':
        cursor.execute("SELECT emp_id FROM hr_team_members WHERE email = (SELECT email FROM users WHERE user_id = %s)", (user_id,))
        emp_row = cursor.fetchone()
        emp_id = emp_row['emp_id'] if emp_row else None
        if not emp_id:
            return JsonResponse({"error": "User data not found"}, status=404)
        cursor.execute("SELECT team_id FROM teams WHERE lead_emp_id = %s", (emp_id,))
        lead_team_ids = [row['team_id'] for row in cursor.fetchall()]
        if not lead_team_ids:
            return JsonResponse({"teams": []})
        format_strings = ','.join(['%s'] * len(lead_team_ids))
        cursor.execute(f"SELECT team_id, team_name FROM teams WHERE team_id IN ({format_strings})", tuple(lead_team_ids))
        
    else:
        cursor.execute("SELECT team_id, team_name FROM teams")
    teams = cursor.fetchall()
    cursor.close()
    conn.close()
    return JsonResponse({"teams": teams})


@csrf_exempt
@role_required(['Admin', 'Team_Lead'], is_api=True)
def team_reports_api(request):
    """
    API endpoint to get team reports.
    """
    conn = None
    cursor = None
    try:
        conn = DataOperations.get_db_connection()
        cursor = conn.cursor(dictionary=True)

        params = {
            "team_id": request.GET.get("team_id"),
            "jd_status": request.GET.get("jd_status"),
            "from_date": request.GET.get("from_date"),
            "to_date": request.GET.get("to_date"),
            "role": request.session.get('role', 'Guest'),
            "user_id": request.session.get('user_id', None)
        }

        # Centralized access control and parameter validation
        access_result, access_error = _check_access_and_get_team_ids(cursor, params)
        if access_error:
            return access_error

        lead_team_ids = access_result
        
        # Consolidate all data fetching into helper functions
        jd_progress = _fetch_jd_progress(cursor, params, lead_team_ids)
        profile_status_chart = _fetch_profile_status_chart(cursor, params, lead_team_ids)
        team_contribution = _fetch_team_contribution(cursor, params, lead_team_ids)
        timeline_chart = _fetch_timeline_chart(cursor, params, lead_team_ids)
        
        # Calculate derived metrics
        jd_completion_chart = _get_jd_completion_chart(jd_progress)

        # Single return statement
        return JsonResponse({
            "jd_progress": jd_progress,
            "profile_status_chart": profile_status_chart,
            "jd_completion_chart": jd_completion_chart,
            "team_contribution": team_contribution,
            "timeline_chart": timeline_chart
        })

    except Exception as e:
        print(f"Error in team_reports_api: {e}")
        return JsonResponse({"error": "Internal Server Error"}, status=500)
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

# --- Helper Functions ---

def _check_access_and_get_team_ids(cursor, params):
    """
    Handles Team_Lead access control and returns a list of team IDs
    or an error response if access is denied.
    """
    lead_team_ids = None
    if params["role"] == 'Team_Lead':
        cursor.execute("SELECT emp_id FROM hr_team_members WHERE email = (SELECT email FROM users WHERE user_id = %s)", (params["user_id"],))
        emp_row = cursor.fetchone()
        emp_id = emp_row['emp_id'] if emp_row else None

        if not emp_id:
            lead_team_ids = []
        else:
            cursor.execute("SELECT team_id FROM teams WHERE lead_emp_id = %s", (emp_id,))
            lead_team_ids = [row['team_id'] for row in cursor.fetchall()]

        if params["team_id"] and (not lead_team_ids or int(params["team_id"]) not in lead_team_ids):
            # Team Lead requesting a team they don't lead
            return None, JsonResponse({
                "jd_progress": [],
                "profile_status_chart": {"labels": ["To Be Screened", "Selected", "Rejected", "On Hold"], "data": [0, 0, 0, 0]},
                "jd_completion_chart": {"labels": [], "data": []},
                "team_contribution": [],
                "timeline_chart": {"labels": [], "datasets": []}
            })
    return lead_team_ids, None

def _build_query_with_filters(base_query, params, lead_team_ids):
    """
    Dynamically builds the WHERE clause and a list of parameters for a query.
    Returns (query_string, query_params_list).
    """
    query_parts = [base_query]
    query_params = []
    
    # Handle team filtering based on role
    if params["role"] == 'Team_Lead':
        if lead_team_ids:
            format_strings = ','.join(['%s'] * len(lead_team_ids))
            if not params["team_id"]:
                query_parts.append(f"AND jd.team_id IN ({format_strings})")
                query_params.extend(lead_team_ids)
            else:
                query_parts.append("AND jd.team_id = %s")
                query_params.append(params["team_id"])
        else: # Team lead with no teams
            query_parts.append("AND 1=2") # Force no results
    elif params["team_id"]:
        query_parts.append("AND jd.team_id = %s")
        query_params.append(params["team_id"])
    
    # Add other filters if present
    if params["jd_status"]:
        query_parts.append("AND jd.jd_status = %s")
        query_params.append(params["jd_status"])
    if params["from_date"]:
        query_parts.append("AND jd.created_at >= %s")
        query_params.append(params["from_date"])
    if params["to_date"]:
        query_parts.append("AND jd.created_at <= %s")
        query_params.append(params["to_date"])
        
    return " ".join(query_parts), query_params

def _fetch_jd_progress(cursor, params, lead_team_ids):
    """Fetches JD progress data based on filters."""
    base_query = """
        SELECT jd.jd_id, jd.jd_summary, t.team_name, jd.jd_status, jd.total_profiles,
            jd.profiles_in_progress, jd.profiles_completed, jd.profiles_selected,
            jd.profiles_rejected, jd.profiles_on_hold
        FROM recruitment_jds jd
        LEFT JOIN teams t ON jd.team_id = t.team_id
        WHERE 1=1
    """
    query, query_params = _build_query_with_filters(base_query, params, lead_team_ids)
    cursor.execute(query, query_params)
    return cursor.fetchall()

def _fetch_profile_status_chart(cursor, params, lead_team_ids):
    """Fetches data for the profile status pie chart."""
    base_query = """
        SELECT c.screen_status, COUNT(*) as cnt
        FROM candidates c
        LEFT JOIN recruitment_jds jd ON c.jd_id = jd.jd_id
        WHERE 1=1
    """
    query, query_params = _build_query_with_filters(base_query, params, lead_team_ids)
    query += " GROUP BY c.screen_status"
    cursor.execute(query, query_params)
    
    status_counts = {row["screen_status"]: row["cnt"] for row in cursor.fetchall()}
    return {
        "labels": ["To Be Screened", "Selected", "Rejected", "On Hold"],
        "data": [
            status_counts.get("toBeScreened", 0),
            status_counts.get("selected", 0),
            status_counts.get("rejected", 0),
            status_counts.get("onHold", 0)
        ]
    }

def _get_jd_completion_chart(jd_progress):
    """Calculates the JD completion chart data from progress data."""
    return {
        "labels": [jd["jd_id"] for jd in jd_progress],
        "data": [
            int((jd["profiles_selected"] / jd["total_profiles"] * 100) if jd["total_profiles"] else 0)
            for jd in jd_progress
        ]
    }

def _fetch_team_contribution(cursor, params, lead_team_ids):
    """Fetches team/member contribution data."""
    base_query = """
        SELECT c.jd_id, t.team_name, CONCAT(m.first_name, ' ', m.last_name) AS member_name,
            COUNT(c.candidate_id) AS profiles_processed,
            SUM(c.screen_status='selected') AS selected,
            SUM(c.screen_status='rejected') AS rejected,
            SUM(c.screen_status='onHold') AS on_hold
        FROM candidates c
        LEFT JOIN teams t ON c.team_id = t.team_id
        LEFT JOIN hr_team_members m ON c.hr_member_id = m.emp_id
        WHERE 1=1
    """
    # Build query for Team_Lead or Admin/other roles
    query_parts = [base_query]
    query_params = []
    
    if params["role"] == 'Team_Lead':
        if lead_team_ids:
            format_strings = ','.join(['%s'] * len(lead_team_ids))
            query_parts.append(f"AND c.team_id IN ({format_strings})")
            query_params.extend(lead_team_ids)
        else:
            query_parts.append("AND 1=2") # No teams, force no results
    
    query_parts.append("GROUP BY c.jd_id, t.team_name, member_name")
    
    cursor.execute(" ".join(query_parts), query_params)
    return cursor.fetchall()

def _fetch_timeline_chart(cursor, params, lead_team_ids):
    """Fetches data for the timeline chart."""
    base_query = """
        SELECT DATE(screened_on) as date, 
            SUM(screen_status='selected') as selected,
            SUM(screen_status='rejected') as rejected,
            SUM(screen_status='onHold') as on_hold,
            COUNT(*) as processed
        FROM candidates
        WHERE screened_on IS NOT NULL
    """
    # Build query for Team_Lead or Admin/other roles
    query_parts = [base_query]
    query_params = []
    
    if params["role"] == 'Team_Lead':
        if lead_team_ids:
            format_strings = ','.join(['%s'] * len(lead_team_ids))
            query_parts.append(f"AND team_id IN ({format_strings})")
            query_params.extend(lead_team_ids)
        else:
            query_parts.append("AND 1=2")
            
    query_parts.append("GROUP BY DATE(screened_on) ORDER BY date ASC")
    
    cursor.execute(" ".join(query_parts), query_params)
    rows = cursor.fetchall()
    
    labels = [r["date"].strftime("%Y-%m-%d") for r in rows]
    return {
        "labels": labels,
        "datasets": [
            {"label": "Processed", "data": [r["processed"] for r in rows], "borderColor": "#2563eb", "fill": False},
            {"label": "Selected", "data": [r["selected"] for r in rows], "borderColor": "#16a34a", "fill": False},
            {"label": "Rejected", "data": [r["rejected"] for r in rows], "borderColor": "#dc2626", "fill": False},
            {"label": "On Hold", "data": [r["on_hold"] for r in rows], "borderColor": "#f59e42", "fill": False}
        ]
    }

def api_jd_detail(request, jd_id):
    """
    API endpoint to get details of a specific job description.
    """
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT c.candidate_id, c.name, c.email, c.phone, c.screen_status as status,
            t.team_name, CONCAT(m.first_name, ' ', m.last_name) AS member_name
        FROM candidates c
        LEFT JOIN teams t ON c.team_id = t.team_id
        LEFT JOIN hr_team_members m ON c.hr_member_id = m.emp_id
        WHERE c.jd_id = %s
    """, (jd_id,))
    candidates = cursor.fetchall()
    cursor.close()
    conn.close()
    return JsonResponse({"candidates": candidates})

def team_reports_export(request):
    """
    API endpoint to export team reports as CSV.
    """
    team_id = request.GET.get("team_id")
    jd_status = request.GET.get("jd_status")
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor()
    query = """
        SELECT jd.jd_id, jd.jd_summary, t.team_name, jd.jd_status, jd.total_profiles,
            jd.profiles_in_progress, jd.profiles_completed, jd.profiles_selected,
            jd.profiles_rejected, jd.profiles_on_hold
        FROM recruitment_jds jd
        LEFT JOIN teams t ON jd.team_id = t.team_id
        WHERE 1=1
    """
    params = []
    if team_id:
        query += " AND jd.team_id = %s"
        params.append(team_id)
    if jd_status:
        query += " AND jd.jd_status = %s"
        params.append(jd_status)
    if from_date:
        query += " AND jd.created_at >= %s"
        params.append(from_date)
    if to_date:
        query += " AND jd.created_at <= %s"
        params.append(to_date)
    cursor.execute(query, params)
    rows = cursor.fetchall()
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = "attachment; filename=task_progress_report.csv"
    writer = csv.writer(response)
    writer.writerow([
        "JD ID", "Summary", "Team", "Status", "Total Profiles",
        "In Progress", "Completed", "Selected", "Rejected", "On Hold"
    ])
    for row in rows:
        writer.writerow(row)
    cursor.close()
    conn.close()
    return response

@role_required(['Admin', 'Team_Lead'])
def candidate_conversion_rates_page(request):
    name = request.session.get('name', 'Guest')
    return render(request, "candidate_conversion_rates.html",{'name': name})

@role_required(['Admin', 'Team_Lead'], is_api=True)
def ccr_filters(request):
    """
    API endpoint to get filters for candidate conversion rates.
    """

    role = request.session.get("role", "Guest")
    user_id = request.session.get("user_id", None)

    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    team_filtering_clause = "1=1"
    if role == "Team_Lead":
        team_filtering_clause = f"team_id IN ({','.join(map(str, DataOperations.get_team_lead_teams(user_id)))})"
    
    cursor.execute(f"SELECT jd_id, jd_summary FROM recruitment_jds WHERE {team_filtering_clause}")
    jds = cursor.fetchall()
    cursor.execute(f"SELECT team_id, team_name FROM teams WHERE {team_filtering_clause}")
    teams = cursor.fetchall()
    cursor.close()
    conn.close()
    return JsonResponse({"jds": jds, "teams": teams})

def _build_ccr_query_params(request, role, user_id):
    """
    Builds the dynamic WHERE clauses and parameters for CCR reports based on user filters and role.
    """
    jd_id = request.GET.get("jd_id")
    team_id = request.GET.get("team_id")
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")

    if not from_date:
        from_date = datetime.now().replace(day=1).strftime('%Y-%m-%d')
    if not to_date:
        to_date = datetime.now().strftime('%Y-%m-%d')
    
    where_clauses = ["1=1"]
    sql_params = []
    
    team_ids = []
    if role == "Team_Lead":
        team_ids = DataOperations.get_team_lead_teams(user_id)
        if not team_ids:
            # Force no results if Team Lead has no assigned teams
            return where_clauses, sql_params, True
        format_strings = ','.join(['%s'] * len(team_ids))
        where_clauses.append(f"c.team_id IN ({format_strings})")
        sql_params.extend(team_ids)

    if jd_id:
        where_clauses.append("c.jd_id = %s")
        sql_params.append(jd_id)
    if team_id:
        if role == "Team_Lead" and int(team_id) not in team_ids:
            # Team Lead requesting an unauthorized team
            return where_clauses, sql_params, True
        where_clauses.append("c.team_id = %s")
        sql_params.append(team_id)
    if from_date:
        where_clauses.append("c.created_at >= %s")
        sql_params.append(from_date)
    if to_date:
        where_clauses.append("c.created_at <= %s")
        sql_params.append(to_date)

    where_clause_str = " AND ".join(where_clauses)
    return where_clause_str, tuple(sql_params), False

def _get_overall_funnel_data(cursor, where_clause, params):
    """
    Fetches the data for the overall funnel report.
    """
    query = f"""
        SELECT
            SUM(screen_status='selected') as screened,
            SUM(screen_status='selected' AND l1_result='selected') as l1,
            SUM(screen_status='selected' AND l1_result='selected' AND l2_result='selected') as l2,
            SUM(screen_status='selected' AND l1_result='selected' AND l2_result='selected' AND l3_result='selected') as l3,
            SUM(screen_status='selected' AND l1_result='selected' AND l2_result='selected' AND l3_result='selected' AND offer_status='accepted') as final_selected
        FROM candidates c
        WHERE {where_clause}
    """
    cursor.execute(query, params)
    return cursor.fetchone()

def _calculate_stage_conversion_rates(funnel_row):
    """
    Calculates conversion percentages between stages from funnel data.
    """
    screened = funnel_row["screened"] or 0
    l1 = funnel_row["l1"] or 0
    l2 = funnel_row["l2"] or 0
    l3 = funnel_row["l3"] or 0
    final_selected = funnel_row["final_selected"] or 0

    stage_data = [
        round((l1 / screened) * 100, 2) if screened else 0,
        round((l2 / l1) * 100, 2) if l1 else 0,
        round((l3 / l2) * 100, 2) if l2 else 0,
        round((final_selected / l3) * 100, 2) if l3 else 0,
    ]
    return stage_data

def _get_trend_analysis_data(cursor, where_clause, params):
    """
    Fetches daily trend data for candidate stages.
    """
    query = f"""
        SELECT DATE(created_at) as date,
            SUM(screen_status='selected') as screened,
            SUM(l1_result='selected') as l1,
            SUM(l2_result='selected') as l2,
            SUM(l3_result='selected') as l3
        FROM candidates c
        WHERE {where_clause}
        GROUP BY date
        ORDER BY date ASC
    """
    cursor.execute(query, params)
    rows = cursor.fetchall()
    
    trend_labels = [str(r["date"]) for r in rows]
    trend_datasets = [
        {"label": "Screened", "data": [r["screened"] for r in rows], "borderColor": "#2563eb", "fill": False},
        {"label": "L1", "data": [r["l1"] for r in rows], "borderColor": "#3b82f6", "fill": False},
        {"label": "L2", "data": [r["l2"] for r in rows], "borderColor": "#0ea5e9", "fill": False},
        {"label": "L3", "data": [r["l3"] for r in rows], "borderColor": "#16a34a", "fill": False},
    ]
    return trend_labels, trend_datasets

def _get_jd_conversion_rates(cursor, where_clause, params):
    """
    Fetches conversion rates for each job description.
    """
    query = f"""
        SELECT jd.jd_id, jd.jd_summary, t.team_name,
            COUNT(c.candidate_id) as total,
            SUM(c.screen_status='selected') as screened,
            SUM(c.l1_result='selected') as l1,
            SUM(c.l2_result='selected') as l2,
            SUM(c.l3_result='selected') as l3,
            SUM(c.l3_result='selected' AND c.offer_status='accepted') as final_selected
        FROM recruitment_jds jd
        LEFT JOIN teams t ON jd.team_id = t.team_id
        LEFT JOIN candidates c ON jd.jd_id = c.jd_id
        WHERE {where_clause.replace('c.', 'jd.')}  -- Use jd. and t. aliases for JD-specific filtering
        GROUP BY jd.jd_id, jd.jd_summary, t.team_name
    """
    # The original query's parameter order was a bit off, so let's adjust the query to be clearer.
    # We can rebuild the where clause for this specific query to ensure parameter order is correct.
    # Note: `where_clause` is generic, so we apply it carefully.
    
    cursor.execute(query, params)
    jd_rates = []
    for r in cursor.fetchall():
        total = r["total"] or 1
        final_selected = r["final_selected"] or 0
        conversion_pct = round((final_selected / total) * 100, 2) if total else 0
        jd_rates.append({
            "jd_summary": r["jd_summary"],
            "team_name": r["team_name"],
            "total": r["total"],
            "screened": r["screened"],
            "l1": r["l1"],
            "l2": r["l2"],
            "l3": r["l3"],
            "final_selected": r["final_selected"],
            "conversion_pct": conversion_pct
        })
    return jd_rates

def _get_team_conversion_performance(cursor, where_clause, params):
    """
    Fetches conversion performance by team and member.
    """
    query = f"""
        SELECT t.team_name, CONCAT(m.first_name, ' ', m.last_name) AS member_name,
            COUNT(c.candidate_id) as total,
            SUM(c.l3_result='selected' AND c.offer_status='accepted') as final_selected
        FROM candidates c
        LEFT JOIN teams t ON c.team_id = t.team_id
        LEFT JOIN hr_team_members m ON c.hr_member_id = m.emp_id
        WHERE {where_clause}
        GROUP BY t.team_name, member_name
    """
    cursor.execute(query, params)
    team_rates = []
    for r in cursor.fetchall():
        total = r["total"] or 1
        conversion_pct = round((r["final_selected"] / total) * 100, 2) if total else 0
        team_rates.append({
            "team_name": r["team_name"],
            "member_name": r["member_name"],
            "total": r["total"],
            "final_selected": r["final_selected"],
            "conversion_pct": conversion_pct
        })
    return team_rates

def _get_time_to_conversion_metrics(cursor, where_clause, params):
    """
    Fetches average time taken for each conversion stage.
    """
    query = f"""
        SELECT jd.jd_summary,
            AVG(DATEDIFF(c.l1_date, c.screened_on)) as screen_l1,
            AVG(DATEDIFF(c.l2_date, c.l1_date)) as l1_l2,
            AVG(DATEDIFF(c.l3_date, c.l2_date)) as l2_l3,
            AVG(DATEDIFF(c.updated_at, c.l3_date)) as l3_final
        FROM candidates c
        LEFT JOIN recruitment_jds jd ON c.jd_id = jd.jd_id
        WHERE {where_clause}
        GROUP BY jd.jd_summary
    """
    cursor.execute(query, params)
    time_metrics = []
    for r in cursor.fetchall():
        time_metrics.append({
            "jd_summary": r["jd_summary"],
            "screen_l1": round(r["screen_l1"] or 0, 1),
            "l1_l2": round(r["l1_l2"] or 0, 1),
            "l2_l3": round(r["l2_l3"] or 0, 1),
            "l3_final": round(r["l3_final"] or 0, 1)
        })
    return time_metrics

@csrf_exempt
@role_required(['Admin', 'Team_Lead'], is_api=True)
def ccr_reports_api(request):
    """
    API endpoint to get candidate conversion reports, refactored for clarity and modularity.
    """
    conn = None
    cursor = None
    try:
        role = request.session.get("role", "Guest")
        user_id = request.session.get("user_id", None)
        
        # Build dynamic WHERE clause and parameters once
        where_clause, params, unauthorized = _build_ccr_query_params(request, role, user_id)
        if unauthorized:
            return JsonResponse({"message": "Unauthorized access or no teams assigned."}, status=403)

        conn = DataOperations.get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # 1. Overall Funnel Report
        funnel_labels = ["Screened", "L1 Cleared", "L2 Cleared", "L3 Cleared", "Final Selected"]
        funnel_row = _get_overall_funnel_data(cursor, where_clause, params)
        funnel_data = [
            funnel_row["screened"] or 0,
            funnel_row["l1"] or 0,
            funnel_row["l2"] or 0,
            funnel_row["l3"] or 0,
            funnel_row["final_selected"] or 0
        ]
        
        # 2. Stage-wise Conversion Rates
        stage_labels = ["Screen → L1", "L1 → L2", "L2 → L3", "L3 → Final"]
        stage_data = _calculate_stage_conversion_rates(funnel_row)

        # 3. Trend Analysis
        trend_labels, trend_datasets = _get_trend_analysis_data(cursor, where_clause, params)
        
        # 4. JD-wise Conversion Rates
        jd_rates = _get_jd_conversion_rates(cursor, where_clause, params)
        
        # 5. Team/Member Conversion Performance
        team_rates = _get_team_conversion_performance(cursor, where_clause, params)

        # 6. Time-to-Conversion Metrics
        time_metrics = _get_time_to_conversion_metrics(cursor, where_clause, params)

        return JsonResponse({
            "funnel": {"labels": funnel_labels, "data": funnel_data},
            "stage_rates": {"labels": stage_labels, "data": stage_data},
            "trend": {"labels": trend_labels, "datasets": trend_datasets},
            "jd_rates": jd_rates,
            "team_rates": team_rates,
            "time_metrics": time_metrics
        })

    except Exception as e:
        print(f"Error generating CCR report: {e}")
        return JsonResponse({"message": "Internal Server Error"}, status=500)
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

def ccr_reports_export(request):
    """
    API endpoint to export candidate conversion reports.
    """
    jd_id = request.GET.get("jd_id")
    team_id = request.GET.get("team_id")
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor()
    query = """
        SELECT jd.jd_summary, t.team_name,
            COUNT(c.candidate_id) as total,
            SUM(c.screen_status='selected') as screened,
            SUM(c.l1_result='selected') as l1,
            SUM(c.l2_result='selected') as l2,
            SUM(c.l3_result='selected') as l3,
            SUM(c.l3_result='selected' AND c.screen_status='selected') as final_selected
        FROM recruitment_jds jd
        LEFT JOIN teams t ON jd.team_id = t.team_id
        LEFT JOIN candidates c ON jd.jd_id = c.jd_id
        WHERE 1=1
    """
    params = []
    if team_id:
        query += " AND jd.team_id = %s"
        params.append(team_id)
    if jd_id:
        query += " AND jd.jd_id = %s"
        params.append(jd_id)
    if from_date:
        query += " AND c.created_at >= %s"
        params.append(from_date)
    if to_date:
        query += " AND c.created_at <= %s"
        params.append(to_date)
    query += " GROUP BY jd.jd_summary, t.team_name"
    cursor.execute(query, params)
    rows = cursor.fetchall()
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = "attachment; filename=candidate_conversion_rates.csv"
    writer = csv.writer(response)
    writer.writerow([
        "JD", "Team", "Total Candidates", "Screened", "L1 Cleared", "L2 Cleared", "L3 Cleared", "Final Selected"
    ])
    for row in rows:
        writer.writerow(row)
    cursor.close()
    conn.close()
    return response


@login_required
def user_profile(request):
    """
    View to render the user profile page.
    """
    try:
        email = request.session.get('email')
        conn = DataOperations.get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT
                htm.emp_id,
                CONCAT(htm.first_name, ' ', htm.last_name) AS name,
                htm.email,
                htm.phone,
                htm.role,
                htm.date_joined,
                GROUP_CONCAT(tm.team_id SEPARATOR ', ') AS team_ids,
                GROUP_CONCAT(t.team_name SEPARATOR ', ') AS team_names
            FROM hr_team_members htm
            LEFT JOIN team_members tm ON htm.emp_id = tm.emp_id
            LEFT JOIN teams t ON tm.team_id = t.team_id
            WHERE htm.email = %s
            LIMIT 1
        """, (email,))
        userDetails = cursor.fetchone() or {}
        cursor.close()
        conn.close()
    except mysql.connector.Error as err:
        return HttpResponse(f"Database error: {err}", status=500)
    return render(request, "user_profile.html", {"user": userDetails})


@role_required('Admin')
def manage_sessions_view(request):
    """
    View to manage user sessions.
    """
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT us.session_id, us.user_id, us.expires_at, u.username, u.role
        FROM user_sessions us
        JOIN users u ON us.user_id = u.user_id
        WHERE us.expires_at > NOW()
        ORDER BY us.expires_at DESC
    """)
    sessions = []
    for row in cursor.fetchall():
        sessions.append({
            'session_id': row['session_id'],
            'user_id': row['user_id'],
            'username': row['username'],
            'role': row['role'],
            'expires_at': row['expires_at'].strftime('%Y-%m-%d %H:%M')
        })
    cursor.close()
    conn.close()
    name = request.session.get('name', 'Guest')
    return render(request, 'manage_sessions.html', {'sessions': json.dumps(sessions), 'name': name})

@csrf_exempt
def logout_session_api(request):
    """
    API endpoint to log out a user session.
    """
    if request.method == 'POST':
        data = json.loads(request.body)
        session_id = data.get('session_id')
        if session_id:
            conn = DataOperations.get_db_connection()
            cursor = conn.cursor()
            cursor.execute("DELETE FROM user_sessions WHERE session_id=%s", (session_id,))
            conn.commit()
            cursor.close()
            conn.close()
            return JsonResponse({'success': True})
    return JsonResponse({'success': False})


@login_required
def access_permissions(request):
    """
    View to manage access permissions.
    """

    is_admin = request.session.get('role') == "Admin"
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT user_id, username, email, role, is_active, created_at FROM users")
    users = cursor.fetchall()
    cursor.close()
    conn.close()
    return render(request, "access_permissions.html", {"users": users, "is_admin": is_admin, "roles": Constants.ROLES})


@csrf_exempt
@login_required
def change_password(request):
    """
    API endpoint to change user password.
    """
    if request.method == "POST":
        data = json.loads(request.body)
        old_password = data.get("old_password")
        new_password = data.get("new_password")
        user_id = request.session.get("user_id")
        conn = DataOperations.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT password_hash FROM users WHERE user_id=%s", (user_id,))
        row = cursor.fetchone()
        if not row or not check_password(old_password, row[0]):
            cursor.close()
            conn.close()
            return JsonResponse({"success": False, "message": "Current password is incorrect."})
        new_password_hash = make_password(new_password)
        cursor.execute("UPDATE users SET password_hash=%s WHERE user_id=%s", (new_password_hash, user_id))
        conn.commit()
        cursor.close()
        conn.close()
        return JsonResponse({"success": True, "message": "Password changed successfully."})
    return JsonResponse({"success": False, "message": "Invalid request."})

@csrf_exempt
@role_required('Admin', is_api=True)
def change_role(request):
    """
    API endpoint to change user role.
    """
    if request.method == "POST":
        data = json.loads(request.body)
        user_id = data.get("user_id")
        role = Constants.ROLES.get(data.get("role"), None)

        if not user_id or not role:
            return JsonResponse({"success": False, "message": "Invalid user ID or role."})
        
        conn = DataOperations.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE users SET role=%s WHERE user_id=%s", (role, user_id))
        conn.commit()
        cursor.close()
        conn.close()

        # change session of role changed user

        if user_id and DataOperations.get_user_settings(user_id).get('notifications_enabled', False):
            MessageProviders.send_notification(user_id, "Role Update", f"Your role has been changed to {role}, please re-login to see the changes.")
        return JsonResponse({"success": True, "message": "Role updated successfully."})
    return JsonResponse({"success": False, "message": "Invalid request."})

@role_required(['Admin'], is_api=True)
def change_user_status(request, user_id, action):
    """
    View to activate or deactivate a user.
    """
    if action not in ['activate', 'deactivate']:
        return HttpResponse("Invalid action.", status=400)
    try:
        emp_id = DataOperations.get_emp_id_from_user_id(user_id)
        is_active = 1 if action == 'activate' else 0
        conn = DataOperations.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE users SET is_active=%s WHERE user_id=%s", (is_active, user_id))
        if not is_active and emp_id:
            cursor.execute("UPDATE hr_team_members SET status='inactive' WHERE emp_id=%s", (emp_id,))
        conn.commit()
    except:
        conn.rollback()
        return HttpResponse("Error updating user status.", status=500)
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'conn' in locals() and conn:
            conn.close()
    return JsonResponse({"success": True, "message": f"User {'activated' if is_active else 'deactivated'} successfully."})

@role_required(['Admin', 'Team_Lead'])
def status_report_page(request):
    """
    View to render the status report page.
    """
    role = request.session.get("role")
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    if role == "Team_Lead":
        lead_team_ids = DataOperations.get_team_lead_teams(request.session.get("user_id"))
        if not lead_team_ids:
            return render(request, 'status_report.html', {"teams": [], "members": []})
        format_strings = ','.join(['%s'] * len(lead_team_ids))
        cursor.execute(f"SELECT team_id, team_name FROM teams WHERE team_id IN ({format_strings})", tuple(lead_team_ids))
    else:
        cursor.execute("SELECT team_id, team_name FROM teams")
    teams = [
        {
            "team_id": t["team_id"], 
            "team_name": t["team_name"]
        } for t in cursor.fetchall()
    ]
    if role == "Team_Lead" and lead_team_ids:
        format_strings = ','.join(['%s'] * len(lead_team_ids))
        cursor.execute(f"""
            SELECT emp_id, first_name, last_name 
            FROM hr_team_members 
            WHERE status='active' AND emp_id IN (
                SELECT emp_id FROM team_members WHERE team_id IN ({format_strings})
            )
        """, tuple(lead_team_ids))
    else:
        cursor.execute("SELECT emp_id, first_name, last_name FROM hr_team_members WHERE status='active'")
    members = [
        {
            "emp_id": m["emp_id"], 
            "first_name": m["first_name"], 
            "last_name": m["last_name"]
        } 
        for m in cursor.fetchall()
    ]
    cursor.close()
    conn.close()
    return render(request, 'status_report.html', {"teams": teams, "members": members})

def _validate_report_params(request_data, role, user_id):
    """Validates and sanitizes report parameters."""
    report_type = request_data.get("report_type")
    team_id = request_data.get("team_id")
    member_id = request_data.get("member_id")
    date_param = request_data.get("date")
    from_date = request_data.get("from_date")
    to_date = request_data.get("to_date")

    if report_type != "custom":
        if not date_param:
            return JsonResponse({"report": [], "message": "date is required."}, status=400)
    elif not from_date or not to_date:
        return JsonResponse({"report": [], "message": "From and To dates are required for custom report."}, status=400)

    # Adjust to_date to include the full day
    if to_date:
        to_date = (datetime.strptime(to_date, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")

    # Authorize and filter by team for Team_Lead role
    if role == "Team_Lead":
        lead_team_ids = DataOperations.get_team_lead_teams(user_id)
        if not lead_team_ids:
            return JsonResponse({"report": [], "message": "No teams assigned."}, status=403)
        if team_id and team_id != "all" and int(team_id) not in lead_team_ids:
            return JsonResponse({"report": [], "message": "Unauthorized team access."}, status=403)

    return {
        "report_type": report_type,
        "team_id": team_id,
        "member_id": member_id,
        "date": date_param,
        "from_date": from_date,
        "to_date": to_date,
        "lead_team_ids": lead_team_ids if role == "Team_Lead" else None
    }

def _build_where_clause(params):
    """Builds the dynamic WHERE clause for the SQL query."""
    where_clauses = []
    sql_params = []

    # Filter by user role
    if params["lead_team_ids"]:
        format_strings = ','.join(['%s'] * len(params["lead_team_ids"]))
        where_clauses.append(f"c.team_id IN ({format_strings})")
        sql_params.extend(params["lead_team_ids"])

    # Filter by team and member if specified
    if params["team_id"] and params["team_id"] != "all":
        where_clauses.append("c.team_id=%s")
        sql_params.append(params["team_id"])
    if params["member_id"] and params["member_id"] != "all":
        where_clauses.append("c.hr_member_id=%s")
        sql_params.append(params["member_id"])

    # Filter by date range
    if params["report_type"] == "daily" and params["date"]:
        where_clauses.append("DATE(c.shared_on)=%s")
        sql_params.append(params["date"])
    elif params["report_type"] == "weekly" and params["date"]:
        where_clauses.append("YEARWEEK(c.shared_on, 1)=YEARWEEK(%s, 1)")
        sql_params.append(params["date"])
    elif params["report_type"] == "custom" and params["from_date"] and params["to_date"]:
        where_clauses.append("DATE(c.shared_on) BETWEEN %s AND %s")
        sql_params.extend([params["from_date"], params["to_date"]])

    where_clause = " AND ".join(where_clauses) if where_clauses else "1=1"
    return where_clause, tuple(sql_params)

def _get_report_summary(cursor, where_clause, sql_params):
    """Fetches and processes data to create the main report and summary."""
    query = f"""
        SELECT
            cu.company_name,
            j.jd_summary,
            c.jd_id,
            COUNT(c.candidate_id) AS profile_count,
            GROUP_CONCAT(c.screened_remarks SEPARATOR ', ') AS feedback
        FROM candidates c
        JOIN recruitment_jds j ON c.jd_id = j.jd_id
        JOIN customers cu ON j.company_id = cu.company_id
        WHERE {where_clause}
        GROUP BY cu.company_name, j.jd_summary, c.jd_id
        ORDER BY MAX(c.shared_on) DESC
    """
    cursor.execute(query, sql_params)
    rows = cursor.fetchall()
    
    report = []
    unique_companies = set()
    unique_jds = set()
    total_profiles = 0

    for r in rows:
        report.append({
            "company_name": r['company_name'],
            "jd_summary": r['jd_summary'],
            "jd_id": r['jd_id'],
            "profile_count": r['profile_count'],
            "feedback": r['feedback'] or ""
        })
        unique_companies.add(r['company_name'])
        unique_jds.add(r['jd_id'])
        total_profiles += r['profile_count'] if r['profile_count'] else 0

    # Add summary line
    report.append({
        "company_name": f"Total Unique Companies: {len(unique_companies)}",
        "jd_summary": f"Total JDs: {len(unique_jds)}",
        "jd_id": "",
        "profile_count": f"Total Shared Profiles: {total_profiles}",
        "feedback": ""
    })
    
    return report

def _get_detailed_candidate_list(cursor, report_summary, params):
    """Fetches detailed candidate information for each JD in the report."""
    list_of_candidates = []
    
    date_or_date_range = params["date"] if params["report_type"] != "custom" else f"{params['from_date']} to {params['to_date']}"
    date_filtering_clause = "AND DATE(shared_on)=%s" if params["report_type"] != "custom" else "AND DATE(shared_on) BETWEEN %s AND %s"
    date_params = (params["date"],) if params["report_type"] != "custom" else (params["from_date"], params["to_date"])

    for row in report_summary[:-1]:  # Exclude summary row
        meta_data = {
            "date_or_date_range": date_or_date_range,
            "company_name": row["company_name"],
            "jd_summary": row["jd_summary"],
        }
        
        query = f"""
            SELECT 
                candidate_id, name, email, phone, 
                experience, current_ctc, expected_ctc, 
                notice_period, previous_job_profile as profile, 
                location, recruiter_comments
            FROM candidates 
            WHERE jd_id=%s {date_filtering_clause}
        """
        cursor.execute(query, (row["jd_id"], *date_params))
        candidates = cursor.fetchall()
        
        list_of_candidates.append({
            "metadata": meta_data,
            "candidates": candidates
        })
        
    return list_of_candidates

@csrf_exempt
@role_required(['Admin', 'Team_Lead'], is_api=True)
def generate_status_report(request):
    """
    API endpoint to generate a status report by modularizing the process.
    """
    if request.method != 'POST':
        return JsonResponse({"report": [], "list_of_candidates": [], "message": "Invalid request."}, status=405)

    conn = None
    cursor = None
    # if True:
    try:
        request_data = {
            "report_type": request.POST.get("report_type"),
            "team_id": request.POST.get("team_id"),
            "member_id": request.POST.get("member_id"),
            "date": request.POST.get("date"),
            "from_date": request.POST.get("from_date"),
            "to_date": request.POST.get("to_date"),        
        }
        
        # Step 1: Validate and sanitize request parameters
        params = _validate_report_params(request_data, request.session.get("role"), request.session.get("user_id"))
        if isinstance(params, JsonResponse):
            return params

        conn = DataOperations.get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Step 2: Build the dynamic SQL WHERE clause
        where_clause, sql_params = _build_where_clause(params)

        # Step 3: Get the main report summary
        report_summary = _get_report_summary(cursor, where_clause, sql_params)
        
        # Step 4: Get the detailed list of candidates
        detailed_candidates = _get_detailed_candidate_list(cursor, report_summary, params)

        return JsonResponse({
            "report": report_summary,
            "list_of_candidates": detailed_candidates,
            "message": "Report generated successfully."
        })

    except Exception as e:
        print(f"Error generating report: {e}")
        return JsonResponse({"report": [], "list_of_candidates": [], "message": "Internal Server Error"}, status=500)
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@csrf_exempt
def export_teams_excel(request):
    """
    API endpoint to export team data as Excel.
    """
    if request.method == "POST":
        data = json.loads(request.body)
        teams = data.get("teams", [])
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Teams"
        headers = ["Team Id", "Team Name", "Total Strength", "Created On"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="1F497D")
            cell.alignment = Alignment(horizontal="center")
        for team in teams:
            ws.append([
                team.get("team_id", ""),
                team.get("team_name", ""),
                team.get("strength", ""),
                team.get("created_at", "")
            ])
        # Auto-width columns
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2
        # Save to response
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=teams_export.xlsx"
        wb.save(response)
        return response
    return JsonResponse({"error": "Invalid request"}, status=400)

@csrf_exempt
def export_team_reports_excel(request):
    """
    API endpoint to export team reports as Excel.
    """
    print("Received request for exporting team reports to Excel.")
    if request.method == "POST":
        data = json.loads(request.body)
        print("Data received for export:", data)
        wb = openpyxl.Workbook()
        thin = Side(border_style="thin", color="000000")

        # Sheet mapping: key -> sheet name
        sheet_map = {
            "teamOverviewData": "Team Overview",
            "recruitmentMetricsData": "Recruitment Metrics",
            "candidatePipelineData": "Candidate Pipeline",
            "memberContributionData": "Member Contribution",
            "customerDistributionData": "Customer Distribution"
        }

        # Remove default sheet if not used
        default_sheet = wb.active
        default_sheet.title = "DeleteMe"
        used = False

        for key, sheet_name in sheet_map.items():
            rows = data.get(key, [])
            if not rows:
                continue
            if not used:
                ws = default_sheet
                ws.title = sheet_name
                used = True
            else:
                ws = wb.create_sheet(title=sheet_name)
            for i, row in enumerate(rows):
                ws.append(row)
                for j, cell in enumerate(ws[i+1]):
                    cell.font = Font(bold=(i==0), color="1F497D" if i==0 else "000000")
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    if i == 0:
                        cell.fill = PatternFill("solid", fgColor="D9E1F2")
            # Auto-width columns
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2

        # Remove default sheet if unused
        if not used:
            wb.remove(default_sheet)

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=team_reports.xlsx"
        wb.save(response)
        return response
    return HttpResponse(status=405)

def get_user_id_by_username(username):
    """
    Fetch user ID from the database using username.
    """
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT user_id FROM users WHERE username=%s", (username,))
    row = cursor.fetchone()
    cursor.close()
    conn.close()
    return row[0] if row else None

def notification_settings(request):
    """
    View to manage notification settings.
    """
    print("Accessing notification settings page.")
    #fetch user id for username
    username = request.session.get("username")
    user_id = get_user_id_by_username(username)
    print("Current user ID:", username, "for:", user_id)
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
   # python
    cursor.execute("SELECT notifications_enabled FROM user_settings WHERE user_id=%s", [user_id])
    row = cursor.fetchone()
    notifications_enabled = row["notifications_enabled"] if row else False
    print(f"User {username} notifications enabled: {notifications_enabled}")

    # get all notifications for user
    cursor.execute("SELECT * FROM notifications WHERE user_id=%s and is_read=%s", [user_id, False])
    notifications = cursor.fetchall()
    print(f"User {username} notifications: {notifications}")

    cursor.close()
    conn.close()
    name = request.session.get('name', 'Guest')
    print(f"User {username} notifications: {notifications}")
    return render(request, "settings_notification.html", {
        "notifications_enabled": notifications_enabled,
        "name": name,
        "notifications": notifications
    })

def mark_as_read_notification(request, notification_id):
    """
    API endpoint to mark a notification as read.
    """
    print("Accessing mark as read notification API.")
    if request.method == "POST":
        username = request.session.get("username")
        user_id = get_user_id_by_username(username)
        print("Marking notification as read for user ID:", username)

        conn = DataOperations.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE notifications SET is_read=true WHERE user_id=%s AND notification_id=%s", [user_id, notification_id])
        conn.commit()
        cursor.close()
        conn.close()
        return JsonResponse({"status": "success"})
    return JsonResponse({"error": "Invalid request"}, status=400)

def clear_all_notifications(request):
    user_name = request.session.get("username")
    try:
        user_id = get_user_id_by_username(user_name)
        print("Clearing all notifications for user ID:", user_name)

        conn = DataOperations.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM notifications WHERE user_id=%s", [user_id])
        conn.commit()
        cursor.close()
        conn.close()
        return JsonResponse({"status": "success"})
    except Exception as e:
        print("Error clearing notifications for user ID:", user_name, "Error:", e)
        return JsonResponse({"status": "error", "message": str(e)})

def notification_count(request):
    """
    API endpoint to get the count of unread notifications.
    """
    if request.method == "GET":
        username = request.session.get("username")
        user_id = get_user_id_by_username(username)
        print("Fetching notification count for user ID:", username)

        conn = DataOperations.get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT COUNT(*) as count FROM notifications WHERE user_id=%s AND is_read=false", [user_id])
        row = cursor.fetchone()
        cursor.close()
        conn.close()

        count = row["count"] if row else 0
        return JsonResponse({"notification_count": count})
    return JsonResponse({"error": "Invalid request"}, status=400)


@csrf_exempt
def toggle_notification(request):
    """
    API endpoint to toggle notification settings.
    """
    if request.method == "POST":
        # fetch username from session or request
        username = request.session.get("username")
        user_id = get_user_id_by_username(username)
        print("Toggling notification for user ID:", username)
        data = json.loads(request.body)
        enabled = data.get("enabled", True)

        conn = DataOperations.get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id FROM user_settings WHERE user_id=%s", [user_id])
        if cursor.fetchone():
            cursor.execute("UPDATE user_settings SET notifications_enabled=%s WHERE user_id=%s", [enabled, user_id])
        else:
            cursor.execute("INSERT INTO user_settings (user_id, notifications_enabled) VALUES (%s, %s)", [user_id, enabled])
        status_text = "Notifications are ON" if enabled else "Notifications are OFF"
        conn.commit()
        cursor.close()
        conn.close()
        return JsonResponse({"status_text": status_text})
    return JsonResponse({"error": "Invalid request"}, status=400)


@login_required
def save_email_config(request):
    session = request.session
    useremail = session.get('username')
    if not useremail:
        return redirect('login')
    
    mail_providers = MessageProviders.MAIL_SERVICE_PROVIDERS

    if request.method == "POST":
        print("save_email_config -> Saving email config for user:", request.session.get('username'))
        email = request.POST.get("email_address")
        email_host_password = request.POST.get("email_host_password")
        # Checkbox: present as 'on' if checked, None if not
        email_provider = request.POST.get("email_provider", "gmail")
        smtp_host = mail_providers.get(email_provider, {}).get("smtp_host", "smtp.gmail.com")
        smtp_port = mail_providers.get(email_provider, {}).get("smtp_port", 587)
        # Validate email_host and email_host_password
        if not email:
            messages.error(request, "Email should not be blank.")
            return redirect('save_email_config')
        elif DataValidators.is_valid_email(email) is False:
            messages.error(request, "Invalid email format.")
            return redirect('save_email_config')
        
        if not email_host_password:
            messages.error(request, "Host Password should not be blank.")
            return redirect('save_email_config')
        
        # Get user_id from users table
        user_id = get_user_id_by_username(useremail)

        if not user_id:
            return render(request, "email_config.html", {"user": {"email": useremail}, "error": "User not found."})
        print("Saving email config for user ID:", useremail, "->", user_id)
        # Check email credentials by sending a test mail
        from .utils import encrypt_password

        test_subject = "[QuantumNxt] Email Configuration Test"
        test_body = "<p>Your email configuration was tested and is working! If you did not request this, please ignore.</p>"
        test_result = MessageProviders.send_email(
            from_email=email,
            app_password=email_host_password,
            to_email=email,  # send to self
            subject=test_subject,
            html_body=test_body,
            smtp_host=smtp_host,
            smtp_port=smtp_port
        )
        
        if not test_result:
            messages.error(request, "Could not send test email. Please check your email address and app password.")
            return redirect('save_email_config')

        # Encrypt the password before saving
        encrypted_password = encrypt_password(email_host_password)

        try:
            conn = DataOperations.get_db_connection()
            cursor = conn.cursor(dictionary=True)
            # Upsert into email_config table (user_id, email, email_host_password)
            cursor.execute("""
                INSERT INTO email_config (user_id, email, email_smtp_host, email_smtp_port, email_host_password)
                VALUES (%s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE email=%s, email_smtp_host=%s, email_smtp_port=%s, email_host_password=%s
            """, (user_id, email, smtp_host, smtp_port, encrypted_password, email, smtp_host, smtp_port, encrypted_password))
            conn.commit()
        except Exception as e:
            print("save_email_config -> Error saving email config:", str(e))
            messages.error(request, "Error saving email configuration. Please try again.")
            return redirect('save_email_config')
        finally:
            cursor.close()
            conn.close()

        messages.success(request, "Email configuration successful!")
        return redirect('save_email_config')
    return render(request, "email_config.html", {"user": {"email": useremail}, "email_providers": mail_providers, "is_gmail": False})

def notifications_list(request):
    """
    View to render the notifications list page.
    """
    return render(request, 'notifications_list.html', {})

def get_email_configs(user_id):
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM email_config WHERE user_id=%s", [user_id])
    email_configs = cursor.fetchone()
    cursor.close()
    conn.close()
    return email_configs


def generate_daily_report(report_type, user_id, **kwargs):
    emp_id = DataOperations.get_emp_id_from_user_id(user_id)
    if not emp_id:
        print("Employee ID not found for user ID:", user_id)
        return None
    conn = DataOperations.get_db_connection()
    cursor = conn.cursor(dictionary=True)
    # get team ids for emp_id
    cursor.execute("""
        SELECT t.team_id, t.team_name FROM teams t
        left join team_members tm on tm.team_id=t.team_id
        WHERE emp_id=%s""", [emp_id])
    team_ids = cursor.fetchall()

    if report_type not in ["daily", "weekly", "monthly"]:
        print("Invalid report type:", report_type)
        return None

    if report_type == "daily":
        date = datetime.now().strftime("%Y-%m-%d")
        from_date = datetime.now().replace(hour=6, minute=0, second=0)
        to_date = datetime.now().replace(hour=23, minute=59, second=59)
    
    elif report_type == "weekly":
        # from date will be start date of ongoing week (Monday)
        date = datetime.now().strftime("%Y-%m-%d")
        from_date = datetime.now() - timedelta(days=datetime.now().weekday())
        to_date = (from_date + timedelta(days=6)).replace(hour=23, minute=59, second=59)

    elif report_type == "monthly":
        # from date will be start date of ongoing month (1st)
        date = datetime.now().strftime("%Y-%m-%d")
        from_date = datetime.now().replace(day=1, hour=0, minute=0, second=0)
        next_month = from_date.replace(day=28) + timedelta(days=4)  # this will never fail
        to_date = (next_month - timedelta(days=next_month.day)).replace(hour=23, minute=59, second=59)

    elif report_type == "custom":
        from_date_str = kwargs.get("from_date")
        to_date_str = kwargs.get("to_date")
        if not from_date_str or not to_date_str:
            print("From date and To date are required for custom report.")
            return None
        try:
            from_date = datetime.strptime(from_date_str, "%Y-%m-%d")
            to_date = datetime.strptime(to_date_str, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
        except ValueError:
            print("Invalid date format for custom report. Use YYYY-MM-DD.")
            return None

    request_data = {
                "report_type": 'custom',
                "team_id": None,
                "member_id": emp_id,
                "date": datetime.now().strftime("%Y-%m-%d"),
                "from_date": from_date.strftime("%Y-%m-%d"),
                "to_date": to_date.strftime("%Y-%m-%d"),
                "lead_team_ids": None
        }
    
    all_teams_data = []

    for team in team_ids:
        request_data['team_id'] = team['team_id']
        where_clause, sql_params = _build_where_clause(request_data)
        report_summary = _get_report_summary(cursor, where_clause, sql_params)
        detailed_candidates = _get_detailed_candidate_list(cursor, report_summary, request_data)
        all_teams_data.append({
            "team_name": team['team_name'],
            "report_summary": report_summary,
            "detailed_candidates": detailed_candidates
        })

    return all_teams_data



def generate_recruitment_report_excel(data):
    """
    Generates a single-sheet Excel file from a recruitment report data structure.

    The sheet contains a summary table followed by individual tables for detailed
    candidate data, with metadata for each table.

    Args:
        data (list): A list of dictionaries containing recruitment report data.

    Returns:
        HttpResponse: A Django HTTP response containing the Excel file.
    """
    # Create an in-memory byte stream to store the Excel file
    output = io.BytesIO()
    workbook = Workbook()
    main_sheet = workbook.active
    main_sheet.title = "Recruitment Report"

    # Define common styles
    header_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(bold=True, size=12)
    border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))

    current_row = 1

    # --- Create the Summary Table ---
    main_sheet.cell(row=current_row, column=1, value="Report Summary").font = title_font
    current_row += 2

    summary_headers = ['Company Name', 'JD Summary', 'JD ID', 'Profile Count', 'Feedback']
    main_sheet.append(summary_headers)
    for cell in main_sheet[current_row]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    current_row += 1

    for team_report in data:
        for summary_row in team_report.get('report_summary', []):
            row_data = [
                summary_row.get('company_name', ''),
                summary_row.get('jd_summary', ''),
                summary_row.get('jd_id', ''),
                summary_row.get('profile_count', ''),
                summary_row.get('feedback', '')
            ]
            main_sheet.append(row_data)
            for cell in main_sheet[current_row]:
                cell.border = border
            current_row += 1

    # Add a blank row for separation
    current_row += 2
    
    # --- Create Detailed Candidate Tables ---
    main_sheet.cell(row=current_row, column=1, value="Detailed Candidates").font = title_font
    current_row += 2

    for team_report in data:
        team_name = team_report.get('team_name', 'Unnamed Team')
        main_sheet.cell(row=current_row, column=1, value=f"Team: {team_name}").font = subtitle_font
        main_sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=11)
        current_row += 2
        
        for detailed_report in team_report.get('detailed_candidates', []):
            metadata = detailed_report.get('metadata', {})
            company_name = metadata.get('company_name', '')
            jd_summary = metadata.get('jd_summary', '')
            date_range = metadata.get('date_or_date_range', '')
            
            # Metadata row with larger font
            main_sheet.cell(row=current_row, column=1, value=f"JD: {jd_summary} for {company_name} ({date_range})").font = subtitle_font
            main_sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=11)
            current_row += 1
            
            # Detailed candidate headers
            detailed_headers = [
                'Candidate ID', 'Name', 'Email', 'Phone', 'Experience',
                'Current CTC', 'Expected CTC', 'Notice Period', 'Profile',
                'Location', 'Recruiter Comments'
            ]
            main_sheet.append(detailed_headers)
            for cell in main_sheet[current_row]:
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            current_row += 1

            for candidate in detailed_report.get('candidates', []):
                candidate_data = [
                    candidate.get('candidate_id', ''),
                    candidate.get('name', ''),
                    candidate.get('email', ''),
                    candidate.get('phone', ''),
                    candidate.get('experience', ''),
                    candidate.get('current_ctc', ''),
                    candidate.get('expected_ctc', ''),
                    candidate.get('notice_period', ''),
                    candidate.get('profile', ''),
                    candidate.get('location', ''),
                    candidate.get('recruiter_comments', '')
                ]
                main_sheet.append(candidate_data)
                for cell in main_sheet[current_row]:
                    cell.border = border
                current_row += 1
            
            # Add a blank row for separation
            current_row += 1

    # Autofit columns
    for column in main_sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        if adjusted_width > 50:
            adjusted_width = 50
        main_sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    # Save the workbook to the in-memory stream
    workbook.save(output)
    output.seek(0)

    # Prepare the Django HttpResponse
    
    return output


@login_required
def generate_export_report(request):
    """
    API endpoint to generate and export the recruitment report as an Excel file.
    """
    if request.method != "GET":
        return JsonResponse({"error": "Invalid request method."}, status=405)

    user_id = request.session.get("user_id")
    report_type = request.GET.get("report_type", "daily")
    from_date = None
    to_date = None
    if report_type == "custom":
        from_date = request.GET.get("from_date")
        to_date = request.GET.get("to_date")
        if not from_date or not to_date:
            return JsonResponse({"error": "From date and To date are required for custom report."}, status=400)
    print(f"Generating export report for user ID: {user_id} with report type: {report_type}")

    report_data = generate_daily_report(report_type, user_id, from_date=from_date, to_date=to_date)
    if report_data is None:
        return JsonResponse({"error": "Could not generate report data."}, status=500)

    # Generate the Excel file in memory
    excel_file = generate_recruitment_report_excel(report_data)

    # Prepare the Django HttpResponse to serve the file for download
    response = HttpResponse(
        excel_file.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=recruitment_report.xlsx'
    
    return response

@login_required
def report_and_logout(request):

    if request.method != "POST":
        return JsonResponse({"error": "Invalid request method."}, status=405)
    print("Report and logout initiated for user:", request.session.get('username'))
    post_data = request.POST
    mail_list = post_data.getlist("emails[]")
    user_id = request.session.get("user_id")

    if mail_list:
        email_list = [email.strip() for email in mail_list if email.strip()]
    else:
        messages.error(request, "Please provide at least one email address.")
        return redirect("prelogout")
    
    report_type = "daily"
    report_data = generate_daily_report(report_type, user_id)

    
    excel_file = generate_recruitment_report_excel(report_data)
    email_configs = DataOperations.get_email_configs(user_id)
    if not email_configs:
        messages.error(request, "Email configuration not found. Please set up your email configuration.")
        return redirect("prelogout")

    from .utils import decrypt_password

    from_email = email_configs['email']
    app_password = decrypt_password(email_configs['email_host_password'])
    smtp_host = email_configs.get('email_smtp_host', 'smtp.gmail.com')
    smtp_port = email_configs.get('email_smtp_port', 587)
    subject = "Recruitment Report"
    html_body = "Hi,<br><br>Please find the attached Daily recruitment report for your review.<br><br>Thanks."
    email_sent = MessageProviders.send_email(
        from_email=from_email,
        app_password=app_password,
        to_email=email_list,
        subject=subject,
        html_body=html_body,
        in_memory_attachments=[("recruitment_report.xlsx", excel_file)],
        smtp_host=smtp_host,
        smtp_port=smtp_port
    )

    if email_sent:
        return redirect("logout")
    else:
        messages.error(request, "Failed to send recruitment report.")
    return redirect("prelogout")

@login_required
def prelogout_page(request):
    """
    View to render the pre-logout page.
    """
    # get email ids of all admins and self team lead if exists
    user_id = request.session.get("user_id")
    admin_and_teamlead_emails = []
    try:
        conn = DataOperations.get_db_connection()
        cursor = conn.cursor(dictionary=True)
        emp_id = DataOperations.get_emp_id_from_user_id(user_id)
        cursor.execute("""
            SELECT
                u.email
            FROM
                users u
            WHERE
                u.role = 'Admin'

            UNION

            SELECT
                u_lead.email
            FROM
                hr_team_members htm_lead
            INNER JOIN
                teams t ON htm_lead.emp_id = t.lead_emp_id
            INNER JOIN
                team_members tm_current ON t.team_id = tm_current.team_id
            INNER JOIN
                users u_lead ON htm_lead.email = u_lead.email
            WHERE
                tm_current.emp_id = %s;
        """, (emp_id,))
        
        emails = [row['email'] for row in cursor.fetchall()]
        admin_and_teamlead_emails.extend(emails)
    except Exception as e:
        print("Error fetching admin emails:", e)
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

    return render(request, "prelogout.html", {"email_list": admin_and_teamlead_emails})